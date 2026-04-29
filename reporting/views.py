import csv
import io
import json
import calendar
import ssl
import xmlrpc.client
from collections import defaultdict
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer)
    from reportlab.pdfgen import canvas as rl_canvas
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'reportlab'])
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer)
    from reportlab.pdfgen import canvas as rl_canvas

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings

# Catégorie Odoo pour le carburant (BIEN MATERIEL / ENERGIE / CARBURANT)
CARBURANT_CATEG_ID = 262

# Distinction métier confirmée via equipment_id.category_id (Odoo).
# Transport: camions, bennes, véhicules routiers.
CATEGORIES_TRANSPORT = [18, 19, 21, 23, 43, 48, 49, 50]
# Production: pelles, bulldozers, foreuses, compresseurs, engins carrière.
CATEGORIES_PRODUCTION = [25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 39, 40, 41, 42, 44, 45, 53]

SITES_LIST = [
    'AIN JEMAA', 'CIMAR AIT BAHA', 'CIMAR AMSKROUD', 'CIMAT BENI MELLAL',
    'GRABEMARO BENSLIMANE', 'LH BENSLIMANE', 'LH MEKNES', 'LH OUJDA',
    'SETTAT DÉPÔT', 'Virtual Locations/Consommation', 'YOUSSOUFIA',
    'LHOUJ/Stock', 'LHMEK/Stock',
]


def format_number(value):
    """Formate un nombre en français: milliers espace insécable, décimales virgule."""
    if value is None:
        return '0,00'
    try:
        value = float(value)
        formatted = f'{value:,.2f}'
        integer, decimal = formatted.split('.')
        integer = integer.replace(',', '\u00a0')
        return f'{integer},{decimal}'
    except (TypeError, ValueError):
        return str(value)


def format_number_decimals(value, decimals=2):
    """Formate un nombre avec décimales fixes et séparateurs français."""
    if value is None:
        if decimals <= 0:
            return '0'
        return '0,' + ('0' * decimals)
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)
    fmt = f'{{:,.{decimals}f}}'
    formatted = fmt.format(v)
    if decimals <= 0:
        return formatted.replace(',', '\u00a0')
    integer, decimal = formatted.split('.')
    integer = integer.replace(',', '\u00a0')
    return f'{integer},{decimal}'


def extract_matricule(name):
    if not name:
        return ''
    return str(name).split('/', 1)[0].strip()


def _activity_bucket_from_picking(picking, ouvrage_text=''):
    """
    Détection métier robuste:
    - transport
    - voiture_service
    - production
    """
    if picking.get('service_car'):
        return 'voiture_service'
    if picking.get('transport_logistics'):
        return 'transport'

    parts = [ouvrage_text]
    for field in ('account_analytic_id', 'affectation_id', 'equipment_id', 'location_id'):
        val = picking.get(field)
        if isinstance(val, list) and len(val) > 1 and val[1]:
            parts.append(val[1])
    raw = ' '.join(str(v) for v in parts if v).lower()
    if ('transport' in raw) or ('logist' in raw):
        return 'transport'
    return 'production'


def _build_project_activity_map(uid, models, invoices):
    """
    Construit une map {project_id: 'transport'|'production'} depuis project.project.
    Utilise le booléen transport_logistics quand disponible.
    """
    project_ids = []
    for inv in invoices:
        p = inv.get('project_id')
        if isinstance(p, list) and p:
            project_ids.append(p[0])
    project_ids = sorted({pid for pid in project_ids if pid})
    if not project_ids:
        return {}

    try:
        projects = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'project.project', 'search_read',
            [[('id', 'in', project_ids)]],
            {'fields': ['id', 'name', 'transport_logistics'], 'limit': len(project_ids)},
        )
        out = {}
        for p in projects:
            name = (p.get('name') or '').lower()
            if p.get('transport_logistics') or ('transport' in name) or ('logist' in name):
                out[p['id']] = 'transport'
            else:
                out[p['id']] = 'production'
        return out
    except Exception:
        # Fallback si champ custom absent/inaccessible.
        return {}


def _invoice_activity_bucket(invoice, project_activity_map=None):
    """
    Classe une facture vente dans:
    - transport
    - production
    selon project_id + libellés de référence.
    """
    parts = []
    proj = invoice.get('project_id')
    if project_activity_map and isinstance(proj, list) and proj:
        mapped = project_activity_map.get(proj[0])
        if mapped in {'transport', 'production'}:
            return mapped
    if isinstance(proj, list) and len(proj) > 1 and proj[1]:
        parts.append(proj[1])
    for field in ('invoice_origin', 'ref', 'name'):
        val = invoice.get(field)
        if val:
            parts.append(str(val))
    raw = ' '.join(parts).lower()
    if ('transport' in raw) or ('logist' in raw):
        return 'transport'
    return 'production'


def _enrich_sortie_bon(bon):
    """Ajoute les champs *_fmt pour affichage HTML (après calcul des valeurs brutes)."""
    bon['cpt_initial_fmt'] = format_number_decimals(bon.get('cpt_initial'), 0)
    bon['cpt_actuel_fmt'] = format_number_decimals(bon.get('cpt_actuel'), 0)
    bon['ecart_fmt'] = format_number_decimals(bon.get('ecart'), 1)
    bon['product_qty_fmt'] = format_number_decimals(bon.get('product_qty'), 1)
    c = bon.get('consommation')
    bon['consommation_fmt'] = format_number_decimals(c, 2) if c else ''


def _enrich_entree_bon(bon):
    bon['product_qty_fmt'] = format_number_decimals(bon.get('product_qty'), 1)
    pu = bon.get('price_unit')
    bon['price_unit_fmt'] = format_number_decimals(pu, 2) if pu not in (None, '', False) else ''
    tot = bon.get('total')
    bon['total_fmt'] = format_number_decimals(tot, 2) if tot not in (None, '', False) else ''


def get_odoo_connection():
    """Retourne (uid, models) pour les appels Odoo XML-RPC."""
    verify_ssl = getattr(settings, 'ODOO_SSL_VERIFY', True)
    server_proxy_kwargs = {}
    if settings.ODOO_URL.startswith('https://') and not verify_ssl:
        # Local/dev workaround when corporate/intermediate cert is missing.
        server_proxy_kwargs['context'] = ssl._create_unverified_context()

    common = xmlrpc.client.ServerProxy(
        f'{settings.ODOO_URL}/xmlrpc/2/common',
        **server_proxy_kwargs,
    )
    uid = common.authenticate(settings.ODOO_DB, settings.ODOO_USER, settings.ODOO_PASS, {})
    models = xmlrpc.client.ServerProxy(
        f'{settings.ODOO_URL}/xmlrpc/2/object',
        **server_proxy_kwargs,
    )
    return uid, models



@login_required
def accueil(request):
    today = date.today()
    JOURS_FR = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    MOIS_FR  = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    date_affichee = f"{JOURS_FR[today.weekday()]} {today.day} {MOIS_FR[today.month]} {today.year}"

    return render(request, 'accueil.html', {
        'date_affichee': date_affichee,
    })


# ─────────────────────────────────────────────
#  GASOIL — SORTIES
#  Modèle : stock.picking (picking_type_consumption=True)
#  Produit filtré par catégorie CARBURANT (id=262)
# ─────────────────────────────────────────────
def _build_sorties_domain(date_debut, date_fin, site, chauffeur, ouvrage, anomalie,
                          societe='', categorie_engin='', activite_filtre=''):
    """Construit le domaine Odoo pour les bons de sortie gasoil."""
    domain = [
        ('state', '=', 'done'),
        ('picking_type_consumption', '=', True),
        ('move_ids.product_id.categ_id', '=', CARBURANT_CATEG_ID),
    ]
    if date_debut:
        domain.append(('scheduled_date', '>=', date_debut + ' 00:00:00'))
    if date_fin:
        domain.append(('scheduled_date', '<=', date_fin + ' 23:59:59'))
    if societe:
        domain.append(('company_id.name', '=', societe))
    if site:
        domain.append(('location_id.complete_name', 'ilike', site))
    if chauffeur:
        domain.append(('partner_id.name', 'ilike', chauffeur))
    
    if categorie_engin:
        domain.append(('equipment_id.category_id.name', '=', categorie_engin))
    # Filtres activité via champs booléens natifs Odoo
    if activite_filtre == 'transport':
        domain.append(('transport_logistics', '=', True))
    elif activite_filtre == 'voiture_service':
        domain.append(('service_car', '=', True))
    elif activite_filtre == 'production':
        domain += [('transport_logistics', '=', False), ('service_car', '=', False)]
    return domain


_TRANSPORT_CATS = {'CAMION TRACTEUR', 'CAMION ENGIN', 'SEMI-REMORQUE', 'PICK-UP', 'TRANSPORT PERSONNEL'}
_SERVICE_CATS   = {'VOITURE DE SERVICE', 'VOITURE DE FONCTION'}


def _fetch_sorties_bons(uid, models, domain, limit=1000):
    """Récupère et enrichit les bons de sortie depuis Odoo."""
    pickings = models.execute_kw(
        settings.ODOO_DB, uid, settings.ODOO_PASS,
        'stock.picking', 'search_read',
        [domain],
        {
            'fields': [
                'name', 'scheduled_date', 'date', 'write_date',
                'partner_id', 'user_id', 'company_id', 'location_id',
                'picking_type_id', 'account_analytic_id', 'affectation_id',
                'equipment_id', 'initial_counter', 'actual_counter',
                'move_ids',
                'transport_logistics', 'service_car',   # champs booléens natifs Odoo
            ],
            'order': 'scheduled_date desc',
            'limit': limit,
        }
    )

    # Quantités gasoil par picking
    moves_qty = {}
    if pickings:
        all_move_ids = [mid for p in pickings for mid in p.get('move_ids', [])]
        if all_move_ids:
            moves = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'stock.move', 'search_read',
                [[['id', 'in', all_move_ids],
                  ['product_id.categ_id', '=', CARBURANT_CATEG_ID]]],
                {'fields': ['id', 'picking_id', 'product_qty'], 'limit': 10000}
            )
            for m in moves:
                pid = m['picking_id'][0] if m['picking_id'] else None
                if pid:
                    moves_qty[pid] = moves_qty.get(pid, 0) + (m['product_qty'] or 0)

    bons = []
    for p in pickings:
        raw_date = (p.get('scheduled_date') or p.get('date') or p.get('write_date') or '')
        bon_date = raw_date[:10] if raw_date else '—'
        ecart    = (p.get('actual_counter') or 0) - (p.get('initial_counter') or 0)
        qty      = moves_qty.get(p['id'], 0)
        conso    = round(qty / ecart, 2) if ecart > 0 else 0
        is_anomalie = (ecart < 0) or (conso > 500 and conso > 0)

        engin_val = p.get('equipment_id')

        ouvrage_val = (
            p['account_analytic_id'][1] if isinstance(p.get('account_analytic_id'), list)
            else (p['affectation_id'][1] if p.get('affectation_id') else '')
            or extract_matricule(engin_val[1]) if engin_val
            else '—'
        )

        activity_bucket = _activity_bucket_from_picking(p, ouvrage_val)

        bons.append({
            'id':             p['id'],
            'date':           bon_date,
            'name':           p.get('name', '—'),
            'societe':        p['company_id'][1]           if p.get('company_id')           else '—',
            'chauffeur': (
                p['partner_id'][1] if p.get('partner_id')
                else p['user_id'][1] if p.get('user_id')
                else '—'
            ),
            'site':           p['location_id'][1]          if p.get('location_id')          else '—',
            'type_operation': p['picking_type_id'][1]      if p.get('picking_type_id')      else '—',
            'ouvrage':        ouvrage_val,
            'affectation':    p['affectation_id'][1]       if p.get('affectation_id')       else '—',
            'engin':          extract_matricule(engin_val[1]) if engin_val else '—',
            'categorie': (
                'Transport & Log.'  if activity_bucket == 'transport' else
                'Voiture de serv.'  if activity_bucket == 'voiture_service' else
                'Production'
            ),
            'is_transport':   activity_bucket == 'transport',
            'service_car':    p.get('service_car', False),
            'activite_bucket': activity_bucket,
            'cpt_initial':    p.get('initial_counter') or 0,
            'cpt_actuel':     p.get('actual_counter')  or 0,
            'ecart':          round(ecart, 1),
            'product_qty':    round(qty, 1),
            'consommation':   conso,
            'anomalie':       'Anomalie' if is_anomalie else 'OK',
        })
        _enrich_sortie_bon(bons[-1])
    return bons


def _sorties_pdf_response(bons, filters, total_litres, nb_anomalies, conso_moyenne):
    """Génère un PDF ReportLab A4 paysage — format SOMATRIN officiel."""
    import os
    from datetime import date as _date
    from reportlab.platypus import Image

    NAVY     = colors.HexColor('#1a2c4e')
    ORANGE   = colors.HexColor('#E87722')
    WHITE    = colors.white
    RED      = colors.HexColor('#dc2626')
    GREEN    = colors.HexColor('#16a34a')
    ROW_ALT  = colors.HexColor('#f4f6fb')
    ROW_ANOM = colors.HexColor('#fee2e2')
    GREY_TXT = colors.HexColor('#6b7280')
    BODY_TXT = colors.HexColor('#374151')
    today    = _date.today().strftime('%d/%m/%Y')

    # Chemin logo
    BASE_DIR  = settings.BASE_DIR
    LOGO_PATH = os.path.join(BASE_DIR, 'static', 'images', 'logo_somatrin.png')

    # ── Canvas numéroté ───────────────────────────────────────────────────────
    class _NumberedCanvas(rl_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            rl_canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved = []

        def showPage(self):
            self._saved.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved)
            for state in self._saved:
                self.__dict__.update(state)
                self._draw_footer(total)
                rl_canvas.Canvas.showPage(self)
            rl_canvas.Canvas.save(self)

        def _draw_footer(self, total):
            pw = landscape(A4)[0]
            self.saveState()
            self.setFont('Helvetica', 7)
            self.setFillColor(GREY_TXT)
            self.drawString(18 * mm, 8 * mm, 'SOMATRIN — Document Confidentiel — Usage Interne')
            self.drawRightString(pw - 18 * mm, 8 * mm,
                                 f'Page {self._pageNumber} / {total}  |  {today}')
            self.restoreState()

    # ── Buffer & document ─────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm,  bottomMargin=20 * mm,
    )

    PAGE_W = landscape(A4)[0] - 30 * mm  # largeur utile

    # ── En-tête ───────────────────────────────────────────────────────────────
    s_conf  = ParagraphStyle('conf',  fontName='Helvetica', fontSize=8,
                              textColor=GREY_TXT, alignment=TA_CENTER)
    s_title = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=14,
                              textColor=NAVY, alignment=TA_CENTER)
    s_sub   = ParagraphStyle('sub',   fontName='Helvetica', fontSize=9,
                              textColor=GREY_TXT, alignment=TA_CENTER)
    s_date  = ParagraphStyle('date',  fontName='Helvetica', fontSize=8,
                              textColor=GREY_TXT, alignment=TA_RIGHT)

    # Logo
    if os.path.exists(LOGO_PATH):
        logo = Image(LOGO_PATH, width=28 * mm, height=11 * mm)
    else:
        logo = Paragraph('<b>SOMATRIN</b>',
                         ParagraphStyle('lg', fontName='Helvetica-Bold',
                                        fontSize=12, textColor=NAVY))

    # Filtres actifs pour sous-titre
    filtres = []
    if filters.get('date_debut'): filtres.append(f"Du {filters['date_debut']}")
    if filters.get('date_fin'):   filtres.append(f"au {filters['date_fin']}")
    if filters.get('site'):       filtres.append(f"Site : {filters['site']}")
    if filters.get('societe'):    filtres.append(f"Société : {filters['societe']}")
    sous_titre = '  |  '.join(filtres) if filtres else 'Toutes les données'

    col_w_hdr = [50 * mm, PAGE_W - 100 * mm, 50 * mm]

    hdr_data = [[
        logo,
        [Paragraph('Document Confidentiel — Usage Interne', s_conf),
         Spacer(1, 2 * mm),
         Paragraph('Rapport Sorties Gasoil', s_title),
         Spacer(1, 1 * mm),
         Paragraph(sous_titre, s_sub)],
        Paragraph(f'Page 1 / …<br/>{today}', s_date),
    ]]

    hdr_tbl = Table(hdr_data, colWidths=col_w_hdr)
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',        (0, 0), (0, 0),   'LEFT'),
        ('ALIGN',        (1, 0), (1, 0),   'CENTER'),
        ('ALIGN',        (2, 0), (2, 0),   'RIGHT'),
        ('LINEBELOW',    (0, 0), (-1, 0),  1.5, NAVY),
        ('TOPPADDING',   (0, 0), (-1, 0),  2),
        ('BOTTOMPADDING',(0, 0), (-1, 0),  6),
    ]))

    elems = [hdr_tbl, Spacer(1, 6 * mm)]

    # ── Tableau principal ─────────────────────────────────────────────────────
    COL_MM = [20, 25, 16, 20, 40, 32, 22, 30, 15, 15, 13, 15, 15, 13]
    col_w = [c * mm for c in COL_MM]

    HEADERS = ['Date', 'N° Bon', 'Société', 'Site', 'Ouvrage', 'Engin',
               'Catégorie', 'Chauffeur', 'Cpt. Init', 'Cpt. Act',
               'Écart', 'Qté (L)', 'Conso.', 'Statut']

    s_h  = ParagraphStyle('sh',  fontSize=8, textColor=WHITE,
                           fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_c  = ParagraphStyle('sc',  fontSize=7, textColor=BODY_TXT, fontName='Helvetica')
    s_cr = ParagraphStyle('scr', fontSize=7, textColor=BODY_TXT,
                           fontName='Helvetica', alignment=TA_RIGHT)
    s_cc = ParagraphStyle('scc', fontSize=7, textColor=BODY_TXT,
                           fontName='Helvetica', alignment=TA_CENTER)
    s_ok = ParagraphStyle('sok', fontSize=7, textColor=GREEN,
                           fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_an = ParagraphStyle('san', fontSize=7, textColor=RED,
                           fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_co = ParagraphStyle('sco', fontSize=7,
                           textColor=colors.HexColor('#0ea5e9'),
                           fontName='Helvetica-Bold', alignment=TA_RIGHT)

    def trunc(s, n):
        return (s[:n] + '…') if len(s) > n else s

    rows = [[Paragraph(h, s_h) for h in HEADERS]]
    for bon in bons:
        conso_s = format_number_decimals(bon['consommation'], 2) if bon.get('consommation') else '—'
        statut  = (Paragraph('OK', s_ok)
                   if bon['anomalie'] == 'OK'
                   else Paragraph('Anomalie', s_an))
        rows.append([
            Paragraph(bon['date'],                    s_cc),
            Paragraph(bon['name'],                    s_c),
            Paragraph(trunc(bon['societe'],    12),   s_c),
            Paragraph(trunc(bon['site'],       14),   s_c),
            Paragraph(trunc(bon['ouvrage'],    25),   s_c),
            Paragraph(trunc(bon['engin'],      20),   s_c),
            Paragraph(bon.get('categorie', '—'),      s_c),
            Paragraph(trunc(bon['chauffeur'],  15),   s_c),
            Paragraph(format_number_decimals(bon['cpt_initial'], 0),   s_cr),
            Paragraph(format_number_decimals(bon['cpt_actuel'], 0),    s_cr),
            Paragraph(format_number_decimals(bon['ecart'], 1),          s_cr),
            Paragraph(format_number_decimals(bon['product_qty'], 1),    s_cr),
            Paragraph(conso_s,                        s_co),
            statut,
        ])

    # Ligne TOTAL fond bleu
    total_qty = sum(b['product_qty'] for b in bons)
    s_tot  = ParagraphStyle('stot', fontSize=8, textColor=WHITE,
                             fontName='Helvetica-Bold')
    s_totq = ParagraphStyle('stotq', fontSize=8, textColor=WHITE,
                             fontName='Helvetica-Bold', alignment=TA_RIGHT)
    rows.append([
        Paragraph(f'TOTAL — {len(bons)} bon{"s" if len(bons) != 1 else ""}', s_tot),
        '', '', '', '', '', '', '', '', '', '',
        Paragraph(format_number_decimals(total_qty, 1), s_totq),
        '', '',
    ])

    n_rows = len(rows)
    style  = [
        ('BACKGROUND',    (0, 0),  (-1, 0),  NAVY),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),  WHITE),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0),  (-1, 0),  8),
        ('ALIGN',         (0, 0),  (-1, 0),  'CENTER'),
        ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0),  (-1, 0),  5),
        ('BOTTOMPADDING', (0, 0),  (-1, 0),  5),
        ('FONTSIZE',      (0, 1),  (-1, -2), 7),
        ('TOPPADDING',    (0, 1),  (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1),  (-1, -1), 3),
        ('LEFTPADDING',   (0, 0),  (-1, -1), 3),
        ('RIGHTPADDING',  (0, 0),  (-1, -1), 3),
        ('GRID',          (0, 0),  (-1, -2), 0.4, colors.HexColor('#d1d5db')),
        # Lignes alternées
        *[('BACKGROUND', (0, i), (-1, i), ROW_ALT)
          for i in range(2, n_rows - 1, 2)],
        # Ligne TOTAL
        ('BACKGROUND',    (0, -1), (-1, -1), NAVY),
        ('TEXTCOLOR',     (0, -1), (-1, -1), WHITE),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, -1), (-1, -1), 8),
        ('LINEABOVE',     (0, -1), (-1, -1), 1.5, NAVY),
        ('TOPPADDING',    (0, -1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 5),
        ('SPAN',          (0, -1), (10, -1)),
    ]

    # Anomalies en rouge
    for i, bon in enumerate(bons, start=1):
        if bon['anomalie'] == 'Anomalie':
            style.append(('BACKGROUND', (0, i), (-1, i), ROW_ANOM))

    main_tbl = Table(rows, colWidths=col_w, repeatRows=1)
    main_tbl.setStyle(TableStyle(style))
    elems.append(main_tbl)

    # ── Build PDF ─────────────────────────────────────────────────────────────
    doc.build(elems, canvasmaker=_NumberedCanvas)
    buffer.seek(0)

    fname_parts = ['sorties_gasoil']
    if filters.get('date_debut'): fname_parts.append(filters['date_debut'])
    if filters.get('date_fin'):   fname_parts.append(filters['date_fin'])
    filename = '_'.join(fname_parts) + '.pdf'

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def gasoil_sorties(request):
    date_debut      = request.GET.get('date_debut', '')
    date_fin        = request.GET.get('date_fin', '')
    societe         = request.GET.get('societe', '')
    site            = request.GET.get('site', '')
    chauffeur       = request.GET.get('chauffeur', '').strip()
    ouvrage         = request.GET.get('ouvrage', '').strip()
    anomalie        = request.GET.get('anomalie', '')
    categorie_engin = request.GET.get('categorie_engin', '')
    activite_filtre = request.GET.get('activite', '')
    export          = request.GET.get('export', '')
    page_number     = request.GET.get('page', 1)

    bons             = []
    categories_engin = []
    ouvrages_list    = []
    error            = None
    total_bons       = 0
    total_litres     = 0.0
    nb_anomalies     = 0
    conso_moyenne    = 0.0

    try:
        uid, models = get_odoo_connection()

        # Catégories engins
        try:
            cats = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'maintenance.equipment.category', 'search_read',
                [[]],
                {'fields': ['name'], 'order': 'name asc', 'limit': 100}
            )
            categories_engin = [c['name'] for c in cats]
        except Exception:
            pass

        # Ouvrages / analytiques distincts (liste statique issue de account_analytic_id)
        ouvrages_list = [
            'Alimentation Station De Recyclage Mobile Par Pelle - GRABEMARO',
            'Alimentation station de lavage - LAFARGEHOLCIM MAROC',
            'Chargement Camions Agrégats Asment Béton & Tiers - GRABEMARO',
            'Chargement camions blocks Asment ciment - GRABEMARO',
            'Chargement camions pour concasseur - GRABEMARO',
            'Chargement des camions clients - LAFARGEHOLCIM MAROC',
            'Chargement et transport de la matière 1ère - LAFARGEHOLCIM MAROC',
            'Déstockage produits finis - LAFARGEHOLCIM MAROC',
            'Déstockage stérile au pré criblage - LAFARGEHOLCIM MAROC',
            'Déstockage, Alimentation Station De Lavage Sable - GRABEMARO',
            'Forage Minage Chargement et transport - LAFARGEHOLCIM MAROC',
            'Foration, chargement et tir des mines - GRABEMARO',
            'Mise en décharge - LAFARGEHOLCIM MAROC',
            'Ripage calcaire chargement et transport Zone A  vers HAZ - LAFARGEHOLCIM MAROC',
            'S00002-Arrosage de la piste d\'accès à la carrière (eau) - CIMENTS DU MAROC',
            'S00002-Décapage - CIMENTS DU MAROC',
            'S00002-Entretien de la piste d\'accès à la carrière - CIMENTS DU MAROC',
            'S00002-Extraction, Chargement, Transport & alimentation concasseur - CIMENTS DU MAROC',
            'S00002-Reprise de chargement argile concassée - CIMENTS DU MAROC',
            'S00004-Décapage sans tir - CIMENTS DU MAROC',
            'S00004-Foration, tir, chargement, transport et alimentation concasseur - CIMENTS DU MAROC',
            'S00006-Chargement camion blocs - ASMENT DU CENTRE',
            'S00006-Chargement des granulats - ASMENT DU CENTRE',
            'S00006-Foration et Abattage - ASMENT DU CENTRE',
            'S00006-Transport vers le concasseur - ASMENT DU CENTRE',
            'S00010-Chargement transport argile vers concasseur - LAFARGEHOLCIM MAROC',
            'S00010-Chargement transport calcaire vers concasseur - LAFARGEHOLCIM MAROC',
            'S00010-Foration, Minage carrière Calcaire - LAFARGEHOLCIM MAROC',
            'S00010-Manutention M.P ( Sable, minerai de fer) - LAFARGEHOLCIM MAROC',
            'S00010-Manutention M.P (Pouzzolane, Gypse) - LAFARGEHOLCIM MAROC',
            'S00010-Manutention interne : CIMENT - LAFARGEHOLCIM MAROC',
            'S00010-Ripage carrière Argile - LAFARGEHOLCIM MAROC',
            'S00011-Compensation prix Gasoil LH MEKNES - LAFARGEHOLCIM MAROC',
            'S00011-Forage minage chargement transport - zone D vers HAZ - LAFARGEHOLCIM MAROC',
            'S00011-Forage minage chargement transport -Zone 4 par minage vers HAZ - LAFARGEHOLCIM MAROC',
            'S00011-Location chargeuse 4m3 - LAFARGEHOLCIM MAROC',
            'S00011-Location engins pour manutention interne - LAFARGEHOLCIM MAROC',
            'S00011-Manutention des MP vers concasseurs - Chargement transport Matières premières - LAFARGEHOLCIM MAROC',
            'S00011-Ripage calcaire chargement et transport Zone A  vers HAZ - LAFARGEHOLCIM MAROC',
            'S00011-Ripage calcaire chargement et transport Zone D vers HAZ - LAFARGEHOLCIM MAROC',
            'S00012-Décapage Par Explosif - GRABEMARO',
            'S00012-Fourniture Station D\'eau - GRABEMARO',
            'S00012/ Décapage par explosif - GRABEMARO',
            'S00013-LOCATION CAMION (HRS) - GRABEMARO',
            'S00013-LOCATION CHARGEUSE (HRS) - GRABEMARO',
            'S00013-LOCATION CITERNE D\'EAU - GRABEMARO',
            'S00013-LOCATION D\'ENGINS : PELLE - GRABEMARO',
            'S00044-Manutention des travaux - LAFARGEHOLCIM MAROC',
            'S00071-FOURNITURE DE SCHISTE LH SETTAT - LAFARGEHOLCIM MAROC',
            'S00073-ALIMENTATION CAMION 8*4 - SERVICE LOGISTIQUES -SOMATRIN',
            'S00073-CHARGEMENT PRODUIT FINI - SERVICE LOGISTIQUES -SOMATRIN',
            'S00073-Concassage - SERVICE LOGISTIQUES -SOMATRIN',
            'S00073-DECAPAGE - SERVICE LOGISTIQUES -SOMATRIN',
            'S00073-FORAGE - SERVICE LOGISTIQUES -SOMATRIN',
            'S00073-REDUCTION - SERVICE LOGISTIQUES -SOMATRIN',
            'Transport inter du front au concasseur - GRABEMARO',
        ]

        domain = _build_sorties_domain(
            date_debut, date_fin, site, chauffeur, ouvrage, anomalie,
            societe, categorie_engin, activite_filtre
        )
        limit = 2000 if (date_debut or date_fin) else 500
        bons  = _fetch_sorties_bons(uid, models, domain, limit=limit)

    except Exception as e:
        error = f"Erreur de connexion Odoo : {e}"

    # Filtre ouvrage post-fetch
    print("OUVRAGE FILTRE:", ouvrage)
    if bons:
        print("EXEMPLE OUVRAGE BON:", bons[0].get('ouvrage'))
    # Filtre Statut post-fetch : ok → bons sans anomalie ; anomalie → bons avec anomalie calculée
    if anomalie == 'ok':
        bons = [b for b in bons if b['anomalie'] == 'OK']
    elif anomalie == 'anomalie':
        bons = [b for b in bons if b['anomalie'] == 'Anomalie']

    # Calculs des totaux
    total_bons    = len(bons)
    total_litres  = sum(b['product_qty'] for b in bons)
    nb_anomalies  = sum(1 for b in bons if b['anomalie'] == 'Anomalie')
    conso_vals    = [b['consommation'] for b in bons if b['consommation'] > 0]
    conso_moyenne = round(sum(conso_vals) / len(conso_vals), 2) if conso_vals else 0

    # ── Filtre sélection (export personnalisé par IDs) ───────────────────────
    def _apply_ids_filter(lst):
        ids_param = request.GET.get('ids', '')
        if not ids_param:
            return lst
        id_set = {int(i) for i in ids_param.split(',') if i.strip().isdigit()}
        return [b for b in lst if b['id'] in id_set]

    # ── Export CSV ────────────────────────────────────────────────────────────
    if export == 'csv':
        from datetime import date as _csv_date
        export_bons = _apply_ids_filter(bons)
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        fname_parts = ['sorties_gasoil']
        if date_debut: fname_parts.append(date_debut)
        if date_fin:   fname_parts.append(date_fin)
        response['Content-Disposition'] = (
            f'attachment; filename="{"_".join(fname_parts)}.csv"'
        )
        writer = csv.writer(response, delimiter=';')

        # ── Lignes de métadonnées ─────────────────────────────────────────────
        writer.writerow([f'# SOMATRIN — Rapport Sorties Gasoil'])
        username = request.user.get_full_name() or request.user.username
        writer.writerow([f'# Généré par : {username} le '
                         f'{_csv_date.today().strftime("%d/%m/%Y")}'])
        filtres_actifs = []
        if date_debut:      filtres_actifs.append(f'Date début : {date_debut}')
        if date_fin:        filtres_actifs.append(f'Date fin : {date_fin}')
        if societe:         filtres_actifs.append(f'Société : {societe}')
        if site:            filtres_actifs.append(f'Site : {site}')
        if chauffeur:       filtres_actifs.append(f'Chauffeur : {chauffeur}')
        if ouvrage:         filtres_actifs.append(f'Ouvrage : {ouvrage}')
        if anomalie:
            _lbl = {'ok': 'OK', 'anomalie': 'Anomalie'}.get(anomalie, anomalie)
            filtres_actifs.append(f'Statut : {_lbl}')
        if activite_filtre: filtres_actifs.append(f'Activité : {activite_filtre}')
        writer.writerow([f'# Filtres : '
                         + (', '.join(filtres_actifs) if filtres_actifs else 'Aucun')])
        writer.writerow([])   # Ligne 4 vide

        # ── En-têtes et données ───────────────────────────────────────────────
        writer.writerow(['Date', 'N° Bon', 'Société', 'Site', 'Ouvrage',
                         'Engin', 'Catégorie', 'Chauffeur',
                         'Cpt. initial', 'Cpt. actuel', 'Écart (km)',
                         'Qté (L)', 'Conso. (L/h)', 'Statut'])
        for b in export_bons:
            writer.writerow([
                b['date'], b['name'], b['societe'], b['site'],
                b['ouvrage'], b['engin'], b.get('categorie', ''),
                b['chauffeur'],
                str(b['cpt_initial']).replace('.', ','),
                str(b['cpt_actuel']).replace('.', ','),
                str(b['ecart']).replace('.', ','),
                str(b['product_qty']).replace('.', ','),
                str(b['consommation']).replace('.', ',') if b['consommation'] else '',
                b['anomalie'],
            ])
        return response

    # ── Export PDF ReportLab ──────────────────────────────────────────────────
    if export == 'pdf':
        return _sorties_pdf_response(
            bons=_apply_ids_filter(bons),
            filters={
                'date_debut': date_debut, 'date_fin': date_fin,
                'societe': societe, 'site': site,
                'chauffeur': chauffeur, 'anomalie': anomalie,
                'activite_filtre': activite_filtre,
            },
            total_litres=round(total_litres, 1),
            nb_anomalies=nb_anomalies,
            conso_moyenne=conso_moyenne,
        )

    paginator = Paginator(bons, 50)
    page_obj  = paginator.get_page(page_number)

    from urllib.parse import urlencode
    export_params = {k: v for k, v in request.GET.items() if k != 'page' and k != 'export'}
    export_qs = urlencode(export_params)

    return render(request, 'gasoil/sorties.html', {
        'page_obj': page_obj, 'error': error,
        'date_debut': date_debut, 'date_fin': date_fin,
        'societe': societe, 'site': site,
        'chauffeur': chauffeur, 'ouvrage': ouvrage,
        'anomalie': anomalie,
        'categorie_engin':  categorie_engin,
        'activite_filtre':  activite_filtre,
        'categories_engin': categories_engin,
        'ouvrages_list':    ouvrages_list,
        'sites':            SITES_LIST,
        'total_bons':       total_bons,
        'total_litres':     round(total_litres, 1),
        'nb_anomalies':     nb_anomalies,
        'conso_moyenne':    conso_moyenne,
        'export_qs':        export_qs,
        'total_bons_fmt': format_number_decimals(total_bons, 0),
        'total_litres_fmt': format_number_decimals(total_litres, 0),
        'nb_anomalies_fmt': format_number_decimals(nb_anomalies, 0),
        'conso_moyenne_fmt': format_number_decimals(conso_moyenne, 2),
        'page_start_fmt': format_number_decimals(page_obj.start_index, 0),
        'page_end_fmt': format_number_decimals(page_obj.end_index, 0),
    })


@login_required
def gasoil_sorties_export(request):
    """Export Excel des bons de sortie gasoil (openpyxl)."""
    date_debut      = request.GET.get('date_debut', '')
    date_fin        = request.GET.get('date_fin', '')
    societe         = request.GET.get('societe', '')
    site            = request.GET.get('site', '')
    chauffeur       = request.GET.get('chauffeur', '').strip()
    ouvrage         = request.GET.get('ouvrage', '').strip()
    anomalie        = request.GET.get('anomalie', '')
    activite_filtre = request.GET.get('activite', '')

    try:
        uid, models = get_odoo_connection()
        domain = _build_sorties_domain(date_debut, date_fin, site, chauffeur, ouvrage,
                                       anomalie, societe, activite_filtre=activite_filtre)
        bons   = _fetch_sorties_bons(uid, models, domain, limit=5000)
    except Exception as e:
        return HttpResponse(f"Erreur Odoo : {e}", status=500)

    # ── Post-filtres (anomalie calculée + sélection par IDs) ─────────────────
    if anomalie == 'ok':
        bons = [b for b in bons if b['anomalie'] == 'OK']
    elif anomalie == 'anomalie':
        bons = [b for b in bons if b['anomalie'] == 'Anomalie']

    ids_param = request.GET.get('ids', '')
    if ids_param:
        id_set = {int(i) for i in ids_param.split(',') if i.strip().isdigit()}
        bons   = [b for b in bons if b['id'] in id_set]

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sorties Gasoil"

    NAVY   = "1a2c4e"
    ORANGE = "E87722"
    LIGHT  = "F8F9FB"
    WHITE  = "FFFFFF"

    header_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    header_fill  = PatternFill("solid", fgColor=NAVY)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    total_font  = Font(name="Calibri", bold=True, color=NAVY, size=11)
    total_fill  = PatternFill("solid", fgColor="EEF1F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Titre ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = "Rapport Sorties Gasoil"
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="EEF1F7")
    ws.row_dimensions[1].height = 30

    # Sous-titre filtre
    ws.merge_cells("A2:N2")
    subtitle = []
    if date_debut: subtitle.append(f"Du {date_debut}")
    if date_fin:   subtitle.append(f"au {date_fin}")
    if societe:    subtitle.append(f"Société : {societe}")
    if site:       subtitle.append(f"Site : {site}")
    if anomalie:
        _lbl_a = {'ok': 'OK', 'anomalie': 'Anomalie'}.get(anomalie, anomalie)
        subtitle.append(f"Statut : {_lbl_a}")
    ws["A2"].value = "  |  ".join(subtitle) if subtitle else "Toutes les données"
    ws["A2"].font  = Font(name="Calibri", italic=True, size=10, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6  # Espace

    # ── En-têtes ──────────────────────────────────────────────────────────────
    headers = [
        ("Date",          12),
        ("N° Bon",        18),
        ("Société",       14),
        ("Site",          20),
        ("Ouvrage",       28),
        ("Engin",         20),
        ("Chauffeur",     22),
        ("Cpt. initial",  14),
        ("Cpt. actuel",   14),
        ("Écart (km)",    13),
        ("Qté (L)",       12),
        ("Conso.",        12),
        ("Statut",        13),
    ]

    for col_idx, (hdr, width) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[4].height = 24

    # ── Données ───────────────────────────────────────────────────────────────
    for row_idx, bon in enumerate(bons, start=5):
        is_even     = (row_idx % 2 == 0)
        row_fill    = PatternFill("solid", fgColor=LIGHT) if is_even else PatternFill("solid", fgColor=WHITE)
        anomal_fill = PatternFill("solid", fgColor="FEF2F2")

        use_fill = anomal_fill if bon['anomalie'] == 'Anomalie' else row_fill

        values = [
            bon['date'],
            bon['name'],
            bon['societe'],
            bon['site'],
            bon['ouvrage'],
            bon['engin'],
            bon['chauffeur'],
            bon['cpt_initial'],
            bon['cpt_actuel'],
            bon['ecart'],
            bon['product_qty'],
            bon['consommation'] if bon['consommation'] else '',
            bon['anomalie'],
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = use_fill
            cell.border = border
            cell.font   = Font(name="Calibri", size=10)
            # Alignements spécifiques
            if col_idx in (1,):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (9, 10, 11, 12, 13):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, (int, float)) and val:
                    cell.number_format = '0.0'
            elif col_idx == 14:
                cell.alignment = Alignment(horizontal="center")
                if val == 'Anomalie':
                    cell.font = Font(name="Calibri", size=10, bold=True, color="B91C1C")
                else:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="15803D")

        ws.row_dimensions[row_idx].height = 16

    # ── Ligne total ───────────────────────────────────────────────────────────
    total_row = len(bons) + 5
    ws.merge_cells(f"A{total_row}:G{total_row}")
    total_cell = ws.cell(row=total_row, column=1, value=f"TOTAL  —  {len(bons)} bon(s)")
    total_cell.font      = total_font
    total_cell.fill      = total_fill
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.border    = border

    total_litres = sum(b['product_qty'] for b in bons)
    qty_cell = ws.cell(row=total_row, column=11, value=round(total_litres, 1))
    qty_cell.font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    qty_cell.fill = total_fill
    qty_cell.alignment  = Alignment(horizontal="right")
    qty_cell.number_format = '0.0'
    qty_cell.border = border

    for col in [8, 9, 10, 12, 13]:
        c = ws.cell(row=total_row, column=col)
        c.fill   = total_fill
        c.border = border

    ws.freeze_panes = "A5"

    # ── Réponse HTTP ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename_parts = ["sorties_gasoil"]
    if date_debut: filename_parts.append(date_debut)
    if date_fin:   filename_parts.append(date_fin)
    filename = "_".join(filename_parts) + ".xlsx"

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _entrees_pdf_response(bons, filters, total_litres, total_cout):
    """Génère un PDF ReportLab A4 paysage — Entrées Gasoil."""
    import os
    from datetime import date as _date
    from reportlab.platypus import Image

    NAVY     = colors.HexColor('#1a2c4e')
    WHITE    = colors.white
    GREEN    = colors.HexColor('#16a34a')
    ROW_ALT  = colors.HexColor('#f4f6fb')
    GREY_TXT = colors.HexColor('#6b7280')
    BODY_TXT = colors.HexColor('#374151')
    today    = _date.today().strftime('%d/%m/%Y')

    BASE_DIR  = settings.BASE_DIR
    LOGO_PATH = os.path.join(BASE_DIR, 'static', 'images', 'logo_somatrin.png')

    # ── Canvas numéroté ───────────────────────────────────────────────────────
    class _NumberedCanvas(rl_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            rl_canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved = []

        def showPage(self):
            self._saved.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved)
            for state in self._saved:
                self.__dict__.update(state)
                self._draw_footer(total)
                rl_canvas.Canvas.showPage(self)
            rl_canvas.Canvas.save(self)

        def _draw_footer(self, total):
            pw = landscape(A4)[0]
            self.saveState()
            self.setFont('Helvetica', 7)
            self.setFillColor(GREY_TXT)
            self.drawString(15 * mm, 8 * mm, 'SOMATRIN — Document Confidentiel — Usage Interne')
            self.drawRightString(pw - 15 * mm, 8 * mm,
                                 f'Page {self._pageNumber} / {total}  |  {today}')
            self.restoreState()

    # ── Buffer & document ─────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm,  bottomMargin=20 * mm,
    )

    PAGE_W = landscape(A4)[0] - 30 * mm

    # ── En-tête ───────────────────────────────────────────────────────────────
    s_conf  = ParagraphStyle('conf',  fontName='Helvetica', fontSize=8,
                              textColor=GREY_TXT, alignment=TA_CENTER)
    s_title = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=14,
                              textColor=NAVY, alignment=TA_CENTER)
    s_sub   = ParagraphStyle('sub',   fontName='Helvetica', fontSize=9,
                              textColor=GREY_TXT, alignment=TA_CENTER)
    s_date  = ParagraphStyle('date',  fontName='Helvetica', fontSize=8,
                              textColor=GREY_TXT, alignment=TA_RIGHT)

    if os.path.exists(LOGO_PATH):
        logo = Image(LOGO_PATH, width=28 * mm, height=11 * mm)
    else:
        logo = Paragraph('<b>SOMATRIN</b>',
                         ParagraphStyle('lg', fontName='Helvetica-Bold',
                                        fontSize=12, textColor=NAVY))

    filtres = []
    if filters.get('date_debut'): filtres.append(f"Du {filters['date_debut']}")
    if filters.get('date_fin'):   filtres.append(f"au {filters['date_fin']}")
    if filters.get('fournisseur'): filtres.append(f"Fournisseur : {filters['fournisseur']}")
    sous_titre = '  |  '.join(filtres) if filtres else 'Toutes les données'

    col_w_hdr = [50 * mm, PAGE_W - 100 * mm, 50 * mm]

    hdr_data = [[
        logo,
        [Paragraph('Document Confidentiel — Usage Interne', s_conf),
         Spacer(1, 2 * mm),
         Paragraph('Rapport Entrées Gasoil', s_title),
         Spacer(1, 1 * mm),
         Paragraph(sous_titre, s_sub)],
        Paragraph(f'{today}', s_date),
    ]]

    hdr_tbl = Table(hdr_data, colWidths=col_w_hdr)
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',         (0, 0), (0, 0),   'LEFT'),
        ('ALIGN',         (1, 0), (1, 0),   'CENTER'),
        ('ALIGN',         (2, 0), (2, 0),   'RIGHT'),
        ('LINEBELOW',     (0, 0), (-1, 0),  1.5, NAVY),
        ('TOPPADDING',    (0, 0), (-1, 0),  2),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  6),
    ]))

    elems = [hdr_tbl, Spacer(1, 6 * mm)]

    # ── Tableau principal ─────────────────────────────────────────────────────
    COL_MM = [22, 35, 30, 40, 55, 20, 25, 30, 20]
    col_w  = [c * mm for c in COL_MM]

    HEADERS = ['Date', 'N° Facture', 'Réf. fourn.', 'Fournisseur',
               'Produit', 'Qté (L)', 'Prix unit. HT', 'Total HT (MAD)', 'Statut']

    s_h  = ParagraphStyle('sh',  fontSize=8, textColor=WHITE,
                           fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_c  = ParagraphStyle('sc',  fontSize=7, textColor=BODY_TXT, fontName='Helvetica')
    s_cr = ParagraphStyle('scr', fontSize=7, textColor=BODY_TXT,
                           fontName='Helvetica', alignment=TA_RIGHT)
    s_cc = ParagraphStyle('scc', fontSize=7, textColor=BODY_TXT,
                           fontName='Helvetica', alignment=TA_CENTER)
    s_ok = ParagraphStyle('sok', fontSize=7, textColor=GREEN,
                           fontName='Helvetica-Bold', alignment=TA_CENTER)

    def trunc(s, n):
        return (s[:n] + '…') if len(str(s)) > n else str(s)

    rows = [[Paragraph(h, s_h) for h in HEADERS]]
    for bon in bons:
        rows.append([
            Paragraph(str(bon['date']),                    s_cc),
            Paragraph(trunc(bon['name'], 30),              s_c),
            Paragraph(trunc(bon.get('ref', ''), 25),       s_c),
            Paragraph(trunc(bon['fournisseur'], 35),       s_c),
            Paragraph(trunc(bon['product'], 45),           s_c),
            Paragraph(format_number_decimals(bon['product_qty'], 1),         s_cr),
            Paragraph(format_number_decimals(bon['price_unit'], 2),         s_cr),
            Paragraph(format_number_decimals(bon['total'], 2),              s_cr),
            Paragraph(bon.get('statut', 'Validé'), s_ok),
        ])

    # Ligne TOTAL
    s_tot  = ParagraphStyle('stot', fontSize=8, textColor=WHITE,
                             fontName='Helvetica-Bold')
    s_totq = ParagraphStyle('stotq', fontSize=8, textColor=WHITE,
                             fontName='Helvetica-Bold', alignment=TA_RIGHT)

    rows.append([
        Paragraph(f'TOTAL — {len(bons)} ligne{"s" if len(bons) != 1 else ""}', s_tot),
        '', '', '', '',
        Paragraph(format_number_decimals(total_litres, 1), s_totq),
        '',
        Paragraph(format_number_decimals(total_cout, 2), s_totq),
        '',
    ])

    n_rows = len(rows)
    style  = [
        ('BACKGROUND',    (0, 0),  (-1, 0),  NAVY),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),  WHITE),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0),  (-1, 0),  8),
        ('ALIGN',         (0, 0),  (-1, 0),  'CENTER'),
        ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0),  (-1, 0),  5),
        ('BOTTOMPADDING', (0, 0),  (-1, 0),  5),
        ('FONTSIZE',      (0, 1),  (-1, -2), 7),
        ('TOPPADDING',    (0, 1),  (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1),  3),
        ('LEFTPADDING',   (0, 0),  (-1, -1), 3),
        ('RIGHTPADDING',  (0, 0),  (-1, -1), 3),
        ('GRID',          (0, 0),  (-1, -2), 0.4, colors.HexColor('#d1d5db')),
        *[('BACKGROUND', (0, i), (-1, i), ROW_ALT)
          for i in range(2, n_rows - 1, 2)],
        ('BACKGROUND',    (0, -1), (-1, -1), NAVY),
        ('TEXTCOLOR',     (0, -1), (-1, -1), WHITE),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, -1), (-1, -1), 8),
        ('LINEABOVE',     (0, -1), (-1, -1), 1.5, NAVY),
        ('TOPPADDING',    (0, -1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 5),
        ('SPAN',          (0, -1), (4, -1)),
    ]

    main_tbl = Table(rows, colWidths=col_w, repeatRows=1)
    main_tbl.setStyle(TableStyle(style))
    elems.append(main_tbl)

    # ── Build PDF ─────────────────────────────────────────────────────────────
    doc.build(elems, canvasmaker=_NumberedCanvas)
    buffer.seek(0)

    fname_parts = ['entrees_gasoil']
    if filters.get('date_debut'): fname_parts.append(filters['date_debut'])
    if filters.get('date_fin'):   fname_parts.append(filters['date_fin'])
    filename = '_'.join(fname_parts) + '.pdf'

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────
#  GASOIL — ENTRÉES
#  Modèle : stock.picking (réceptions de carburant)
#  Type : incoming (fournisseur → stock)
# ─────────────────────────────────────────────
@login_required
def gasoil_entrees(request):
    date_debut      = request.GET.get('date_debut', '')
    date_fin        = request.GET.get('date_fin', '')
    site            = request.GET.get('site', '')
    fournisseur     = request.GET.get('fournisseur', '').strip()
    activite_filtre = request.GET.get('activite', '')
    export          = request.GET.get('export', '')

    bons  = []
    error = None

    try:
        uid, models = get_odoo_connection()

        # ── 1. Lignes de factures fournisseurs contenant "GASOIL" ──
        line_domain = [
            ('move_id.move_type', '=', 'in_invoice'),
            ('move_id.state', '=', 'posted'),
            ('display_type', '=', 'product'),  # lignes produit uniquement (pas tax/payment_term)
            '|',
            ('product_id.name', 'ilike', 'gasoil'),
            ('name', 'ilike', 'gasoil'),
        ]
        if date_debut:
            line_domain.append(('move_id.invoice_date', '>=', date_debut))
        if date_fin:
            line_domain.append(('move_id.invoice_date', '<=', date_fin))
        if fournisseur:
            line_domain.append(('move_id.partner_id.name', 'ilike', fournisseur))

        lines = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.move.line', 'search_read',
            [line_domain],
            {
                'fields': ['move_id', 'product_id', 'name',
                           'quantity', 'price_unit', 'price_subtotal'],
                'order': 'move_id desc',
                'limit': 5000,
            }
        )

        # ── 2. En-têtes des factures (date, fournisseur, ref) ──
        move_ids = list({l['move_id'][0] for l in lines if l.get('move_id')})
        move_map = {}
        if move_ids:
            invoices = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'account.move', 'search_read',
                [[['id', 'in', move_ids]]],
                {'fields': ['id', 'name', 'invoice_date', 'partner_id', 'ref', 'state'],
                 'limit': len(move_ids)}
            )
            move_map = {m['id']: m for m in invoices}

        # ── 3. Construction des bons ──
        for line in lines:
            mid = line['move_id'][0] if line.get('move_id') else None
            mv  = move_map.get(mid, {})
            qty = line.get('quantity') or 0
            pu  = line.get('price_unit') or 0
            bons.append({
                'id':          line['id'],
                'date':        mv.get('invoice_date', '—') or '—',
                'name':        mv.get('name', '—') or '—',
                'ref':         mv.get('ref', '') or '',
                'fournisseur': mv['partner_id'][1] if mv.get('partner_id') else '—',
                'product':     (line['product_id'][1] if line.get('product_id')
                                else line.get('name', '—')),
                'description': line.get('name', ''),
                'product_qty': round(qty, 1),
                'price_unit':  round(pu, 2),
                'total':       round(line.get('price_subtotal') or qty * pu, 2),
                'statut': {
                    'draft': 'Brouillon',
                    'posted': 'Validé',
                    'cancel': 'Annulé'
                }.get(mv.get('state'), mv.get('state', '—') or '—'),
            })

    except Exception as e:
        error = f"Erreur de connexion Odoo : {e}"

    # ── Filtre activité post-fetch (les factures n'ont pas transport_logistics) ──
    # On filtre sur le nom du produit : GASOIL 10PPM = carburant pur (production/transport)
    # Citerne / Filtre à gasoil = pièces (traités comme production par défaut)
    if activite_filtre == 'transport':
        bons = [b for b in bons if 'PPM' in b.get('product', '').upper()
                                or 'CARBURANT' in b.get('product', '').upper()]
    elif activite_filtre == 'voiture_service':
        bons = [b for b in bons if 'SERVICE' in b.get('fournisseur', '').upper()
                                or 'VOITURE' in b.get('product', '').upper()]
    # 'production' : pas de filtre supplémentaire — tout ce qui n'est pas transport

    total_bons   = len(bons)
    total_litres = sum(b['product_qty'] for b in bons)
    total_cout   = sum(b['total'] for b in bons)

    # ── Export CSV ────────────────────────────────────────────────────────────
    if export == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        fname_parts = ['entrees_gasoil']
        if date_debut: fname_parts.append(date_debut)
        if date_fin:   fname_parts.append(date_fin)
        response['Content-Disposition'] = (
            f'attachment; filename="{"_".join(fname_parts)}.csv"'
        )
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Date facture', 'N° Facture', 'Réf. fournisseur',
                         'Fournisseur', 'Produit', 'Quantité (L)',
                         'Prix unit. HT', 'Total HT (MAD)', 'Statut'])
        for b in bons:
            writer.writerow([
                b['date'], b['name'], b.get('ref', ''), b['fournisseur'],
                b['product'],
                str(b['product_qty']).replace('.', ','),
                str(b['price_unit']).replace('.', ','),
                str(b['total']).replace('.', ','),
                b.get('statut', '—'),
            ])
        return response

    # ── Export PDF (page impression A4) ──────────────────────────────────────
    if export == 'pdf':
        return _entrees_pdf_response(
            bons=bons,
            filters={
                'date_debut': date_debut,
                'date_fin': date_fin,
                'fournisseur': fournisseur,
            },
            total_litres=round(total_litres, 1),
            total_cout=round(total_cout, 2),
        )

    return render(request, 'gasoil/entrees.html', {
        'bons':            bons,
        'error':           error,
        'date_debut':      date_debut,
        'date_fin':        date_fin,
        'site':            site,
        'fournisseur':     fournisseur,
        'activite_filtre': activite_filtre,
        'total_bons':      total_bons,
        'total_litres':    round(total_litres, 1),
        'total_cout':      round(total_cout, 2),
        'now':            date.today(),
    })


# ─────────────────────────────────────────────
#  GASOIL — BILAN
# ─────────────────────────────────────────────
@login_required
def gasoil_bilan(request):
    annee = request.GET.get('annee', '')
    mois = request.GET.get('mois', '')
    site  = request.GET.get('site', '')
    activite_filtre = request.GET.get('activite_filtre', '').strip()
    anomalie_seulement = request.GET.get('anomalie_seulement', '').strip()
    societe = request.GET.get('societe', '').strip()
    engin = request.GET.get('engin', '').strip()

    error        = None
    entrees_data = []
    sorties_data = []
    entrees_all = []
    sorties_all = []

    def _search_read_all(model_name, domain, fields, batch_size=2000):
        """Read all matching records using paginated search + read."""
        all_rows = []
        offset = 0
        while True:
            ids = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                model_name, 'search',
                [domain],
                {'offset': offset, 'limit': batch_size, 'order': 'id desc'}
            )
            if not ids:
                break
            rows = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                model_name, 'read',
                [ids],
                {'fields': fields}
            )
            all_rows.extend(rows)
            if len(ids) < batch_size:
                break
            offset += batch_size
        return all_rows

    def _read_moves_chunked(move_ids, fields, batch_size=2000):
        """Read stock.move records in chunks to avoid truncation/limits."""
        rows = []
        for i in range(0, len(move_ids), batch_size):
            chunk = move_ids[i:i + batch_size]
            part = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'stock.move', 'search_read',
                [[['id', 'in', chunk], ['product_id.categ_id', '=', CARBURANT_CATEG_ID]]],
                {'fields': fields, 'limit': batch_size}
            )
            rows.extend(part)
        return rows

    try:
        uid, models = get_odoo_connection()

        date_filter = []

        # ── Sorties ──
        domain_s = [
            ('picking_type_consumption', '=', True),
            ('state', '=', 'done'),
            ('move_ids.product_id.categ_id', '=', CARBURANT_CATEG_ID),
        ] + date_filter
        if site:
            domain_s.append(('location_id.complete_name', 'ilike', site))
        if societe:
            domain_s.append(('company_id.name', 'ilike', societe))
        if activite_filtre == 'transport':
            domain_s.append(('transport_logistics', '=', True))
        elif activite_filtre == 'voiture_service':
            domain_s.append(('service_car', '=', True))
        elif activite_filtre == 'production':
            domain_s += [('transport_logistics', '=', False), ('service_car', '=', False)]
        if anomalie_seulement == '1':
            domain_s.append(('picking_type_is_hors_affectation', '=', True))
        if engin:
            domain_s.append(('equipment_id.name', 'ilike', engin))

        pickings_s = _search_read_all(
            'stock.picking',
            domain_s,
            ['name', 'scheduled_date', 'date', 'write_date',
             'location_id', 'company_id', 'move_ids', 'equipment_id',
             'transport_logistics', 'service_car',
             'picking_type_is_hors_affectation'],
        )

        # Quantités sorties
        if pickings_s:
            all_ids = [mid for p in pickings_s for mid in p.get('move_ids', [])]
            qty_map = {}
            if all_ids:
                mvs = _read_moves_chunked(all_ids, ['picking_id', 'product_qty'])
                for m in mvs:
                    pid = m['picking_id'][0] if m['picking_id'] else None
                    if pid:
                        qty_map[pid] = qty_map.get(pid, 0) + (m['product_qty'] or 0)

            for p in pickings_s:
                raw = (p.get('scheduled_date') or p.get('date') or p.get('write_date') or '')
                # Catégorie via booléens natifs Odoo
                if p.get('transport_logistics'):
                    categorie = 'Transport & Logistique'
                elif p.get('service_car'):
                    categorie = 'Voiture de service'
                else:
                    categorie = 'Production'
                engin_val = p.get('equipment_id')
                sorties_data.append({
                    'date':     raw[:10] if raw else '—',
                    'site':     p['location_id'][1] if p.get('location_id') else '—',
                    'societe':  p['company_id'][1] if p.get('company_id') else '—',
                    'qty':      qty_map.get(p['id'], 0),
                    'anomalie': p.get('picking_type_is_hors_affectation', False),
                    'name':     p.get('name', '—'),
                    'engin':    engin_val[1] if engin_val else 'Inconnu',
                    'categorie': categorie,
                })

        # ── Entrées ──
        domain_e = [
            ('state', '=', 'done'),
            ('picking_type_id.code', '=', 'incoming'),
            ('move_ids.product_id.categ_id', '=', CARBURANT_CATEG_ID),
        ] + date_filter
        if site:
            domain_e.append(('location_dest_id.complete_name', 'ilike', site))
        if societe:
            domain_e.append(('company_id.name', 'ilike', societe))

        pickings_e = _search_read_all(
            'stock.picking',
            domain_e,
            ['scheduled_date', 'date', 'write_date',
             'location_dest_id', 'company_id', 'move_ids'],
        )

        if pickings_e:
            all_ids_e = [mid for p in pickings_e for mid in p.get('move_ids', [])]
            qty_map_e = {}
            pu_map_e  = {}
            if all_ids_e:
                mvs_e = _read_moves_chunked(all_ids_e, ['picking_id', 'product_qty', 'price_unit'])
                for m in mvs_e:
                    pid = m['picking_id'][0] if m['picking_id'] else None
                    if pid:
                        qty_map_e[pid] = qty_map_e.get(pid, 0) + (m['product_qty'] or 0)
                        pu_map_e[pid]  = m.get('price_unit') or 0

            for p in pickings_e:
                qty = qty_map_e.get(p['id'], 0)
                pu  = pu_map_e.get(p['id'], 0)
                raw = (p.get('scheduled_date') or p.get('date') or p.get('write_date') or '')
                entrees_data.append({
                    'date': raw[:10] if raw else '—',
                    'societe': p['company_id'][1] if p.get('company_id') else '—',
                    'qty':  qty,
                    'cout': qty * pu,
                })

        # Conserver une copie brute (avant filtres période) pour diagnostics/listes.
        sorties_all = list(sorties_data)
        entrees_all = list(entrees_data)

        # Filtrage de période robuste basé sur la date réellement disponible
        # (scheduled_date ou date ou write_date déjà normalisée en YYYY-MM-DD).
        if annee:
            sorties_data = [s for s in sorties_data if (s.get('date') or '').startswith(f'{annee}-')]
            entrees_data = [e for e in entrees_data if (e.get('date') or '').startswith(f'{annee}-')]
            if mois and len(mois) == 2 and mois.isdigit():
                needle = f'{annee}-{mois}'
                sorties_data = [s for s in sorties_data if (s.get('date') or '').startswith(needle)]
                entrees_data = [e for e in entrees_data if (e.get('date') or '').startswith(needle)]

    except Exception as e:
        error = f"Erreur de connexion Odoo : {e}"

    # ── KPI ──
    total_entrees = sum(e['qty'] for e in entrees_data)
    total_sorties = sum(s['qty'] for s in sorties_data)
    stock_estime  = total_entrees - total_sorties
    total_cout    = sum(e['cout'] for e in entrees_data)
    nb_anomalies  = sum(1 for s in sorties_data if s['anomalie'])

    # ── Graphique mensuel ──
    mois_labels       = ['Jan','Fév','Mar','Avr','Mai','Juin',
                         'Juil','Août','Sep','Oct','Nov','Déc']
    sorties_par_mois  = defaultdict(float)
    entrees_par_mois  = defaultdict(float)

    for s in sorties_data:
        if s['date'] and len(s['date']) >= 7:
            sorties_par_mois[int(s['date'][5:7])] += s['qty']
    for e in entrees_data:
        if e['date'] and len(e['date']) >= 7:
            entrees_par_mois[int(e['date'][5:7])] += e['qty']

    chart_labels       = json.dumps(mois_labels)
    chart_sorties_data = json.dumps([round(sorties_par_mois.get(m, 0), 1) for m in range(1, 13)])
    chart_entrees_data = json.dumps([round(entrees_par_mois.get(m, 0), 1) for m in range(1, 13)])

    # ── Répartition par site ──
    site_data = defaultdict(float)
    for s in sorties_data:
        site_data[s['site']] += s['qty']

    site_labels = json.dumps(list(site_data.keys()))
    site_values = json.dumps([round(v, 1) for v in site_data.values()])

    recap_sites = [
        {
            'site':   s,
            'litres': round(v, 1),
            'pct':    round(v / total_sorties * 100, 1) if total_sorties else 0,
        }
        for s, v in sorted(site_data.items(), key=lambda x: -x[1])
    ]

    # ── 1) Répartition par catégorie d'engin ─────────────────────────────────
    cat_data = defaultdict(float)
    for s in sorties_data:
        cat_data[s['categorie']] += s['qty']
    cat_sorted = sorted(cat_data.items(), key=lambda x: -x[1])[:8]
    categories_labels = json.dumps([c[0] for c in cat_sorted])
    categories_values = json.dumps([round(c[1], 1) for c in cat_sorted])

    # ── 2) Consommation par semaine ───────────────────────────────────────────
    semaines_data = defaultdict(float)
    for s in sorties_data:
        d = s['date']
        if d and len(d) == 10:
            try:
                from datetime import datetime
                dt = datetime.strptime(d, '%Y-%m-%d')
                # Clé = lundi de la semaine ISO (YYYY-Www)
                iso = dt.isocalendar()
                semaine_key = f"{iso[0]}-S{iso[1]:02d}"
                semaine_label = dt.strftime('%d/%m')
                semaines_data[semaine_key] = semaines_data.get(semaine_key, 0) + s['qty']
                semaines_data[f'_lbl_{semaine_key}'] = semaine_label
            except ValueError:
                pass
    # Trier par clé (chronologique) et séparer labels/valeurs
    real_keys = sorted(k for k in semaines_data if not k.startswith('_lbl_'))
    semaines_labels = json.dumps([semaines_data.get(f'_lbl_{k}', k) for k in real_keys])
    semaines_values = json.dumps([round(semaines_data[k], 1) for k in real_keys])

    # ── 3) Top 10 bons de sortie ──────────────────────────────────────────────
    bons_data = defaultdict(float)
    for s in sorties_data:
        bons_data[s['name']] += s['qty']
    bons_sorted = sorted(bons_data.items(), key=lambda x: -x[1])[:10]
    bons_labels = json.dumps([b[0] for b in bons_sorted])
    bons_values = json.dumps([round(b[1], 1) for b in bons_sorted])

    # ── 4) Top 10 matricules ──────────────────────────────────────────────────
    mat_data = defaultdict(float)
    for s in sorties_data:
        mat_data[s['engin']] += s['qty']
    mat_sorted = sorted(mat_data.items(), key=lambda x: -x[1])[:10]
    max_mat = mat_sorted[0][1] if mat_sorted else 1
    matricules_data = [
        {
            'nom':   m[0],
            'total': round(m[1], 1),
            'pct':   round(m[1] / max_mat * 100, 1),
        }
        for m in mat_sorted
    ]

    # ── 5) Équipements actifs ─────────────────────────────────────────────────
    equip_map = defaultdict(lambda: {'categorie': '—', 'nb_sorties': 0, 'total_litres': 0.0})
    for s in sorties_data:
        e = s['engin']
        equip_map[e]['categorie']   = s['categorie']
        equip_map[e]['nb_sorties'] += 1
        equip_map[e]['total_litres'] += s['qty']
    equipements_actifs = sorted(
        [{'matricule': k, **v, 'total_litres': round(v['total_litres'], 1)}
         for k, v in equip_map.items()],
        key=lambda x: -x['total_litres']
    )[:20]
    nb_equipements_actifs = len(equip_map)
    societes_list = sorted({s.get('societe') for s in sorties_all if s.get('societe') and s.get('societe') != '—'})
    engins_list = sorted({s.get('engin') for s in sorties_all if s.get('engin') and s.get('engin') != 'Inconnu'})[:200]
    years_detected = sorted({
        d[:4]
        for d in [*(s.get('date', '') for s in sorties_all), *(e.get('date', '') for e in entrees_all)]
        if isinstance(d, str) and len(d) >= 4 and d[:4].isdigit()
    })
    # N'afficher que les années réellement présentes dans les données Odoo.
    if not years_detected:
        years_detected = [str(date.today().year)]
    no_data_for_filters = (
        (annee or mois or site or societe or activite_filtre or engin or anomalie_seulement)
        and total_entrees == 0
        and total_sorties == 0
    )
    month_names = {
        '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
        '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
        '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre',
    }
    months_detected_for_year = set()
    if annee:
        for d in [*(s.get('date', '') for s in sorties_all), *(e.get('date', '') for e in entrees_all)]:
            if isinstance(d, str) and len(d) >= 7 and d[:4] == annee:
                mm = d[5:7]
                if mm in month_names:
                    months_detected_for_year.add(mm)
    else:
        for d in [*(s.get('date', '') for s in sorties_all), *(e.get('date', '') for e in entrees_all)]:
            if isinstance(d, str) and len(d) >= 7:
                mm = d[5:7]
                if mm in month_names:
                    months_detected_for_year.add(mm)
    mois_list_dynamic = [(m, month_names[m]) for m in sorted(months_detected_for_year)]
    if not mois_list_dynamic:
        mois_list_dynamic = [
            ('01', 'Janvier'), ('02', 'Février'), ('03', 'Mars'), ('04', 'Avril'),
            ('05', 'Mai'), ('06', 'Juin'), ('07', 'Juillet'), ('08', 'Août'),
            ('09', 'Septembre'), ('10', 'Octobre'), ('11', 'Novembre'), ('12', 'Décembre'),
        ]
    filtred_rows_count = len(sorties_data) + len(entrees_data)

    return render(request, 'gasoil/bilan.html', {
        'error': error,
        'annee': annee, 'mois': mois, 'site': site,
        'activite_filtre': activite_filtre,
        'anomalie_seulement': anomalie_seulement,
        'societe': societe,
        'engin': engin,
        'total_entrees': round(total_entrees, 1),
        'total_sorties': round(total_sorties, 1),
        'stock_estime':  round(stock_estime, 1),
        'total_cout':    round(total_cout, 2),
        'nb_anomalies':  nb_anomalies,
        'chart_labels':       chart_labels,
        'chart_sorties_data': chart_sorties_data,
        'chart_entrees_data': chart_entrees_data,
        'site_labels':        site_labels,
        'site_values':        site_values,
        'recap_sites':        recap_sites,
        'annees_list':        years_detected,
        'mois_list': mois_list_dynamic,
        'sites':              SITES_LIST,
        'societes_list': societes_list,
        'engins_list': engins_list,
        'no_data_for_filters': no_data_for_filters,
        'filtred_rows_count': filtred_rows_count,
        # Nouvelles sections analytiques
        'categories_labels_json': categories_labels,
        'categories_values_json': categories_values,
        'semaines_labels_json':   semaines_labels,
        'semaines_values_json':   semaines_values,
        'bons_labels_json':       bons_labels,
        'bons_values_json':       bons_values,
        'matricules_data':        matricules_data,
        'equipements_actifs':     equipements_actifs,
        'nb_equipements_actifs':  nb_equipements_actifs,
    })


# ─────────────────────────────────────────────
#  TRANSPORT & LOGISTIQUE
# ─────────────────────────────────────────────
def _transport_domain_dates(date_debut, date_fin, field_name):
    domain = []
    if date_debut:
        domain.append((field_name, '>=', f'{date_debut} 00:00:00'))
    if date_fin:
        domain.append((field_name, '<=', f'{date_fin} 23:59:59'))
    return domain


@login_required
def transport_bons(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    reference = request.GET.get('reference', '').strip()
    partenaire = request.GET.get('partenaire', '').strip()
    site = request.GET.get('site', '').strip()

    bons = []
    error = None
    try:
        uid, models = get_odoo_connection()
        domain = [('state', 'in', ['done', 'assigned'])]
        domain += _transport_domain_dates(date_debut, date_fin, 'scheduled_date')
        if reference:
            domain.append(('name', 'ilike', reference))
        if partenaire:
            domain.append(('partner_id.name', 'ilike', partenaire))
        if site:
            domain.append(('location_id.complete_name', 'ilike', site))

        try:
            domain_with_flag = domain + [('transport_logistics', '=', True)]
            records = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'stock.picking', 'search_read',
                [domain_with_flag],
                {
                    'fields': [
                        'name', 'scheduled_date', 'state', 'partner_id',
                        'location_id', 'location_dest_id', 'origin',
                    ],
                    'order': 'scheduled_date desc',
                    'limit': 500,
                }
            )
        except Exception:
            records = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'stock.picking', 'search_read',
                [domain],
                {
                    'fields': [
                        'name', 'scheduled_date', 'state', 'partner_id',
                        'location_id', 'location_dest_id', 'origin',
                    ],
                    'order': 'scheduled_date desc',
                    'limit': 500,
                }
            )

        for rec in records:
            origine = (
                rec.get('origin')
                or (rec['location_dest_id'][1] if rec.get('location_dest_id') else '')
                or '—'
            )
            bons.append({
                'date': (rec.get('scheduled_date') or '')[:10] or '—',
                'reference': rec.get('name') or '—',
                'partenaire': rec['partner_id'][1] if rec.get('partner_id') else '—',
                'site': rec['location_id'][1] if rec.get('location_id') else '—',
                'origine': origine,
                'etat': rec.get('state') or '—',
            })
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return render(request, 'transport/bons_transport.html', {
        'rows': bons,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'reference': reference,
        'partenaire': partenaire,
        'site': site,
        'total_rows': len(bons),
    })


@login_required
def transport_gasoil(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    vehicule = request.GET.get('vehicule', '').strip()
    conducteur_id = request.GET.get('conducteur_id', '').strip()
    group_by = request.GET.get('group_by', '').strip()

    rows = []
    vehicules = []
    conducteurs = []
    error = None
    try:
        uid, models = get_odoo_connection()

        # Listes déroulantes (distinctes) chargées au chargement de la page.
        list_domain = _build_sorties_domain(
            date_debut='',
            date_fin='',
            site='',
            chauffeur='',
            ouvrage='',
            anomalie='',
            societe='',
            categorie_engin='',
            activite_filtre='transport',
        )
        pickings_for_filters = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'stock.picking', 'search_read',
            [list_domain],
            {
                'fields': ['equipment_id', 'partner_id', 'user_id'],
                'limit': 5000,
            }
        )
        vehicule_set = set()
        conducteur_map = {}
        for p in pickings_for_filters:
            if p.get('equipment_id'):
                vehicule_set.add(extract_matricule(p['equipment_id'][1]))
            if p.get('partner_id'):
                conducteur_map[p['partner_id'][0]] = p['partner_id'][1]
            elif p.get('user_id'):
                conducteur_map[p['user_id'][0]] = p['user_id'][1]
        vehicules = [
            {
                'value': v,
                'label': extract_matricule(v),
            }
            for v in sorted(vehicule_set)
        ]
        conducteurs = [
            {'id': cid, 'name': cname}
            for cid, cname in sorted(conducteur_map.items(), key=lambda item: item[1].lower())
        ]

        # Reprise de la logique du module gasoil existant:
        # stock.picking + stock.move (catégorie carburant) avec activité transport.
        domain = _build_sorties_domain(
            date_debut=date_debut,
            date_fin=date_fin,
            site='',
            chauffeur='',
            ouvrage='',
            anomalie='',
            societe='',
            categorie_engin='',
            activite_filtre='transport',
        )
        if conducteur_id.isdigit():
            cid = int(conducteur_id)
            domain += ['|', ('partner_id', '=', cid), ('user_id', '=', cid)]

        limit = 2000 if (date_debut or date_fin) else 500
        bons = _fetch_sorties_bons(uid, models, domain, limit=limit)

        # Montant approximé depuis stock.move (qty * price_unit) par picking.
        move_domain = [
            ['picking_id', 'in', [b['id'] for b in bons]],
            ['product_id.categ_id', '=', CARBURANT_CATEG_ID],
        ]
        move_lines = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'stock.move', 'search_read',
            [move_domain],
            {
                'fields': ['picking_id', 'product_qty', 'price_total', 'unit_price', 'partner_id'],
                'limit': 10000,
            }
        )

        amount_by_picking = defaultdict(float)
        driver_by_picking = {}
        for mv in move_lines:
            pid = mv['picking_id'][0] if mv.get('picking_id') else None
            if not pid:
                continue
            price_total = mv.get('price_total')
            if price_total not in (None, False):
                amount_by_picking[pid] += price_total or 0
            else:
                qty = mv.get('product_qty') or 0
                pu = mv.get('unit_price') or 0
                amount_by_picking[pid] += qty * pu
            if not driver_by_picking.get(pid) and mv.get('partner_id'):
                driver_by_picking[pid] = mv['partner_id'][1]

        # Garantit que la page Transport n'affiche que le périmètre transport/logistique.
        bons = [b for b in bons if b.get('activite_bucket') == 'transport']

        if vehicule:
            bons = [b for b in bons if (b.get('engin') or '') == vehicule]

        for bon in bons:
            vehicle_name = bon.get('engin') or '—'
            vehicle_display = extract_matricule(vehicle_name)
            rows.append({
                'date': bon.get('date') or '—',
                'vehicule': vehicle_name,
                'vehicule_display': vehicle_display,
                'conducteur': bon.get('chauffeur') or driver_by_picking.get(bon['id']) or '—',
                'litres': round(bon.get('product_qty') or 0, 2),
                'montant': round(amount_by_picking.get(bon['id'], 0), 2),
                'compteur': bon.get('cpt_actuel') or 0,
            })
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    total_litres = round(sum(r['litres'] for r in rows), 2)
    total_montant = round(sum(r['montant'] for r in rows), 2)
    total_rows = len(rows)

    grouped_rows = []
    group_by_columns = []
    if group_by:
        groups = {}
        for row in rows:
            if group_by == 'mois':
                key = (row['date'][:7] if row.get('date') and row['date'] != '—' else '—',)
            elif group_by == 'vehicule':
                key = (row['vehicule_display'],)
            elif group_by == 'conducteur':
                key = (row['conducteur'],)
            elif group_by == 'mois_vehicule':
                key = (
                    row['date'][:7] if row.get('date') and row['date'] != '—' else '—',
                    row['vehicule_display'],
                )
            elif group_by == 'mois_conducteur':
                key = (
                    row['date'][:7] if row.get('date') and row['date'] != '—' else '—',
                    row['conducteur'],
                )
            else:
                key = ()

            if key not in groups:
                groups[key] = {'nb_lignes': 0, 'litres': 0.0, 'montant': 0.0}
            groups[key]['nb_lignes'] += 1
            groups[key]['litres'] += row['litres']
            groups[key]['montant'] += row['montant']

        grouped_rows = [
            {'keys': key, 'nb_lignes': val['nb_lignes'], 'litres': round(val['litres'], 2), 'montant': round(val['montant'], 2)}
            for key, val in sorted(groups.items(), key=lambda item: item[0])
        ]

        if group_by == 'mois':
            group_by_columns = ['Mois']
        elif group_by == 'vehicule':
            group_by_columns = ['Véhicule']
        elif group_by == 'conducteur':
            group_by_columns = ['Conducteur']
        elif group_by == 'mois_vehicule':
            group_by_columns = ['Mois', 'Véhicule']
        elif group_by == 'mois_conducteur':
            group_by_columns = ['Mois', 'Conducteur']

    return render(request, 'transport/gasoil.html', {
        'rows': rows,
        'grouped_rows': grouped_rows,
        'group_by': group_by,
        'group_by_columns': group_by_columns,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'vehicule': vehicule,
        'conducteur_id': conducteur_id,
        'vehicules': vehicules,
        'conducteurs': conducteurs,
        'total_rows': total_rows,
        'total_litres': total_litres,
        'total_montant': total_montant,
    })


def _domain_transport_analytic_line(
    date_debut, date_fin, vehicule,
    product_id=None, company_id=None,
):
    """Lignes analytiques transport : picking avec consommation OU transport_logistics."""
    domain = []
    if date_debut:
        domain.append(('date', '>=', date_debut))
    if date_fin:
        domain.append(('date', '<=', date_fin))
    domain.extend([
        '|',
        ('transfer_consumption_id', '!=', False),
        ('transfer_consumption_id.transport_logistics', '=', True),
    ])
    if vehicule:
        v = str(vehicule).strip()
        domain.extend([
            '|',
            ('transfer_consumption_id.equipment_id.name', 'ilike', v),
            ('transfer_consumption_id.affectation_id.name', 'ilike', v),
        ])
    if product_id:
        domain.append(('product_id', '=', int(product_id)))
    if company_id:
        domain.append(('company_id', '=', int(company_id)))
    return domain


@login_required
def transport_couts_nature(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    nature = request.GET.get('nature', '').strip()
    vehicule = request.GET.get('vehicule', '').strip()
    product_id = request.GET.get('product_id', '').strip()
    company_id = request.GET.get('company_id', '').strip()
    group_by = request.GET.get('group_by', '').strip()

    rows = []
    grouped_rows = []
    nature_options = []
    vehicules = []
    product_options = []
    company_options = []
    error = None
    total_montant = 0.0

    try:
        uid, models = get_odoo_connection()

        vehicules_raw = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'maintenance.equipment', 'search_read',
            [[]],
            {'fields': ['id', 'name'], 'order': 'name', 'limit': 500},
        )
        vehicules = sorted({
            extract_matricule(v.get('name'))
            for v in vehicules_raw
            if v.get('name')
        })

        base_filter_domain = _domain_transport_analytic_line(
            date_debut, date_fin,
            vehicule,
        )
        options_domain = _domain_transport_analytic_line(
            date_debut, date_fin,
            None,
        )
        domain = _domain_transport_analytic_line(
            date_debut, date_fin,
            vehicule,
            product_id if product_id.isdigit() else None,
            company_id if company_id.isdigit() else None,
        )

        def _distinct_many2one_options(field_name, option_domain=None):
            current_domain = option_domain if option_domain is not None else base_filter_domain
            lines_for_field = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'account.analytic.line', 'search_read',
                [current_domain],
                {'fields': [field_name], 'limit': False},
            )
            seen = {}
            for ln in lines_for_field:
                val = ln.get(field_name)
                if isinstance(val, list) and len(val) >= 2:
                    seen[val[0]] = val[1]
            vals = [{'id': k, 'name': v} for k, v in seen.items()]
            vals.sort(key=lambda x: (x['name'] or '').lower())
            return vals

        product_options = _distinct_many2one_options('product_id', option_domain=options_domain)
        # Société: liste globale, sans filtre transport_logistics.
        company_options = _distinct_many2one_options('company_id', option_domain=[])

        models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.analytic.line', 'search_read',
            [domain],
            {
                'fields': ['date', 'amount', 'nature_id', 'transfer_consumption_id'],
                'order': 'date desc, id desc',
                'limit': 5,
            },
        )

        lines = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.analytic.line', 'search_read',
            [domain],
            {
                'fields': [
                    'date', 'amount', 'nature_id', 'transfer_consumption_id',
                    'product_categ_id', 'general_account_id', 'account_id', 'name',
                ],
                'order': 'date desc, id desc',
                'limit': 5000,
            },
        )

        picking_ids = list({
            ln['transfer_consumption_id'][0]
            for ln in lines
            if ln.get('transfer_consumption_id')
        })
        pickings = {}
        chunk_size = 200
        for i in range(0, len(picking_ids), chunk_size):
            chunk = picking_ids[i : i + chunk_size]
            for p in models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'stock.picking', 'read',
                [chunk],
                {'fields': ['equipment_id', 'affectation_id', 'name']},
            ):
                pickings[p['id']] = p

        def vehicule_label(p):
            if not p:
                return '—'
            if p.get('equipment_id'):
                return extract_matricule(p['equipment_id'][1])
            if p.get('affectation_id'):
                return extract_matricule(p['affectation_id'][1])
            return '—'

        def nature_label(ln):
            if ln.get('nature_id'):
                return ln['nature_id'][1]
            if ln.get('product_categ_id'):
                return ln['product_categ_id'][1]
            if ln.get('general_account_id'):
                return ln['general_account_id'][1]
            if ln.get('account_id'):
                return ln['account_id'][1]
            return ln.get('name') or '—'

        for ln in lines:
            amt = abs(float(ln.get('amount') or 0))
            pid = ln['transfer_consumption_id'][0] if ln.get('transfer_consumption_id') else None
            pk = pickings.get(pid) if pid else None
            nat = nature_label(ln)
            d = (ln.get('date') or '')[:10] or '—'
            rows.append({
                'date': d,
                'vehicule': vehicule_label(pk),
                'nature': nat,
                'montant': round(amt, 2),
            })

        lines_for_nature = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.analytic.line', 'search_read',
            [options_domain],
            {'fields': ['nature_id', 'product_categ_id', 'general_account_id', 'account_id', 'name'], 'limit': False},
        )
        nature_options = sorted({
            nature_label(ln)
            for ln in lines_for_nature
            if nature_label(ln) and nature_label(ln) != '—'
        })
        if nature:
            rows = [r for r in rows if r['nature'] == nature]

        total_montant = round(sum(r['montant'] for r in rows), 2)

        if group_by == 'nature':
            buckets = defaultdict(lambda: {'count': 0, 'montant': 0.0})
            for r in rows:
                k = r['nature']
                buckets[k]['count'] += 1
                buckets[k]['montant'] += r['montant']
            grouped_rows = [
                {'label': k, 'count': v['count'], 'montant': round(v['montant'], 2)}
                for k, v in sorted(buckets.items(), key=lambda kv: (-kv[1]['montant'], kv[0]))
            ]
        elif group_by == 'month':
            buckets = defaultdict(lambda: {'count': 0, 'montant': 0.0})
            for r in rows:
                k = r['date'][:7] if len(r['date']) >= 7 else r['date']
                buckets[k]['count'] += 1
                buckets[k]['montant'] += r['montant']
            grouped_rows = [
                {'label': k, 'count': v['count'], 'montant': round(v['montant'], 2)}
                for k, v in sorted(buckets.items(), key=lambda kv: kv[0], reverse=True)
            ]
        elif group_by == 'vehicle':
            buckets = defaultdict(lambda: {'count': 0, 'montant': 0.0})
            for r in rows:
                k = r['vehicule']
                buckets[k]['count'] += 1
                buckets[k]['montant'] += r['montant']
            grouped_rows = [
                {'label': k, 'count': v['count'], 'montant': round(v['montant'], 2)}
                for k, v in sorted(buckets.items(), key=lambda kv: (-kv[1]['montant'], kv[0]))
            ]

    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return render(request, 'transport/couts_nature.html', {
        'rows': rows,
        'grouped_rows': grouped_rows,
        'group_by': group_by,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'nature': nature,
        'vehicule': vehicule,
        'nature_options': nature_options,
        'vehicules': vehicules,
        'product_id': product_id,
        'company_id': company_id,
        'product_options': product_options,
        'company_options': company_options,
        'total_rows': len(rows),
        'total_montant': total_montant,
    })


@login_required
def transport_facturation_client(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    numero_facture = request.GET.get('numero_facture', '').strip()
    client_id = request.GET.get('client_id', '').strip()
    shipping_id = request.GET.get('shipping_id', '').strip()
    company_id = request.GET.get('company_id', '').strip()
    group_by = request.GET.get('group_by', '').strip()

    rows = []
    clients = []
    companies = []
    shippings = []
    grouped_rows = []
    kpi_paid_rate = 0.0
    kpi_overdue_rate = 0.0
    kpi_ticket_moyen = 0.0
    kpi_top_client = '—'
    kpi_top_client_share = 0.0
    kpi_ca_evolution = 0.0
    error = None
    try:
        uid, models = get_odoo_connection()
        base_domain = [('move_type', '=', 'out_invoice')]

        # Dropdowns depuis Odoo: client, lieu de livraison, société.
        inv_for_filters = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.move', 'search_read',
            [base_domain],
            {'fields': ['partner_id', 'partner_shipping_id', 'company_id', 'project_id', 'invoice_origin', 'ref', 'name'], 'limit': False},
        )
        project_activity_map_filters = _build_project_activity_map(uid, models, inv_for_filters)
        seen_clients = {}
        seen_shippings = {}
        seen_companies = {}
        # Listes de filtres complètes (toutes factures), même si les résultats
        # de la page restent strictement filtrés sur le périmètre transport.
        for inv in inv_for_filters:
            p = inv.get('partner_id')
            if isinstance(p, list) and len(p) >= 2:
                seen_clients[p[0]] = p[1]
            s = inv.get('partner_shipping_id')
            if isinstance(s, list) and len(s) >= 2:
                seen_shippings[s[0]] = s[1]
            c = inv.get('company_id')
            if isinstance(c, list) and len(c) >= 2:
                seen_companies[c[0]] = c[1]
        clients = [{'id': cid, 'name': cname} for cid, cname in seen_clients.items()]
        clients.sort(key=lambda x: (x['name'] or '').lower())
        shippings = [{'id': sid, 'name': sname} for sid, sname in seen_shippings.items()]
        shippings.sort(key=lambda x: (x['name'] or '').lower())
        companies = [{'id': k, 'name': v} for k, v in seen_companies.items()]
        companies.sort(key=lambda x: (x['name'] or '').lower())

        domain = [('move_type', '=', 'out_invoice')]
        domain += _transport_domain_dates(date_debut, date_fin, 'invoice_date')
        if client_id.isdigit():
            domain.append(('partner_id', '=', int(client_id)))
        if shipping_id.isdigit():
            domain.append(('partner_shipping_id', '=', int(shipping_id)))
        if company_id.isdigit():
            domain.append(('company_id', '=', int(company_id)))
        if numero_facture:
            # Recherche élargie pour les cas où l'utilisateur saisit un code projet
            # (ex: S00068) au lieu du numéro exact de facture.
            domain += [
                '|', '|', '|',
                ('name', 'ilike', numero_facture),
                ('ref', 'ilike', numero_facture),
                ('invoice_origin', 'ilike', numero_facture),
                ('project_id.name', 'ilike', numero_facture),
            ]

        invoices = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.move', 'search_read',
            [domain],
            {
                'fields': [
                    'name', 'invoice_date', 'partner_id', 'partner_shipping_id',
                    'company_id', 'amount_untaxed', 'amount_total',
                    'payment_state', 'invoice_date_due',
                    'project_id', 'invoice_origin', 'ref',
                ],
                'order': 'invoice_date desc',
                'limit': 1000,
            }
        )
        project_activity_map = _build_project_activity_map(uid, models, invoices)
        invoices = [inv for inv in invoices if _invoice_activity_bucket(inv, project_activity_map) == 'transport']

        def _fmt_date_fr(iso_date):
            if not iso_date:
                return '—'
            s = str(iso_date)[:10]
            if len(s) == 10 and s[4] == '-' and s[7] == '-':
                return f'{s[8:10]}/{s[5:7]}/{s[0:4]}'
            return s

        for inv in invoices:
            iso_date = (inv.get('invoice_date') or '')[:10]
            ht = round(inv.get('amount_untaxed') or 0, 2)
            ttc = round(inv.get('amount_total') or 0, 2)
            tva = round(ttc - ht, 2)
            projet_ref = '—'
            proj = inv.get('project_id')
            if isinstance(proj, list) and len(proj) >= 2 and proj[1]:
                projet_ref = proj[1]
            elif inv.get('invoice_origin'):
                projet_ref = inv.get('invoice_origin')
            elif inv.get('ref'):
                projet_ref = inv.get('ref')
            rows.append({
                'date': _fmt_date_fr(iso_date),
                'month_key': iso_date[:7] if len(iso_date) >= 7 else '—',
                'numero': inv.get('name') or '—',
                'projet_reference': projet_ref,
                'client': inv['partner_id'][1] if inv.get('partner_id') else '—',
                'lieu_livraison': inv['partner_shipping_id'][1] if inv.get('partner_shipping_id') else '—',
                'societe': inv['company_id'][1] if inv.get('company_id') else '—',
                'ht': ht,
                'tva': tva,
                'ttc': ttc,
                'payment_state': inv.get('payment_state') or '',
                'invoice_date_due': (inv.get('invoice_date_due') or '')[:10],
            })

        if group_by:
            buckets = defaultdict(lambda: {'count': 0, 'ht': 0.0, 'tva': 0.0, 'ttc': 0.0})
            for r in rows:
                if group_by == 'month':
                    key = r['month_key']
                elif group_by == 'company':
                    key = r['societe']
                elif group_by == 'month_company':
                    mois = r['month_key']
                    key = f'{mois} | {r["societe"]}'
                elif group_by == 'month_delivery':
                    mois = r['month_key']
                    key = f'{mois} | {r["lieu_livraison"]}'
                elif group_by == 'client':
                    key = r['client']
                else:
                    key = '—'
                buckets[key]['count'] += 1
                buckets[key]['ht'] += r['ht']
                buckets[key]['tva'] += r['tva']
                buckets[key]['ttc'] += r['ttc']

            mois_fr = {
                '01': 'Janvier', '02': 'Fevrier', '03': 'Mars', '04': 'Avril',
                '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Aout',
                '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Decembre',
            }

            def _month_label(ym):
                if isinstance(ym, str) and len(ym) == 7 and ym[4] == '-':
                    y = ym[:4]
                    m = ym[5:7]
                    return f"{mois_fr.get(m, m)} {y}"
                return ym

            grouped_rows = [
                {
                    'label': k,
                    'count': v['count'],
                    'ht': round(v['ht'], 2),
                    'tva': round(v['tva'], 2),
                    'ttc': round(v['ttc'], 2),
                    'sort_key': k,
                }
                for k, v in buckets.items()
            ]
            if group_by == 'month':
                for g in grouped_rows:
                    g['label'] = _month_label(g['sort_key'])
                grouped_rows.sort(key=lambda x: x['sort_key'])
            elif group_by in {'month_company', 'month_delivery'}:
                for g in grouped_rows:
                    parts = str(g['sort_key']).split(' | ', 1)
                    ym = parts[0]
                    suffix = parts[1] if len(parts) > 1 else ''
                    g['label'] = f"{_month_label(ym)} | {suffix}" if suffix else _month_label(ym)
                grouped_rows.sort(key=lambda x: x['sort_key'])
            else:
                grouped_rows.sort(key=lambda x: (-x['ttc'], x['label']))
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    total_ht = round(sum(r['ht'] for r in rows), 2)
    total_tva = round(sum(r['tva'] for r in rows), 2)
    total_ttc = round(sum(r['ttc'] for r in rows), 2)
    total_rows = len(rows)
    kpi_ticket_moyen = round((total_ttc / total_rows), 2) if total_rows else 0.0

    paid_count = sum(1 for r in rows if r.get('payment_state') == 'paid')
    kpi_paid_rate = round((paid_count / total_rows) * 100, 1) if total_rows else 0.0

    today_iso = date.today().isoformat()
    overdue_count = sum(
        1 for r in rows
        if (r.get('invoice_date_due') and r.get('invoice_date_due') < today_iso and r.get('payment_state') != 'paid')
    )
    kpi_overdue_rate = round((overdue_count / total_rows) * 100, 1) if total_rows else 0.0

    # KPI client leader (part du CA TTC)
    client_totals = defaultdict(float)
    for r in rows:
        client_totals[r.get('client') or '—'] += float(r.get('ttc') or 0)
    if client_totals and total_ttc > 0:
        best_client, best_value = max(client_totals.items(), key=lambda kv: kv[1])
        kpi_top_client = best_client
        kpi_top_client_share = round((best_value / total_ttc) * 100, 1)

    # KPI évolution CA mensuel (% M vs M-1)
    monthly_totals = defaultdict(float)
    for r in rows:
        mk = r.get('month_key') or '—'
        if mk != '—':
            monthly_totals[mk] += float(r.get('ttc') or 0)
    if len(monthly_totals) >= 2:
        keys = sorted(monthly_totals.keys())
        prev_v = monthly_totals[keys[-2]]
        curr_v = monthly_totals[keys[-1]]
        if prev_v > 0:
            kpi_ca_evolution = round(((curr_v - prev_v) / prev_v) * 100, 1)

    # Section séparée: CA par lieu de livraison.
    delivery_buckets = defaultdict(lambda: {'count': 0, 'ht': 0.0, 'tva': 0.0, 'ttc': 0.0})
    for r in rows:
        k = r['lieu_livraison']
        delivery_buckets[k]['count'] += 1
        delivery_buckets[k]['ht'] += r['ht']
        delivery_buckets[k]['tva'] += r['tva']
        delivery_buckets[k]['ttc'] += r['ttc']
    delivery_rows = [
        {
            'lieu': k,
            'count': v['count'],
            'ht': round(v['ht'], 2),
            'tva': round(v['tva'], 2),
            'ttc': round(v['ttc'], 2),
        }
        for k, v in sorted(delivery_buckets.items(), key=lambda kv: (-kv[1]['ttc'], kv[0]))
    ]
    delivery_total_count = sum(r['count'] for r in delivery_rows)
    delivery_total_ht = round(sum(r['ht'] for r in delivery_rows), 2)
    delivery_total_tva = round(sum(r['tva'] for r in delivery_rows), 2)
    delivery_total_ttc = round(sum(r['ttc'] for r in delivery_rows), 2)

    return render(request, 'transport/facturation_client.html', {
        'rows': rows,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'numero_facture': numero_facture,
        'client_id': client_id,
        'clients': clients,
        'shipping_id': shipping_id,
        'shippings': shippings,
        'company_id': company_id,
        'companies': companies,
        'group_by': group_by,
        'grouped_rows': grouped_rows,
        'total_rows': total_rows,
        'total_ht': total_ht,
        'total_tva': total_tva,
        'total_ttc': total_ttc,
        'kpi_paid_rate': kpi_paid_rate,
        'kpi_overdue_rate': kpi_overdue_rate,
        'kpi_ticket_moyen': kpi_ticket_moyen,
        'kpi_top_client': kpi_top_client,
        'kpi_top_client_share': kpi_top_client_share,
        'kpi_ca_evolution': kpi_ca_evolution,
        'delivery_rows': delivery_rows,
        'delivery_total_count': delivery_total_count,
        'delivery_total_ht': delivery_total_ht,
        'delivery_total_tva': delivery_total_tva,
        'delivery_total_ttc': delivery_total_ttc,
    })


@login_required
def production_index(request):
    return render(request, 'production/production.html', {})


@login_required
def production_gasoil(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    societe = request.GET.get('societe', '').strip()
    site = request.GET.get('site', '').strip()
    statut = request.GET.get('statut', '').strip()
    chauffeur = request.GET.get('chauffeur', '').strip()
    ouvrage = request.GET.get('ouvrage', '').strip()
    activite = request.GET.get('activite', 'production').strip() or 'production'
    export = request.GET.get('export', '').strip().lower()

    rows = []
    error = None
    societes = []
    sites = []
    chauffeurs = []
    ouvrages = []
    total_litres = 0.0
    total_montant = 0.0
    total_rows = 0
    conso_moyenne = 0.0
    nb_anomalies = 0
    try:
        uid, models = get_odoo_connection()

        list_domain = _build_sorties_domain(
            date_debut='',
            date_fin='',
            site='',
            chauffeur='',
            ouvrage='',
            anomalie='',
            societe='',
            categorie_engin='',
            activite_filtre='production',
        )
        list_domain.append(('equipment_id.category_id', 'in', CATEGORIES_PRODUCTION))
        bons_for_filters = _fetch_sorties_bons(uid, models, list_domain, limit=5000)
        societes = sorted({b['societe'] for b in bons_for_filters if b.get('societe') and b.get('societe') != '—'})
        sites = sorted({b['site'] for b in bons_for_filters if b.get('site') and b.get('site') != '—'})
        chauffeurs = sorted({b['chauffeur'] for b in bons_for_filters if b.get('chauffeur') and b.get('chauffeur') != '—'})
        ouvrages = sorted({b['ouvrage'] for b in bons_for_filters if b.get('ouvrage') and b.get('ouvrage') != '—'})

        domain = _build_sorties_domain(
            date_debut=date_debut,
            date_fin=date_fin,
            site=site,
            chauffeur=chauffeur,
            ouvrage='',
            anomalie='',
            societe=societe,
            categorie_engin='',
            activite_filtre='production',
        )
        domain.append(('equipment_id.category_id', 'in', CATEGORIES_PRODUCTION))
        limit = 10000 if (date_debut or date_fin or societe or site or chauffeur) else 2000
        bons = _fetch_sorties_bons(uid, models, domain, limit=limit)

        # Garantit que la page Production n'affiche que le périmètre exploitation/production.
        bons = [b for b in bons if b.get('activite_bucket') == 'production']

        if ouvrage:
            o = ouvrage.lower()
            bons = [b for b in bons if o in (b.get('ouvrage') or '').lower()]

        if statut == 'ok':
            bons = [b for b in bons if b.get('anomalie') == 'OK']
        elif statut == 'anomalie':
            bons = [b for b in bons if b.get('anomalie') == 'Anomalie']

        move_domain = [
            ['picking_id', 'in', [b['id'] for b in bons]],
            ['product_id.categ_id', '=', CARBURANT_CATEG_ID],
        ]
        move_lines = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'stock.move', 'search_read',
            [move_domain],
            {'fields': ['picking_id', 'product_qty', 'price_total', 'unit_price'], 'limit': 20000},
        )
        amount_by_picking = defaultdict(float)
        for mv in move_lines:
            pid = mv['picking_id'][0] if mv.get('picking_id') else None
            if not pid:
                continue
            price_total = mv.get('price_total')
            if price_total not in (None, False):
                amount_by_picking[pid] += price_total or 0
            else:
                amount_by_picking[pid] += (mv.get('product_qty') or 0) * (mv.get('unit_price') or 0)

        def _fmt_date_fr(iso_date):
            if not iso_date or iso_date == '—' or len(iso_date) != 10:
                return iso_date or '—'
            return f'{iso_date[8:10]}/{iso_date[5:7]}/{iso_date[0:4]}'

        rows = [{
            'date': _fmt_date_fr(b.get('date') or '—'),
            'vehicule': extract_matricule(b.get('engin') or '—'),
            'conducteur': b.get('chauffeur') or '—',
            'ouvrage': b.get('ouvrage') or '—',
            'societe': b.get('societe') or '—',
            'site': b.get('site') or '—',
            'litres': round(b.get('product_qty') or 0, 2),
            'montant': round(amount_by_picking.get(b['id'], 0), 2),
            'compteur': b.get('cpt_actuel') or 0,
            'anomalie': b.get('anomalie') or 'OK',
        } for b in bons]

        total_rows = len(rows)
        total_litres = round(sum(r['litres'] for r in rows), 2)
        total_montant = round(sum(r['montant'] for r in rows), 2)
        nb_anomalies = sum(1 for b in bons if b.get('anomalie') == 'Anomalie')
        conso_vals = [b.get('consommation') or 0 for b in bons if (b.get('consommation') or 0) > 0]
        conso_moyenne = round(sum(conso_vals) / len(conso_vals), 2) if conso_vals else 0.0

        if export in {'excel', 'csv'}:
            if export == 'csv':
                out = io.StringIO()
                writer = csv.writer(out, delimiter=';')
                writer.writerow(['Date', 'Véhicule', 'Conducteur/Opérateur', 'Société', 'Site', 'Ouvrage', 'Litres', 'Montant', 'Compteur', 'Statut'])
                for r in rows:
                    writer.writerow([r['date'], r['vehicule'], r['conducteur'], r['societe'], r['site'], r['ouvrage'], r['litres'], r['montant'], r['compteur'], r['anomalie']])
                resp = HttpResponse(out.getvalue(), content_type='text/csv; charset=utf-8')
                resp['Content-Disposition'] = 'attachment; filename=production_gasoil_sorties.csv'
                return resp

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Production Gasoil'
            ws.append(['Date', 'Véhicule', 'Conducteur/Opérateur', 'Société', 'Site', 'Ouvrage', 'Litres', 'Montant', 'Compteur', 'Statut'])
            for r in rows:
                ws.append([r['date'], r['vehicule'], r['conducteur'], r['societe'], r['site'], r['ouvrage'], r['litres'], r['montant'], r['compteur'], r['anomalie']])
            ws.append(['TOTAL', '', '', '', '', '', total_litres, total_montant, '', ''])
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(
                bio.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            resp['Content-Disposition'] = 'attachment; filename=production_gasoil_sorties.xlsx'
            return resp
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return render(request, 'production/gasoil.html', {
        'rows': rows,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'societe': societe,
        'site': site,
        'statut': statut,
        'chauffeur': chauffeur,
        'ouvrage': ouvrage,
        'activite': activite,
        'societes': societes,
        'sites': sites,
        'chauffeurs': chauffeurs,
        'ouvrages': ouvrages,
        'total_rows': total_rows,
        'total_litres': total_litres,
        'total_montant': total_montant,
        'conso_moyenne': conso_moyenne,
        'nb_anomalies': nb_anomalies,
    })


def _domain_production_analytic_line(date_debut, date_fin, vehicule, product_id=None, company_id=None):
    domain = []
    if date_debut:
        domain.append(('date', '>=', date_debut))
    if date_fin:
        domain.append(('date', '<=', date_fin))
    domain.extend([
        ('transfer_consumption_id', '!=', False),
        ('transfer_consumption_id.transport_logistics', '=', False),
        ('transfer_consumption_id.service_car', '=', False),
        ('transfer_consumption_id.equipment_id.category_id', 'in', CATEGORIES_PRODUCTION),
    ])
    if vehicule:
        v = str(vehicule).strip()
        domain.extend([
            '|',
            ('transfer_consumption_id.equipment_id.name', 'ilike', v),
            ('transfer_consumption_id.affectation_id.name', 'ilike', v),
        ])
    if product_id:
        domain.append(('product_id', '=', int(product_id)))
    if company_id:
        domain.append(('company_id', '=', int(company_id)))
    return domain


@login_required
def production_couts_nature(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    nature = request.GET.get('nature', '').strip()
    vehicule = request.GET.get('vehicule', '').strip()
    product_id = request.GET.get('product_id', '').strip()
    company_id = request.GET.get('company_id', '').strip()
    group_by = request.GET.get('group_by', '').strip()

    rows = []
    grouped_rows = []
    nature_options = []
    vehicules = []
    product_options = []
    company_options = []
    error = None
    total_montant = 0.0

    try:
        uid, models = get_odoo_connection()

        vehicules_raw = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'maintenance.equipment', 'search_read',
            [[]],
            {'fields': ['id', 'name'], 'order': 'name', 'limit': 500},
        )
        vehicules = sorted({extract_matricule(v.get('name')) for v in vehicules_raw if v.get('name')})

        base_filter_domain = _domain_production_analytic_line(date_debut, date_fin, vehicule)
        options_domain = _domain_production_analytic_line(date_debut, date_fin, None)
        domain = _domain_production_analytic_line(
            date_debut, date_fin, vehicule,
            product_id if product_id.isdigit() else None,
            company_id if company_id.isdigit() else None,
        )

        def _distinct_many2one_options(field_name, option_domain=None):
            current_domain = option_domain if option_domain is not None else base_filter_domain
            lines_for_field = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'account.analytic.line', 'search_read',
                [current_domain],
                {'fields': [field_name], 'limit': False},
            )
            seen = {}
            for ln in lines_for_field:
                val = ln.get(field_name)
                if isinstance(val, list) and len(val) >= 2:
                    seen[val[0]] = val[1]
            vals = [{'id': k, 'name': v} for k, v in seen.items()]
            vals.sort(key=lambda x: (x['name'] or '').lower())
            return vals

        product_options = _distinct_many2one_options('product_id', option_domain=options_domain)
        company_options = _distinct_many2one_options('company_id', option_domain=[])

        lines = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.analytic.line', 'search_read',
            [domain],
            {
                'fields': ['date', 'amount', 'nature_id', 'transfer_consumption_id', 'product_categ_id', 'general_account_id', 'account_id', 'name'],
                'order': 'date desc, id desc',
                'limit': 5000,
            },
        )

        picking_ids = list({ln['transfer_consumption_id'][0] for ln in lines if ln.get('transfer_consumption_id')})
        pickings = {}
        for i in range(0, len(picking_ids), 200):
            for p in models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'stock.picking', 'read',
                [picking_ids[i:i + 200]],
                {'fields': ['equipment_id', 'affectation_id', 'name']},
            ):
                pickings[p['id']] = p

        def vehicule_label(p):
            if not p:
                return '—'
            if p.get('equipment_id'):
                return extract_matricule(p['equipment_id'][1])
            if p.get('affectation_id'):
                return extract_matricule(p['affectation_id'][1])
            return '—'

        def nature_label(ln):
            if ln.get('nature_id'):
                return ln['nature_id'][1]
            if ln.get('product_categ_id'):
                return ln['product_categ_id'][1]
            if ln.get('general_account_id'):
                return ln['general_account_id'][1]
            if ln.get('account_id'):
                return ln['account_id'][1]
            return ln.get('name') or '—'

        for ln in lines:
            amt = abs(float(ln.get('amount') or 0))
            pid = ln['transfer_consumption_id'][0] if ln.get('transfer_consumption_id') else None
            pk = pickings.get(pid) if pid else None
            nat = nature_label(ln)
            rows.append({
                'date': (ln.get('date') or '')[:10] or '—',
                'vehicule': vehicule_label(pk),
                'nature': nat,
                'montant': round(amt, 2),
            })

        lines_for_nature = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.analytic.line', 'search_read',
            [options_domain],
            {'fields': ['nature_id', 'product_categ_id', 'general_account_id', 'account_id', 'name'], 'limit': False},
        )
        nature_options = sorted({nature_label(ln) for ln in lines_for_nature if nature_label(ln) and nature_label(ln) != '—'})
        if nature:
            rows = [r for r in rows if r['nature'] == nature]

        total_montant = round(sum(r['montant'] for r in rows), 2)

        if group_by == 'nature':
            buckets = defaultdict(lambda: {'count': 0, 'montant': 0.0})
            for r in rows:
                buckets[r['nature']]['count'] += 1
                buckets[r['nature']]['montant'] += r['montant']
            grouped_rows = [{'label': k, 'count': v['count'], 'montant': round(v['montant'], 2)} for k, v in sorted(buckets.items(), key=lambda kv: (-kv[1]['montant'], kv[0]))]
        elif group_by == 'month':
            buckets = defaultdict(lambda: {'count': 0, 'montant': 0.0})
            for r in rows:
                k = r['date'][:7] if len(r['date']) >= 7 else r['date']
                buckets[k]['count'] += 1
                buckets[k]['montant'] += r['montant']
            grouped_rows = [{'label': k, 'count': v['count'], 'montant': round(v['montant'], 2)} for k, v in sorted(buckets.items(), key=lambda kv: kv[0], reverse=True)]
        elif group_by == 'vehicle':
            buckets = defaultdict(lambda: {'count': 0, 'montant': 0.0})
            for r in rows:
                buckets[r['vehicule']]['count'] += 1
                buckets[r['vehicule']]['montant'] += r['montant']
            grouped_rows = [{'label': k, 'count': v['count'], 'montant': round(v['montant'], 2)} for k, v in sorted(buckets.items(), key=lambda kv: (-kv[1]['montant'], kv[0]))]
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return render(request, 'production/couts_nature.html', {
        'rows': rows,
        'grouped_rows': grouped_rows,
        'group_by': group_by,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'nature': nature,
        'vehicule': vehicule,
        'nature_options': nature_options,
        'vehicules': vehicules,
        'product_id': product_id,
        'company_id': company_id,
        'product_options': product_options,
        'company_options': company_options,
        'total_rows': len(rows),
        'total_montant': total_montant,
    })


@login_required
def production_facturation_ventes(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    numero_facture = request.GET.get('numero_facture', '').strip()
    client_id = request.GET.get('client_id', '').strip()
    shipping_id = request.GET.get('shipping_id', '').strip()
    company_id = request.GET.get('company_id', '').strip()
    etat = request.GET.get('etat', '').strip()
    paiement = request.GET.get('paiement', '').strip()
    commercial_id = request.GET.get('commercial_id', '').strip()
    due_date = request.GET.get('due_date', '').strip()
    group_by = request.GET.get('group_by', '').strip()

    rows = []
    clients = []
    companies = []
    shippings = []
    commercials = []
    grouped_rows = []
    kpi_paid_rate = 0.0
    kpi_overdue_rate = 0.0
    kpi_ticket_moyen = 0.0
    kpi_top_client = '—'
    kpi_top_client_share = 0.0
    kpi_ca_evolution = 0.0
    error = None
    try:
        uid, models = get_odoo_connection()
        # Base commune factures de vente; séparation métier faite ensuite en Python.
        base_domain = [('move_type', '=', 'out_invoice')]

        inv_for_filters = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.move', 'search_read',
            [base_domain],
            {'fields': ['partner_id', 'partner_shipping_id', 'company_id', 'invoice_user_id', 'project_id', 'invoice_origin', 'ref', 'name'], 'limit': False},
        )
        project_activity_map_filters = _build_project_activity_map(uid, models, inv_for_filters)
        seen_clients = {}
        seen_shippings = {}
        seen_companies = {}
        seen_commercials = {}
        # Listes de filtres complètes (toutes factures), même si les résultats
        # de la page restent strictement filtrés sur le périmètre production.
        for inv in inv_for_filters:
            p = inv.get('partner_id')
            if isinstance(p, list) and len(p) >= 2:
                seen_clients[p[0]] = p[1]
            s = inv.get('partner_shipping_id')
            if isinstance(s, list) and len(s) >= 2:
                seen_shippings[s[0]] = s[1]
            c = inv.get('company_id')
            if isinstance(c, list) and len(c) >= 2:
                seen_companies[c[0]] = c[1]
            u = inv.get('invoice_user_id')
            if isinstance(u, list) and len(u) >= 2:
                seen_commercials[u[0]] = u[1]
        clients = [{'id': cid, 'name': cname} for cid, cname in seen_clients.items()]
        clients.sort(key=lambda x: (x['name'] or '').lower())
        shippings = [{'id': sid, 'name': sname} for sid, sname in seen_shippings.items()]
        shippings.sort(key=lambda x: (x['name'] or '').lower())
        companies = [{'id': cid, 'name': cname} for cid, cname in seen_companies.items()]
        companies.sort(key=lambda x: (x['name'] or '').lower())
        commercials = [{'id': uid_, 'name': uname} for uid_, uname in seen_commercials.items()]
        commercials.sort(key=lambda x: (x['name'] or '').lower())

        domain = list(base_domain)
        if date_debut:
            domain.append(('invoice_date', '>=', date_debut))
        if date_fin:
            domain.append(('invoice_date', '<=', date_fin))
        if client_id.isdigit():
            domain.append(('partner_id', '=', int(client_id)))
        if shipping_id.isdigit():
            domain.append(('partner_shipping_id', '=', int(shipping_id)))
        if company_id.isdigit():
            domain.append(('company_id', '=', int(company_id)))
        if commercial_id.isdigit():
            domain.append(('invoice_user_id', '=', int(commercial_id)))
        if due_date:
            domain.append(('invoice_date_due', '=', due_date))
        if numero_facture:
            # Recherche élargie pour les cas où l'utilisateur saisit un code projet
            # (ex: S00068) au lieu du numéro exact de facture.
            domain += [
                '|', '|', '|',
                ('name', 'ilike', numero_facture),
                ('ref', 'ilike', numero_facture),
                ('invoice_origin', 'ilike', numero_facture),
                ('project_id.name', 'ilike', numero_facture),
            ]
        if etat in {'posted', 'draft', 'cancel'}:
            domain.append(('state', '=', etat))
        if paiement in {'paid', 'not_paid', 'in_payment'}:
            domain.append(('payment_state', '=', paiement))

        invoices = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.move', 'search_read',
            [domain],
            {
                'fields': [
                    'name', 'invoice_date', 'partner_id', 'partner_shipping_id',
                    'company_id', 'invoice_user_id', 'invoice_date_due',
                    'amount_untaxed', 'amount_total', 'state', 'payment_state',
                    'project_id', 'invoice_origin', 'ref',
                ],
                'order': 'invoice_date desc',
                'limit': 2000,
            }
        )
        project_activity_map = _build_project_activity_map(uid, models, invoices)
        invoices = [inv for inv in invoices if _invoice_activity_bucket(inv, project_activity_map) == 'production']

        def _fmt_date_fr(iso_date):
            if not iso_date:
                return '—'
            s = str(iso_date)[:10]
            if len(s) == 10 and s[4] == '-' and s[7] == '-':
                return f'{s[8:10]}/{s[5:7]}/{s[0:4]}'
            return s

        for inv in invoices:
            iso_date = (inv.get('invoice_date') or '')[:10]
            ht = round(inv.get('amount_untaxed') or 0, 2)
            ttc = round(inv.get('amount_total') or 0, 2)
            tva = round(ttc - ht, 2)
            projet_ref = '—'
            proj = inv.get('project_id')
            if isinstance(proj, list) and len(proj) >= 2 and proj[1]:
                projet_ref = proj[1]
            elif inv.get('invoice_origin'):
                projet_ref = inv.get('invoice_origin')
            elif inv.get('ref'):
                projet_ref = inv.get('ref')
            rows.append({
                'numero': inv.get('name') or '—',
                'date': _fmt_date_fr(iso_date),
                'month_key': iso_date[:7] if len(iso_date) >= 7 else '—',
                'projet_reference': projet_ref,
                'client': inv['partner_id'][1] if inv.get('partner_id') else '—',
                'lieu_livraison': inv['partner_shipping_id'][1] if inv.get('partner_shipping_id') else '—',
                'societe': inv['company_id'][1] if inv.get('company_id') else '—',
                'ht': ht,
                'tva': tva,
                'ttc': ttc,
                'payment_state': inv.get('payment_state') or '',
                'invoice_date_due': (inv.get('invoice_date_due') or '')[:10],
            })

        if group_by:
            buckets = defaultdict(lambda: {'count': 0, 'ht': 0.0, 'tva': 0.0, 'ttc': 0.0})
            for r in rows:
                if group_by == 'month':
                    key = r['month_key']
                elif group_by == 'company':
                    key = r['societe']
                elif group_by == 'month_company':
                    key = f'{r["month_key"]} | {r["societe"]}'
                elif group_by == 'month_delivery':
                    key = f'{r["month_key"]} | {r["lieu_livraison"]}'
                else:
                    key = '—'
                buckets[key]['count'] += 1
                buckets[key]['ht'] += r['ht']
                buckets[key]['tva'] += r['tva']
                buckets[key]['ttc'] += r['ttc']

            mois_fr = {
                '01': 'Janvier', '02': 'Fevrier', '03': 'Mars', '04': 'Avril',
                '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Aout',
                '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Decembre',
            }

            def _month_label(ym):
                if isinstance(ym, str) and len(ym) == 7 and ym[4] == '-':
                    return f"{mois_fr.get(ym[5:7], ym[5:7])} {ym[:4]}"
                return ym

            grouped_rows = [
                {
                    'label': k,
                    'sort_key': k,
                    'count': v['count'],
                    'ht': round(v['ht'], 2),
                    'tva': round(v['tva'], 2),
                    'ttc': round(v['ttc'], 2),
                }
                for k, v in buckets.items()
            ]
            if group_by == 'month':
                for g in grouped_rows:
                    g['label'] = _month_label(g['sort_key'])
                grouped_rows.sort(key=lambda x: x['sort_key'])
            elif group_by in {'month_company', 'month_delivery'}:
                for g in grouped_rows:
                    ym, suffix = str(g['sort_key']).split(' | ', 1)
                    g['label'] = f'{_month_label(ym)} | {suffix}'
                grouped_rows.sort(key=lambda x: x['sort_key'])
            else:
                grouped_rows.sort(key=lambda x: (-x['ttc'], x['label']))
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    total_ht = round(sum(r['ht'] for r in rows), 2)
    total_tva = round(sum(r['tva'] for r in rows), 2)
    total_ttc = round(sum(r['ttc'] for r in rows), 2)
    total_rows = len(rows)
    kpi_ticket_moyen = round((total_ttc / total_rows), 2) if total_rows else 0.0

    paid_count = sum(1 for r in rows if r.get('payment_state') == 'paid')
    kpi_paid_rate = round((paid_count / total_rows) * 100, 1) if total_rows else 0.0

    today_iso = date.today().isoformat()
    overdue_count = sum(
        1 for r in rows
        if (r.get('invoice_date_due') and r.get('invoice_date_due') < today_iso and r.get('payment_state') != 'paid')
    )
    kpi_overdue_rate = round((overdue_count / total_rows) * 100, 1) if total_rows else 0.0

    # KPI client leader (part du CA TTC)
    client_totals = defaultdict(float)
    for r in rows:
        client_totals[r.get('client') or '—'] += float(r.get('ttc') or 0)
    if client_totals and total_ttc > 0:
        best_client, best_value = max(client_totals.items(), key=lambda kv: kv[1])
        kpi_top_client = best_client
        kpi_top_client_share = round((best_value / total_ttc) * 100, 1)

    # KPI évolution CA mensuel (% M vs M-1)
    monthly_totals = defaultdict(float)
    for r in rows:
        mk = r.get('month_key') or '—'
        if mk != '—':
            monthly_totals[mk] += float(r.get('ttc') or 0)
    if len(monthly_totals) >= 2:
        keys = sorted(monthly_totals.keys())
        prev_v = monthly_totals[keys[-2]]
        curr_v = monthly_totals[keys[-1]]
        if prev_v > 0:
            kpi_ca_evolution = round(((curr_v - prev_v) / prev_v) * 100, 1)

    delivery_buckets = defaultdict(lambda: {'count': 0, 'ht': 0.0, 'tva': 0.0, 'ttc': 0.0})
    for r in rows:
        k = r['lieu_livraison']
        delivery_buckets[k]['count'] += 1
        delivery_buckets[k]['ht'] += r['ht']
        delivery_buckets[k]['tva'] += r['tva']
        delivery_buckets[k]['ttc'] += r['ttc']
    delivery_rows = [
        {
            'lieu': k,
            'count': v['count'],
            'ht': round(v['ht'], 2),
            'tva': round(v['tva'], 2),
            'ttc': round(v['ttc'], 2),
        }
        for k, v in sorted(delivery_buckets.items(), key=lambda kv: (-kv[1]['ttc'], kv[0]))
    ]
    delivery_total_count = sum(r['count'] for r in delivery_rows)
    delivery_total_ht = round(sum(r['ht'] for r in delivery_rows), 2)
    delivery_total_tva = round(sum(r['tva'] for r in delivery_rows), 2)
    delivery_total_ttc = round(sum(r['ttc'] for r in delivery_rows), 2)

    return render(request, 'production/facturation_ventes.html', {
        'rows': rows,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'numero_facture': numero_facture,
        'client_id': client_id,
        'shipping_id': shipping_id,
        'company_id': company_id,
        'etat': etat,
        'paiement': paiement,
        'commercial_id': commercial_id,
        'due_date': due_date,
        'clients': clients,
        'shippings': shippings,
        'companies': companies,
        'commercials': commercials,
        'group_by': group_by,
        'grouped_rows': grouped_rows,
        'total_rows': total_rows,
        'total_ht': total_ht,
        'total_tva': total_tva,
        'total_ttc': total_ttc,
        'kpi_paid_rate': kpi_paid_rate,
        'kpi_overdue_rate': kpi_overdue_rate,
        'kpi_ticket_moyen': kpi_ticket_moyen,
        'kpi_top_client': kpi_top_client,
        'kpi_top_client_share': kpi_top_client_share,
        'kpi_ca_evolution': kpi_ca_evolution,
        'delivery_rows': delivery_rows,
        'delivery_total_count': delivery_total_count,
        'delivery_total_ht': delivery_total_ht,
        'delivery_total_tva': delivery_total_tva,
        'delivery_total_ttc': delivery_total_ttc,
    })


@login_required
def production_sites(request):
    date_debut = request.GET.get('date_debut', '').strip()
    date_fin = request.GET.get('date_fin', '').strip()
    site = request.GET.get('site', '').strip()
    societe = request.GET.get('societe', '').strip()
    tri = request.GET.get('tri', 'optimisation').strip() or 'optimisation'

    rows = []
    sites = []
    societes = []
    kpi_sites = 0
    kpi_bons = 0
    kpi_litres = 0.0
    kpi_montant = 0.0
    kpi_anomalies = 0
    chart_labels = []
    chart_values = []
    top_site_name = '—'
    top_site_montant = 0.0
    top_site_score = 0.0
    worst_site_name = '—'
    worst_site_ratio = 0.0
    no_data_for_filters = False
    error = None

    try:
        uid, models = get_odoo_connection()

        list_domain = _build_sorties_domain(
            date_debut='',
            date_fin='',
            site='',
            chauffeur='',
            ouvrage='',
            anomalie='',
            societe='',
            categorie_engin='',
            activite_filtre='production',
        )
        list_domain.append(('equipment_id.category_id', 'in', CATEGORIES_PRODUCTION))
        bons_for_filters = _fetch_sorties_bons(uid, models, list_domain, limit=6000)
        bons_for_filters = [b for b in bons_for_filters if b.get('activite_bucket') == 'production']
        sites = sorted({b['site'] for b in bons_for_filters if b.get('site') and b.get('site') != '—'})
        societes = sorted({b['societe'] for b in bons_for_filters if b.get('societe') and b.get('societe') != '—'})

        domain = _build_sorties_domain(
            date_debut=date_debut,
            date_fin=date_fin,
            site=site,
            chauffeur='',
            ouvrage='',
            anomalie='',
            societe=societe,
            categorie_engin='',
            activite_filtre='production',
        )
        domain.append(('equipment_id.category_id', 'in', CATEGORIES_PRODUCTION))
        bons = _fetch_sorties_bons(uid, models, domain, limit=12000)
        bons = [b for b in bons if b.get('activite_bucket') == 'production']

        move_domain = [
            ['picking_id', 'in', [b['id'] for b in bons]],
            ['product_id.categ_id', '=', CARBURANT_CATEG_ID],
        ]
        move_lines = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'stock.move', 'search_read',
            [move_domain],
            {'fields': ['picking_id', 'product_qty', 'price_total', 'unit_price'], 'limit': 30000},
        )
        amount_by_picking = defaultdict(float)
        for mv in move_lines:
            pid = mv['picking_id'][0] if mv.get('picking_id') else None
            if not pid:
                continue
            price_total = mv.get('price_total')
            if price_total not in (None, False):
                amount_by_picking[pid] += price_total or 0
            else:
                amount_by_picking[pid] += (mv.get('product_qty') or 0) * (mv.get('unit_price') or 0)

        site_bucket = defaultdict(lambda: {
            'site': '—', 'societes': set(), 'nb_bons': 0, 'litres': 0.0, 'montant': 0.0, 'anomalies': 0
        })
        for b in bons:
            site_name = b.get('site') or '—'
            cur = site_bucket[site_name]
            cur['site'] = site_name
            cur['nb_bons'] += 1
            cur['litres'] += float(b.get('product_qty') or 0)
            cur['montant'] += float(amount_by_picking.get(b['id'], 0))
            cur['anomalies'] += 1 if b.get('anomalie') == 'Anomalie' else 0
            if b.get('societe') and b.get('societe') != '—':
                cur['societes'].add(b['societe'])

        rows = []
        for _, item in site_bucket.items():
            litres = round(item['litres'], 2)
            montant = round(item['montant'], 2)
            nb_bons = item['nb_bons']
            anomalies = item['anomalies']
            ratio_anomalies = round((anomalies / nb_bons) * 100, 1) if nb_bons else 0
            litres_par_bon = round((litres / nb_bons), 2) if nb_bons else 0.0
            rows.append({
                'site': item['site'],
                'societe': ', '.join(sorted(item['societes'])) if item['societes'] else '—',
                'nb_bons': nb_bons,
                'litres': litres,
                'montant': montant,
                'anomalies': anomalies,
                'ratio_anomalies': ratio_anomalies,
                'litres_par_bon': litres_par_bon,
                'optimisation_score': 0.0,
            })

        # ── Score d'optimisation pondéré (explicite) ─────────────────────────
        # 60% consommation (litres/bon bas = meilleur)
        # 30% anomalies (taux bas = meilleur)
        # 10% volume d'activité (plus de bons = meilleur)
        if rows:
            max_lpb = max(r['litres_par_bon'] for r in rows) or 1.0
            max_anom = max(r['ratio_anomalies'] for r in rows) or 1.0
            max_bons = max(r['nb_bons'] for r in rows) or 1.0

            for r in rows:
                consommation_norm = max(0.0, 100 * (1 - (r['litres_par_bon'] / max_lpb)))
                anomalies_norm = max(0.0, 100 * (1 - (r['ratio_anomalies'] / max_anom))) if max_anom > 0 else 100.0
                volume_norm = max(0.0, 100 * (r['nb_bons'] / max_bons))
                score = (0.60 * consommation_norm) + (0.30 * anomalies_norm) + (0.10 * volume_norm)
                r['optimisation_score'] = round(score, 2)

        if tri == 'optimisation':
            rows.sort(key=lambda r: (-r['optimisation_score'], r['site']))
        elif tri == 'bons':
            rows.sort(key=lambda r: (-r['nb_bons'], r['site']))
        elif tri == 'litres':
            rows.sort(key=lambda r: (-r['litres'], r['site']))
        elif tri == 'anomalies':
            rows.sort(key=lambda r: (-r['anomalies'], r['site']))
        else:
            rows.sort(key=lambda r: (-r['montant'], r['site']))

        kpi_sites = len(rows)
        kpi_bons = sum(r['nb_bons'] for r in rows)
        kpi_litres = round(sum(r['litres'] for r in rows), 2)
        kpi_montant = round(sum(r['montant'] for r in rows), 2)
        kpi_anomalies = sum(r['anomalies'] for r in rows)

        # Top 8 toujours basé sur l'optimisation (pas sur le montant)
        top_chart = sorted(rows, key=lambda r: (-r['optimisation_score'], r['site']))[:8]
        chart_labels = [r['site'] for r in top_chart]
        chart_values = [r['optimisation_score'] for r in top_chart]

        if rows:
            best = max(rows, key=lambda r: (r['optimisation_score'], r['nb_bons']))
            top_site_name = best['site']
            top_site_montant = best['montant']
            top_site_score = best['optimisation_score']
            worst = max(rows, key=lambda r: (r['ratio_anomalies'], r['anomalies']))
            worst_site_name = worst['site']
            worst_site_ratio = worst['ratio_anomalies']
        no_data_for_filters = len(rows) == 0
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return render(request, 'production/sites.html', {
        'rows': rows,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'site': site,
        'societe': societe,
        'tri': tri,
        'sites': sites,
        'societes': societes,
        'kpi_sites': kpi_sites,
        'kpi_bons': kpi_bons,
        'kpi_litres': kpi_litres,
        'kpi_montant': kpi_montant,
        'kpi_anomalies': kpi_anomalies,
        'top_site_name': top_site_name,
        'top_site_montant': top_site_montant,
        'top_site_score': top_site_score,
        'worst_site_name': worst_site_name,
        'worst_site_ratio': worst_site_ratio,
        'no_data_for_filters': no_data_for_filters,
        'chart_has_data': len(chart_labels) > 0,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_values_json': json.dumps(chart_values),
    })


def _render_achats_module(request, current_key, page_title, page_subtitle):
    menu_items = [
        {'key': 'overview', 'label': "Vue d'ensemble", 'url': 'achats_overview'},
        {'key': 'purchase_requests', 'label': "Demandes d'achat", 'url': 'achats_purchase_requests'},
        {'key': 'rfq', 'label': 'Demandes de prix', 'url': 'achats_rfq'},
        {'key': 'purchase_orders', 'label': 'Bons de commande', 'url': 'achats_purchase_orders'},
        {'key': 'delivery_tracking', 'label': 'Suivi livraisons', 'url': 'achats_delivery_tracking'},
        {'key': 'suppliers', 'label': 'Fournisseurs', 'url': 'achats_suppliers'},
    ]

    return render(request, 'achats/module.html', {
        'page_title': page_title,
        'page_subtitle': page_subtitle,
        'module_key': current_key,
        'menu_items': menu_items,
    })


@login_required
def achats_overview(request):
    return _render_achats_module(
        request,
        current_key='overview',
        page_title="Vue d'ensemble",
        page_subtitle="Vision globale du flux achats : demandes, commandes, livraisons et fournisseurs.",
    )


@login_required
def achats_purchase_requests(request):
    date_debut  = request.GET.get('date_debut',  '').strip()
    date_fin    = request.GET.get('date_fin',    '').strip()
    departement = request.GET.get('departement', '').strip()
    etat        = request.GET.get('etat',        '').strip()
    demandeur   = request.GET.get('demandeur',   '').strip()
    export      = request.GET.get('export',      '').strip().lower()

    STATE_LABELS_PR = {
        'draft':      'Brouillon',
        'to_approve': 'Confirmé',
        'confirmed':  'Confirmé',
        'approved':   'Approuvé',
        'rejected':   'Rejeté',
        'done':       'Terminé',
        'sent':       'Envoyé',
        'purchase':   'Bon de commande',
        'cancel':     'Annulé',
    }

    rows = []
    error = None
    departements = []
    demandeurs_list = []
    total_demandes = 0
    total_montant  = 0.0
    nb_attente     = 0
    nb_approuvees  = 0
    use_fallback   = False
    odoo_model_label = 'purchase.request'

    def _fmt_date_pr(d):
        if not d or len(d) < 10:
            return '—'
        return f'{d[8:10]}/{d[5:7]}/{d[0:4]}'

    state_options = [
        ('draft', 'Brouillon'),
        ('to_approve', 'Confirmé'),
        ('approved', 'Approuvé'),
        ('rejected', 'Rejeté'),
    ]

    try:
        uid, models = get_odoo_connection()

        # ── Essai purchase.request ──────────────────────────────
        pr_exists = False
        try:
            models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'purchase.request', 'search_count', [[]], {},
            )
            pr_exists = True
        except Exception:
            pr_exists = False

        if pr_exists:
            fields_meta = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'purchase.request', 'fields_get', [], {'attributes': ['type', 'string', 'relation']},
            ) or {}
            available_fields = set(fields_meta.keys())

            date_candidates = [f for f in ('date_start', 'date_required', 'date', 'create_date') if f in available_fields]
            requester_candidates = [f for f in ('requested_by', 'requested_by_id', 'user_id', 'create_uid') if f in available_fields]
            department_candidates = [f for f in ('department_id', 'x_department_id', 'service_id', 'company_id') if f in available_fields]
            description_candidates = [f for f in ('description', 'origin', 'notes') if f in available_fields]
            amount_candidates = [f for f in ('estimated_cost', 'amount_total', 'amount_untaxed') if f in available_fields]
            line_field = next((f for f in ('line_ids', 'request_line_ids', 'order_line') if f in available_fields), None)

            date_field = date_candidates[0] if date_candidates else None

            domain = []
            if date_debut and date_field:
                domain.append((date_field, '>=', date_debut))
            if date_fin and date_field:
                domain.append((date_field, '<=', date_fin + ' 23:59:59'))
            if etat and 'state' in available_fields:
                domain.append(('state', '=', etat))

            read_fields = ['name']
            read_fields.extend(date_candidates)
            read_fields.extend(requester_candidates)
            read_fields.extend(department_candidates)
            read_fields.extend(description_candidates)
            read_fields.extend(amount_candidates)
            if 'state' in available_fields:
                read_fields.append('state')
            if line_field:
                read_fields.append(line_field)
            read_fields = sorted(set(read_fields))

            records = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'purchase.request', 'search_read',
                [domain],
                {'fields': read_fields,
                 'limit': 5000, 'order': f'{date_field} desc' if date_field else 'id desc'},
            )

            # Si la description principale est vide, on tente de la reconstruire depuis les lignes.
            line_desc_map = {}
            if line_field:
                try:
                    line_ids = []
                    for rec in records:
                        vals = rec.get(line_field) or []
                        if isinstance(vals, list):
                            line_ids.extend([int(v) for v in vals if isinstance(v, int)])
                    line_ids = sorted(set(line_ids))
                    line_model = (fields_meta.get(line_field) or {}).get('relation') or 'purchase.request.line'
                    if line_ids and line_model:
                        line_fields_meta = models.execute_kw(
                            settings.ODOO_DB, uid, settings.ODOO_PASS,
                            line_model, 'fields_get', [], {'attributes': ['type']},
                        ) or {}
                        line_available = set(line_fields_meta.keys())
                        line_read_fields = [f for f in ('name', 'description', 'product_id', 'product_qty', 'product_uom_qty') if f in line_available]
                        line_read_fields = sorted(set(line_read_fields)) if line_read_fields else ['id']
                        line_records = models.execute_kw(
                            settings.ODOO_DB, uid, settings.ODOO_PASS,
                            line_model, 'read',
                            [line_ids],
                            {'fields': line_read_fields},
                        )
                        for ln in line_records:
                            txt = ''
                            if 'description' in ln and ln.get('description'):
                                txt = str(ln.get('description')).strip()
                            elif 'name' in ln and ln.get('name'):
                                txt = str(ln.get('name')).strip()
                            if not txt and isinstance(ln.get('product_id'), list) and len(ln.get('product_id')) > 1:
                                qty = ln.get('product_qty') if ln.get('product_qty') not in (None, '') else ln.get('product_uom_qty')
                                txt = f"{ln['product_id'][1]} x {qty}" if qty not in (None, '') else str(ln['product_id'][1])
                            line_desc_map[ln.get('id')] = txt
                except Exception:
                    line_desc_map = {}

            all_rows = []
            for r in records:
                date_raw = ''
                for f in date_candidates:
                    val = r.get(f)
                    if val:
                        date_raw = str(val)[:10]
                        break

                req = None
                for f in requester_candidates:
                    val = r.get(f)
                    if val:
                        req = val
                        break

                dept = None
                for f in department_candidates:
                    val = r.get(f)
                    if val:
                        dept = val
                        break

                desc_val = None
                for f in description_candidates:
                    val = r.get(f)
                    if val:
                        desc_val = val
                        break
                if not desc_val and line_field:
                    line_ids_row = r.get(line_field) or []
                    if isinstance(line_ids_row, list):
                        line_texts = [line_desc_map.get(i) for i in line_ids_row if line_desc_map.get(i)]
                        if line_texts:
                            desc_val = ' | '.join(line_texts[:2])

                amount_val = 0
                for f in amount_candidates:
                    val = r.get(f)
                    if val not in (None, ''):
                        amount_val = val
                        break

                dept_name = dept[1] if isinstance(dept, list) and len(dept) > 1 else '—'
                req_name = req[1] if isinstance(req, list) and len(req) > 1 else str(req or '—')
                all_rows.append({
                    'name':        r.get('name') or '—',
                    'date':        _fmt_date_pr(date_raw),
                    'date_raw':    date_raw,
                    'demandeur':   req_name,
                    'departement': dept_name,
                    'description': (str(desc_val or '—'))[:140],
                    'montant':     float(amount_val or 0),
                    'etat_raw':    r.get('state') or 'draft',
                    'etat':        STATE_LABELS_PR.get(r.get('state') or 'draft', r.get('state') or '—'),
                })

            departements    = sorted({row['departement'] for row in all_rows if row['departement'] != '—'})
            demandeurs_list = sorted({row['demandeur']   for row in all_rows if row['demandeur']   != '—'})

            # Filtres côté Python
            rows = all_rows
            if departement:
                rows = [r for r in rows if departement.lower() in r['departement'].lower()]
            if demandeur:
                rows = [r for r in rows if demandeur.lower() in r['demandeur'].lower()]

        else:
            # ── Fallback purchase.order (draft/sent) ───────────
            po_exists = False
            try:
                models.execute_kw(
                    settings.ODOO_DB, uid, settings.ODOO_PASS,
                    'purchase.order', 'search_count', [[]], {},
                )
                po_exists = True
            except Exception:
                po_exists = False

            if not po_exists:
                error = "Modèle non disponible dans cette instance Odoo"
                return render(request, 'achats/demandes_achat.html', {
                    'rows': [],
                    'error': error,
                    'date_debut': date_debut,
                    'date_fin': date_fin,
                    'departement': departement,
                    'etat': etat,
                    'demandeur': demandeur,
                    'departements': [],
                    'demandeurs': [],
                    'total_demandes': 0,
                    'total_montant': 0.0,
                    'nb_attente': 0,
                    'nb_approuvees': 0,
                    'use_fallback': True,
                    'odoo_model_label': 'purchase.request / purchase.order',
                    'state_options': state_options,
                })

            use_fallback = True
            odoo_model_label = 'purchase.order (draft/sent)'
            domain = [('state', 'in', ['draft', 'sent'])]
            if date_debut:
                domain.append(('date_order', '>=', date_debut))
            if date_fin:
                domain.append(('date_order', '<=', date_fin + ' 23:59:59'))
            if etat and etat in ('draft', 'sent'):
                domain.append(('state', '=', etat))

            records = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'purchase.order', 'search_read',
                [domain],
                {'fields': ['name', 'date_order', 'partner_id', 'amount_total',
                             'state', 'user_id', 'notes'],
                 'limit': 5000, 'order': 'date_order desc'},
            )
            all_rows = []
            for r in records:
                partner   = r.get('partner_id')
                dept_name = partner[1] if isinstance(partner, list) and len(partner) > 1 else '—'
                user      = r.get('user_id')
                user_name = user[1]    if isinstance(user,    list) and len(user)    > 1 else '—'
                all_rows.append({
                    'name':        r.get('name') or '—',
                    'date':        _fmt_date_pr((r.get('date_order') or '')[:10]),
                    'date_raw':    (r.get('date_order') or '')[:10],
                    'demandeur':   user_name,
                    'departement': dept_name,
                    'description': (r.get('notes') or '—')[:120],
                    'montant':     float(r.get('amount_total') or 0),
                    'etat_raw':    r.get('state') or 'draft',
                    'etat':        STATE_LABELS_PR.get(r.get('state') or 'draft', '—'),
                })

            departements    = sorted({row['departement'] for row in all_rows if row['departement'] != '—'})
            demandeurs_list = sorted({row['demandeur']   for row in all_rows if row['demandeur']   != '—'})

            rows = all_rows
            if departement:
                rows = [r for r in rows if departement.lower() in r['departement'].lower()]
            if demandeur:
                rows = [r for r in rows if demandeur.lower() in r['demandeur'].lower()]

        # ── KPI ────────────────────────────────────────────────
        total_demandes = len(rows)
        total_montant  = round(sum(r['montant'] for r in rows), 2)
        nb_attente     = sum(1 for r in rows if r['etat_raw'] in ('draft', 'to_approve', 'sent'))
        nb_approuvees  = sum(1 for r in rows if r['etat_raw'] in ('approved', 'purchase', 'done'))
        # ── Exports ────────────────────────────────────────────
        if export == 'csv':
            out = io.StringIO()
            w = csv.writer(out, delimiter=';')
            w.writerow(['N° Demande', 'Date', 'Demandeur', 'Département', 'Description', 'Montant estimé', 'État'])
            for r in rows:
                w.writerow([r['name'], r['date'], r['demandeur'], r['departement'],
                             r['description'], r['montant'], r['etat']])
            resp = HttpResponse(out.getvalue(), content_type='text/csv; charset=utf-8-sig')
            resp['Content-Disposition'] = 'attachment; filename=demandes_achat.csv'
            return resp

        if export == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Demandes d'achat"
            header = ['N° Demande', 'Date', 'Demandeur', 'Département', 'Description', 'Montant estimé', 'État']
            ws.append(header)
            hdr_fill = PatternFill('solid', fgColor='1A2C4E')
            hdr_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for r in rows:
                ws.append([r['name'], r['date'], r['demandeur'], r['departement'],
                            r['description'], r['montant'], r['etat']])
            ws.append(['TOTAL', '', '', '', '', total_montant, ''])

            last_row = ws.max_row
            total_fill = PatternFill('solid', fgColor='1A2C4E')
            total_font = Font(color='FFFFFF', bold=True)
            for cell in ws[last_row]:
                cell.fill = total_fill
                cell.font = total_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Mise en forme des colonnes pour un rendu lisible dans Excel.
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 24
            ws.column_dimensions['E'].width = 52
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 16

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f'A1:G{last_row}'

            for row_idx in range(2, last_row):
                ws[f'F{row_idx}'].number_format = '#,##0.00'
                ws[f'F{row_idx}'].alignment = Alignment(horizontal='right')
                ws[f'B{row_idx}'].alignment = Alignment(horizontal='center')

            ws[f'F{last_row}'].number_format = '#,##0.00'
            ws[f'F{last_row}'].alignment = Alignment(horizontal='right')

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(
                bio.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            resp['Content-Disposition'] = 'attachment; filename=demandes_achat.xlsx'
            return resp

        if export == 'pdf':
            bio_pdf = io.BytesIO()
            logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_somatrin.png'
            pw, ph = landscape(A4)

            class _PagedCanvasDA(rl_canvas.Canvas):
                def __init__(self, *args, **kw):
                    super().__init__(*args, **kw)
                    self._saved_states = []
                def showPage(self):
                    self._saved_states.append(dict(self.__dict__))
                    self._startPage()
                def save(self):
                    total = len(self._saved_states)
                    for state in self._saved_states:
                        self.__dict__.update(state)
                        self.setFont('Helvetica', 7.5)
                        self.setFillColor(colors.HexColor('#6B7280'))
                        self.drawRightString(pw - 15 * mm, 10 * mm,
                                             f'Page {self._pageNumber} / {total}')
                        super().showPage()
                    super().save()

            def _header_da(c, d):
                c.saveState()
                logo_x = 15 * mm
                logo_y = ph - 22 * mm
                logo_w = 32 * mm
                logo_h = 11 * mm
                if logo_path.is_file():
                    c.drawImage(str(logo_path), logo_x, logo_y,
                                width=logo_w, height=logo_h,
                                preserveAspectRatio=True, mask='auto')
                # Alignement vertical sur le centre du logo
                header_y = logo_y + (logo_h / 2.0) - (3 * mm)
                c.setFont('Helvetica-Bold', 11)
                c.setFillColor(colors.HexColor('#1A2C4E'))
                c.drawString(logo_x + logo_w + (6 * mm), header_y, "Demandes d'achat — SOMATRIN")
                c.setFont('Helvetica', 8)
                c.setFillColor(colors.HexColor('#6B7280'))
                c.drawRightString(pw - 15 * mm, header_y, 'Document Confidentiel')
                c.setStrokeColor(colors.HexColor('#E87722'))
                c.setLineWidth(1.2)
                c.line(15 * mm, ph - 26 * mm, pw - 15 * mm, ph - 26 * mm)
                c.restoreState()

            wrap_da = ParagraphStyle('wda', fontSize=7, leading=9)
            navy_da = colors.HexColor('#1A2C4E')
            lgray   = colors.HexColor('#F8FAFC')

            pdf_data = [['N° Demande', 'Date', 'Demandeur', 'Département',
                          'Description', 'Montant estimé', 'État']]
            for r in rows:
                pdf_data.append([
                    r['name'], r['date'], r['demandeur'], r['departement'],
                    Paragraph(str(r['description'])[:80], wrap_da),
                    f"{r['montant']:,.2f}", r['etat'],
                ])
            pdf_data.append(['TOTAL', '', '', '', '',
                              f"{total_montant:,.2f}", f"{total_demandes} demande(s)"])

            cw_da = [35 * mm, 22 * mm, 40 * mm, 40 * mm, 68 * mm, 30 * mm, 22 * mm]
            tbl_da = Table(pdf_data, colWidths=cw_da, repeatRows=1)
            tbl_da.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  navy_da),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),  8),
                ('ALIGN',         (0, 0),  (-1, 0),  'CENTER'),
                ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
                ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
                ('FONTSIZE',      (0, 1),  (-1, -2), 7.5),
                ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, lgray]),
                ('ALIGN',         (5, 1),  (5, -2),  'RIGHT'),
                ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor('#E5E7EB')),
                ('BACKGROUND',    (0, -1), (-1, -1), navy_da),
                ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
                ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, -1), (-1, -1), 8),
                ('ALIGN',         (5, -1), (5, -1),  'RIGHT'),
            ]))

            doc_da = SimpleDocTemplate(
                bio_pdf, pagesize=landscape(A4),
                leftMargin=15 * mm, rightMargin=15 * mm,
                topMargin=32 * mm, bottomMargin=22 * mm,
            )
            doc_da.build([tbl_da], onFirstPage=_header_da, onLaterPages=_header_da,
                         canvasmaker=_PagedCanvasDA)
            bio_pdf.seek(0)
            resp = HttpResponse(bio_pdf.getvalue(), content_type='application/pdf')
            resp['Content-Disposition'] = "attachment; filename=demandes_achat.pdf"
            return resp

    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return render(request, 'achats/demandes_achat.html', {
        'rows':             rows,
        'error':            error,
        'date_debut':       date_debut,
        'date_fin':         date_fin,
        'departement':      departement,
        'etat':             etat,
        'demandeur':        demandeur,
        'departements':     departements,
        'demandeurs':       demandeurs_list,
        'total_demandes':   total_demandes,
        'total_montant':    total_montant,
        'nb_attente':       nb_attente,
        'nb_approuvees':    nb_approuvees,
        'use_fallback':     use_fallback,
        'odoo_model_label': odoo_model_label,
        'state_options':    state_options,
    })


@login_required
def achats_rfq(request):
    date_debut  = request.GET.get('date_debut',   '').strip()
    date_fin    = request.GET.get('date_fin',     '').strip()
    fournisseur = request.GET.get('fournisseur',  '').strip()
    etat        = request.GET.get('etat',         '').strip()
    responsable = request.GET.get('responsable',  '').strip()
    export      = request.GET.get('export',       '').strip().lower()

    STATE_LABELS_RFQ = {
        'draft':    'Brouillon',
        'sent':     'Envoyé',
        'purchase': 'Bon de commande',
        'cancel':   'Annulé',
    }

    rows = []
    error = None
    fournisseurs = []
    responsables = []
    total_dp     = 0
    total_montant = 0.0
    nb_attente    = 0
    nb_expirees   = 0

    def _fmt_date_rfq(d):
        if not d or len(d) < 10:
            return '—'
        return f'{d[8:10]}/{d[5:7]}/{d[0:4]}'

    try:
        from datetime import datetime as _dt, timedelta as _td
        uid, models = get_odoo_connection()

        po_exists = False
        try:
            models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'purchase.order', 'search_count', [[]], {},
            )
            po_exists = True
        except Exception:
            po_exists = False

        if not po_exists:
            return render(request, 'achats/demandes_prix.html', {
                'rows': [],
                'error': "Modèle non disponible dans cette instance Odoo",
                'date_debut': date_debut,
                'date_fin': date_fin,
                'fournisseur': fournisseur,
                'etat': etat,
                'responsable': responsable,
                'fournisseurs': [],
                'responsables': [],
                'total_dp': 0,
                'total_montant': 0.0,
                'nb_attente': 0,
                'nb_expirees': 0,
            })

        domain = []
        if date_debut:
            domain.append(('date_order', '>=', date_debut))
        if date_fin:
            domain.append(('date_order', '<=', date_fin + ' 23:59:59'))
        if etat:
            domain.append(('state', '=', etat))

        # Champs disponibles (pour utiliser une vraie date d'expiration si présente)
        po_fields = {}
        try:
            po_fields = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'purchase.order', 'fields_get', [], {'attributes': ['type']},
            ) or {}
        except Exception:
            po_fields = {}
        po_available = set(po_fields.keys())

        read_fields = ['name', 'date_order', 'partner_id', 'amount_total', 'state', 'currency_id', 'user_id']
        # Certaines instances Odoo exposent une date de validité/expiration des RFQ
        if 'validity_date' in po_available:
            read_fields.append('validity_date')

        records = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'purchase.order', 'search_read',
            [domain],
            {'fields': read_fields,
             'limit': 5000, 'order': 'date_order desc'},
        )

        today = date.today()
        all_rows = []
        for r in records:
            partner    = r.get('partner_id')
            fournisseur_name = partner[1] if isinstance(partner, list) and len(partner) > 1 else '—'
            user       = r.get('user_id')
            user_name  = user[1]    if isinstance(user,    list) and len(user)    > 1 else '—'
            currency   = r.get('currency_id')
            devise     = currency[1] if isinstance(currency, list) and len(currency) > 1 else 'MAD'

            raw_date = (r.get('date_order') or '')[:10]
            validity_raw = (r.get('validity_date') or '')[:10] if 'validity_date' in po_available else ''
            expire_raw = ''
            is_expired = False
            if raw_date and len(raw_date) == 10:
                try:
                    order_date = _dt.strptime(raw_date, '%Y-%m-%d').date()
                    if validity_raw and len(validity_raw) == 10:
                        exp_date = _dt.strptime(validity_raw, '%Y-%m-%d').date()
                        expire_raw = validity_raw
                    else:
                        exp_date = order_date + _td(days=15)
                        expire_raw = exp_date.isoformat()
                    if r.get('state') == 'sent':
                        is_expired = today > exp_date
                except Exception:
                    pass

            all_rows.append({
                'name':        r.get('name') or '—',
                'date':        raw_date,
                'expire_raw':  expire_raw,
                'has_validity_date': bool(validity_raw),
                'fournisseur': fournisseur_name,
                'responsable': user_name,
                'montant':     float(r.get('amount_total') or 0),
                'devise':      devise,
                'etat_raw':    r.get('state') or 'draft',
                'etat':        STATE_LABELS_RFQ.get(r.get('state') or 'draft', '—'),
                'is_expired':  is_expired,
            })

        fournisseurs = sorted({row['fournisseur'] for row in all_rows if row['fournisseur'] != '—'})
        responsables = sorted({row['responsable'] for row in all_rows if row['responsable'] != '—'})

        # Filtres côté Python
        rows = all_rows
        if fournisseur:
            rows = [r for r in rows if fournisseur.lower() in r['fournisseur'].lower()]
        if responsable:
            rows = [r for r in rows if responsable.lower() in r['responsable'].lower()]

        # ── KPI ────────────────────────────────────────────────
        total_dp      = len(rows)
        total_montant = round(sum(r['montant'] for r in rows), 2)
        # "En attente réponse fournisseur" = RFQ envoyées au fournisseur.
        nb_attente    = sum(1 for r in rows if r['etat_raw'] == 'sent')
        nb_expirees   = sum(1 for r in rows if r.get('is_expired'))

        # Formatage dates après calcul is_expired
        for r in rows:
            r['date'] = _fmt_date_rfq(r['date'])
            r['expire'] = _fmt_date_rfq(r.get('expire_raw') or '')

        # ── Exports ────────────────────────────────────────────
        if export == 'csv':
            out = io.StringIO()
            w = csv.writer(out, delimiter=';')
            w.writerow(['N° Demande', 'Date', 'Fournisseur', 'Responsable', 'Montant HT', 'Devise', 'État'])
            for r in rows:
                w.writerow([r['name'], r['date'], r['fournisseur'], r['responsable'],
                             r['montant'], r['devise'], r['etat']])
            resp = HttpResponse(out.getvalue(), content_type='text/csv; charset=utf-8-sig')
            resp['Content-Disposition'] = 'attachment; filename=demandes_prix.csv'
            return resp

        if export == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Demandes de prix'
            header = ['N° Demande', 'Date', 'Fournisseur', 'Responsable', 'Montant HT', 'Devise', 'État']
            ws.append(header)
            hdr_fill = PatternFill('solid', fgColor='1A2C4E')
            hdr_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for r in rows:
                ws.append([r['name'], r['date'], r['fournisseur'], r['responsable'],
                            r['montant'], r['devise'], r['etat']])
            ws.append(['TOTAL', '', '', '', total_montant, '', ''])

            last_row = ws.max_row
            total_fill = PatternFill('solid', fgColor='1A2C4E')
            total_font = Font(color='FFFFFF', bold=True)
            for cell in ws[last_row]:
                cell.fill = total_fill
                cell.font = total_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 24
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 14
            ws.column_dimensions['G'].width = 14

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f'A1:G{last_row}'

            for row_idx in range(2, last_row):
                ws[f'E{row_idx}'].number_format = '#,##0.00'
                ws[f'E{row_idx}'].alignment = Alignment(horizontal='right')
                ws[f'B{row_idx}'].alignment = Alignment(horizontal='center')

            ws[f'E{last_row}'].number_format = '#,##0.00'
            ws[f'E{last_row}'].alignment = Alignment(horizontal='right')

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(
                bio.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            resp['Content-Disposition'] = 'attachment; filename=demandes_prix.xlsx'
            return resp

        if export == 'pdf':
            bio_pdf = io.BytesIO()
            logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_somatrin.png'
            pw, ph = landscape(A4)

            class _PagedCanvasDP(rl_canvas.Canvas):
                def __init__(self, *args, **kw):
                    super().__init__(*args, **kw)
                    self._saved_states = []
                def showPage(self):
                    self._saved_states.append(dict(self.__dict__))
                    self._startPage()
                def save(self):
                    total = len(self._saved_states)
                    for state in self._saved_states:
                        self.__dict__.update(state)
                        self.setFont('Helvetica', 7.5)
                        self.setFillColor(colors.HexColor('#6B7280'))
                        self.drawRightString(pw - 15 * mm, 10 * mm,
                                             f'Page {self._pageNumber} / {total}')
                        super().showPage()
                    super().save()

            def _header_dp(c, d):
                c.saveState()
                logo_x = 15 * mm
                logo_y = ph - 22 * mm
                logo_w = 32 * mm
                logo_h = 11 * mm
                if logo_path.is_file():
                    c.drawImage(str(logo_path), logo_x, logo_y,
                                width=logo_w, height=logo_h,
                                preserveAspectRatio=True, mask='auto')
                header_y = logo_y + (logo_h / 2.0) - (3 * mm)
                c.setFont('Helvetica-Bold', 11)
                c.setFillColor(colors.HexColor('#1A2C4E'))
                c.drawString(logo_x + logo_w + (6 * mm), header_y, 'Demandes de prix — SOMATRIN')
                c.setFont('Helvetica', 8)
                c.setFillColor(colors.HexColor('#6B7280'))
                c.drawRightString(pw - 15 * mm, header_y, 'Document Confidentiel')
                c.setStrokeColor(colors.HexColor('#E87722'))
                c.setLineWidth(1.2)
                c.line(15 * mm, ph - 26 * mm, pw - 15 * mm, ph - 26 * mm)
                c.restoreState()

            navy_dp = colors.HexColor('#1A2C4E')
            lgray_dp = colors.HexColor('#F8FAFC')

            pdf_data = [['N° Demande', 'Date', 'Fournisseur',
                          'Responsable', 'Montant HT', 'Devise', 'État']]
            for r in rows:
                pdf_data.append([
                    r['name'], r['date'], r['fournisseur'],
                    r['responsable'], f"{r['montant']:,.2f}", r['devise'], r['etat'],
                ])
            pdf_data.append(['TOTAL', '', '', '',
                              f"{total_montant:,.2f}", '', f"{total_dp} demande(s)"])

            cw_dp = [38 * mm, 24 * mm, 55 * mm, 45 * mm, 32 * mm, 18 * mm, 22 * mm]
            tbl_dp = Table(pdf_data, colWidths=cw_dp, repeatRows=1)
            tbl_dp.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  navy_dp),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),  8),
                ('ALIGN',         (0, 0),  (-1, 0),  'CENTER'),
                ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
                ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
                ('FONTSIZE',      (0, 1),  (-1, -2), 7.5),
                ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, lgray_dp]),
                ('ALIGN',         (4, 1),  (4, -2),  'RIGHT'),
                ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor('#E5E7EB')),
                ('BACKGROUND',    (0, -1), (-1, -1), navy_dp),
                ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
                ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, -1), (-1, -1), 8),
                ('ALIGN',         (4, -1), (4, -1),  'RIGHT'),
            ]))

            doc_dp = SimpleDocTemplate(
                bio_pdf, pagesize=landscape(A4),
                leftMargin=15 * mm, rightMargin=15 * mm,
                topMargin=32 * mm, bottomMargin=22 * mm,
            )
            doc_dp.build([tbl_dp], onFirstPage=_header_dp, onLaterPages=_header_dp,
                         canvasmaker=_PagedCanvasDP)
            bio_pdf.seek(0)
            resp = HttpResponse(bio_pdf.getvalue(), content_type='application/pdf')
            resp['Content-Disposition'] = "attachment; filename=demandes_prix.pdf"
            return resp

    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return render(request, 'achats/demandes_prix.html', {
        'rows':         rows,
        'error':        error,
        'date_debut':   date_debut,
        'date_fin':     date_fin,
        'fournisseur':  fournisseur,
        'etat':         etat,
        'responsable':  responsable,
        'fournisseurs': fournisseurs,
        'responsables': responsables,
        'total_dp':     total_dp,
        'total_montant': total_montant,
        'nb_attente':   nb_attente,
        'nb_expirees':  nb_expirees,
    })


@login_required
def achats_purchase_orders(request):
    date_debut  = request.GET.get('date_debut',   '').strip()
    date_fin    = request.GET.get('date_fin',     '').strip()
    fournisseur = request.GET.get('fournisseur',  '').strip()
    etat        = request.GET.get('etat',         '').strip()
    responsable = request.GET.get('responsable',  '').strip()
    societe     = request.GET.get('societe',      '').strip()
    export      = request.GET.get('export',       '').strip().lower()

    STATE_LABELS_PO = {
        'draft':    'Brouillon',
        'sent':     'Envoyé',
        'purchase': 'Confirmé',
        'done':     'Terminé',
        'cancel':   'Annulé',
    }

    rows        = []
    error       = None
    fournisseurs = []
    responsables = []
    societes    = []
    total_bons  = 0
    total_ht    = 0.0
    nb_confirmes = 0
    nb_attente  = 0

    def _fmt_date_po(d):
        if not d or len(d) < 10:
            return '—'
        return f'{d[8:10]}/{d[5:7]}/{d[0:4]}'

    try:
        uid, models = get_odoo_connection()

        po_exists = False
        try:
            models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'purchase.order', 'search_count', [[]], {},
            )
            po_exists = True
        except Exception:
            po_exists = False

        if not po_exists:
            return render(request, 'achats/bons_commande.html', {
                'rows': [], 'error': "Modèle purchase.order non disponible dans cette instance Odoo",
                'date_debut': date_debut, 'date_fin': date_fin,
                'fournisseur': fournisseur, 'etat': etat,
                'responsable': responsable, 'societe': societe,
                'fournisseurs': [], 'responsables': [], 'societes': [],
                'total_bons': 0, 'total_ht': 0.0, 'nb_confirmes': 0, 'nb_attente': 0,
            })

        domain = []
        if date_debut:
            domain.append(('date_order', '>=', date_debut))
        if date_fin:
            domain.append(('date_order', '<=', date_fin + ' 23:59:59'))
        if etat:
            domain.append(('state', '=', etat))

        records = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'purchase.order', 'search_read',
            [domain],
            {'fields': ['name', 'date_order', 'partner_id', 'user_id',
                        'amount_untaxed', 'amount_tax', 'amount_total',
                        'state', 'currency_id', 'date_planned', 'company_id'],
             'limit': 5000, 'order': 'date_order desc'},
        )

        all_rows = []
        for r in records:
            partner      = r.get('partner_id')
            fourn_name   = partner[1]  if isinstance(partner,  list) and len(partner)  > 1 else '—'
            user         = r.get('user_id')
            user_name    = user[1]     if isinstance(user,     list) and len(user)     > 1 else '—'
            currency     = r.get('currency_id')
            devise       = currency[1] if isinstance(currency, list) and len(currency) > 1 else 'MAD'
            company      = r.get('company_id')
            company_name = company[1]  if isinstance(company,  list) and len(company)  > 1 else '—'
            all_rows.append({
                'name':         r.get('name') or '—',
                'date':         (r.get('date_order') or '')[:10],
                'date_planned': (r.get('date_planned') or '')[:10],
                'fournisseur':  fourn_name,
                'responsable':  user_name,
                'societe':      company_name,
                'montant_ht':   float(r.get('amount_untaxed') or 0),
                'montant_tax':  float(r.get('amount_tax')      or 0),
                'montant_ttc':  float(r.get('amount_total')    or 0),
                'devise':       devise,
                'etat_raw':     r.get('state') or 'draft',
                'etat':         STATE_LABELS_PO.get(r.get('state') or 'draft', '—'),
            })

        fournisseurs = sorted({row['fournisseur'] for row in all_rows if row['fournisseur'] != '—'})
        responsables = sorted({row['responsable'] for row in all_rows if row['responsable'] != '—'})
        societes     = sorted({row['societe']     for row in all_rows if row['societe']     != '—'})

        rows = all_rows
        if fournisseur:
            rows = [r for r in rows if fournisseur.lower() in r['fournisseur'].lower()]
        if responsable:
            rows = [r for r in rows if responsable.lower() in r['responsable'].lower()]
        if societe:
            rows = [r for r in rows if societe == r['societe']]

        total_bons   = len(rows)
        total_ht     = round(sum(r['montant_ht']  for r in rows), 2)
        nb_confirmes = sum(1 for r in rows if r['etat_raw'] in ('purchase', 'done'))
        nb_attente   = sum(1 for r in rows if r['etat_raw'] in ('draft', 'sent'))

        for r in rows:
            r['date']         = _fmt_date_po(r['date'])
            r['date_planned'] = _fmt_date_po(r['date_planned'])

        if export == 'csv':
            out = io.StringIO()
            w = csv.writer(out, delimiter=';')
            w.writerow(['N° Commande', 'Date', 'Date prévue', 'Fournisseur', 'Responsable',
                        'Société', 'Montant HT', 'Taxes', 'Montant TTC', 'Devise', 'État'])
            for r in rows:
                w.writerow([r['name'], r['date'], r['date_planned'], r['fournisseur'],
                             r['responsable'], r['societe'], r['montant_ht'],
                            r['montant_tax'], r['montant_ttc'], r['devise'], r['etat']])
            resp = HttpResponse(out.getvalue(), content_type='text/csv; charset=utf-8-sig')
            resp['Content-Disposition'] = 'attachment; filename=bons_commande.csv'
            return resp

        if export == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Bons de commande'
            header = ['N° Commande', 'Date', 'Date prévue', 'Fournisseur', 'Responsable',
                      'Société', 'Montant HT', 'Taxes', 'Montant TTC', 'Devise', 'État']
            ws.append(header)
            hdr_fill = PatternFill('solid', fgColor='1A2C4E')
            hdr_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for r in rows:
                ws.append([r['name'], r['date'], r['date_planned'], r['fournisseur'],
                            r['responsable'], r['societe'], r['montant_ht'],
                            r['montant_tax'], r['montant_ttc'], r['devise'], r['etat']])
            ws.append(['TOTAL', '', '', '', '', '', total_ht, '', '', '', ''])

            last_row = ws.max_row
            for cell in ws[last_row]:
                cell.fill = PatternFill('solid', fgColor='1A2C4E')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            col_widths = {'A': 18, 'B': 14, 'C': 14, 'D': 30, 'E': 24,
                          'F': 20, 'G': 16, 'H': 16, 'I': 16, 'J': 10, 'K': 14}
            for col, w_val in col_widths.items():
                ws.column_dimensions[col].width = w_val

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f'A1:K{last_row}'

            for row_idx in range(2, last_row):
                for col in ('G', 'H', 'I'):
                    ws[f'{col}{row_idx}'].number_format = '#,##0.00'
                    ws[f'{col}{row_idx}'].alignment = Alignment(horizontal='right')
                ws[f'B{row_idx}'].alignment = Alignment(horizontal='center')
                ws[f'C{row_idx}'].alignment = Alignment(horizontal='center')
            for col in ('G', 'H', 'I'):
                ws[f'{col}{last_row}'].number_format = '#,##0.00'
                ws[f'{col}{last_row}'].alignment = Alignment(horizontal='right')

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(
                bio.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            resp['Content-Disposition'] = 'attachment; filename=bons_commande.xlsx'
            return resp

        if export == 'pdf':
            bio_pdf = io.BytesIO()
            logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_somatrin.png'
            pw, ph = landscape(A4)

            class _PagedCanvasPO(rl_canvas.Canvas):
                def __init__(self, *args, **kw):
                    super().__init__(*args, **kw)
                    self._saved_states = []
                def showPage(self):
                    self._saved_states.append(dict(self.__dict__))
                    self._startPage()
                def save(self):
                    total = len(self._saved_states)
                    for state in self._saved_states:
                        self.__dict__.update(state)
                        self.setFont('Helvetica', 7.5)
                        self.setFillColor(colors.HexColor('#6B7280'))
                        self.drawRightString(pw - 15 * mm, 10 * mm, f'Page {self._pageNumber} / {total}')
                        super().showPage()
                    super().save()

            def _header_po(c, d):
                c.saveState()
                logo_x = 15 * mm
                logo_y = ph - 22 * mm
                logo_w = 32 * mm
                logo_h = 11 * mm
                if logo_path.is_file():
                    c.drawImage(str(logo_path), logo_x, logo_y,
                                width=logo_w, height=logo_h,
                                preserveAspectRatio=True, mask='auto')
                header_y = logo_y + (logo_h / 2.0) - (3 * mm)
                c.setFont('Helvetica-Bold', 11)
                c.setFillColor(colors.HexColor('#1A2C4E'))
                c.drawString(logo_x + logo_w + (6 * mm), header_y, 'Bons de commande — SOMATRIN')
                c.setFont('Helvetica', 8)
                c.setFillColor(colors.HexColor('#6B7280'))
                c.drawRightString(pw - 15 * mm, header_y, 'Document Confidentiel')
                c.setStrokeColor(colors.HexColor('#E87722'))
                c.setLineWidth(1.2)
                c.line(15 * mm, ph - 26 * mm, pw - 15 * mm, ph - 26 * mm)
                c.restoreState()

            navy_po = colors.HexColor('#1A2C4E')
            lgray_po = colors.HexColor('#F8FAFC')
            pdf_data = [['N° Commande', 'Date', 'Date prévue', 'Fournisseur', 'Responsable',
                         'Société', 'Montant HT', 'Taxes', 'Montant TTC', 'État']]
            for r in rows:
                pdf_data.append([
                    r['name'], r['date'], r['date_planned'], r['fournisseur'], r['responsable'],
                    r['societe'], f"{r['montant_ht']:,.2f}", f"{r['montant_tax']:,.2f}",
                    f"{r['montant_ttc']:,.2f}", r['etat'],
                ])
            pdf_data.append(['TOTAL', '', '', '', '', '', f"{total_ht:,.2f}", '', '', f"{total_bons} bon(s)"])

            cw_po = [33 * mm, 20 * mm, 24 * mm, 42 * mm, 32 * mm, 30 * mm, 24 * mm, 20 * mm, 24 * mm, 20 * mm]
            tbl_po = Table(pdf_data, colWidths=cw_po, repeatRows=1)
            tbl_po.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  navy_po),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),  8),
                ('ALIGN',         (0, 0),  (-1, 0),  'CENTER'),
                ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
                ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
                ('FONTSIZE',      (0, 1),  (-1, -2), 7.2),
                ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, lgray_po]),
                ('ALIGN',         (6, 1),  (8, -1),  'RIGHT'),
                ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor('#E5E7EB')),
                ('BACKGROUND',    (0, -1), (-1, -1), navy_po),
                ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
                ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, -1), (-1, -1), 8),
            ]))

            doc_po = SimpleDocTemplate(
                bio_pdf, pagesize=landscape(A4),
                leftMargin=15 * mm, rightMargin=15 * mm,
                topMargin=32 * mm, bottomMargin=22 * mm,
            )
            doc_po.build([tbl_po], onFirstPage=_header_po, onLaterPages=_header_po, canvasmaker=_PagedCanvasPO)
            bio_pdf.seek(0)
            resp = HttpResponse(bio_pdf.getvalue(), content_type='application/pdf')
            resp['Content-Disposition'] = "attachment; filename=bons_commande.pdf"
            return resp

    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return render(request, 'achats/bons_commande.html', {
        'rows':         rows,
        'error':        error,
        'date_debut':   date_debut,
        'date_fin':     date_fin,
        'fournisseur':  fournisseur,
        'etat':         etat,
        'responsable':  responsable,
        'societe':      societe,
        'fournisseurs': fournisseurs,
        'responsables': responsables,
        'societes':     societes,
        'total_bons':   total_bons,
        'total_ht':     total_ht,
        'nb_confirmes': nb_confirmes,
        'nb_attente':   nb_attente,
    })


@login_required
def achats_delivery_tracking(request):
    date_debut = request.GET.get('date_debut', '').strip()
    date_fin = request.GET.get('date_fin', '').strip()
    fournisseur = request.GET.get('fournisseur', '').strip()
    statut = request.GET.get('statut', '').strip()
    responsable = request.GET.get('responsable', '').strip()
    export = request.GET.get('export', '').strip().lower()

    rows = []
    error = None
    fournisseurs = []
    responsables = []
    total_livraisons = 0
    total_ttc = 0.0
    nb_recues = 0
    nb_retard = 0
    nb_en_cours = 0

    def _fmt_date(v):
        if not v or len(v) < 10:
            return '—'
        return f'{v[8:10]}/{v[5:7]}/{v[0:4]}'

    try:
        from datetime import datetime as _dt

        uid, models = get_odoo_connection()
        po_exists = False
        try:
            models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'purchase.order', 'search_count', [[]], {},
            )
            po_exists = True
        except Exception:
            po_exists = False

        if not po_exists:
            return render(request, 'achats/suivi_livraisons.html', {
                'rows': [],
                'error': "Modèle purchase.order non disponible dans cette instance Odoo",
                'date_debut': date_debut,
                'date_fin': date_fin,
                'fournisseur': fournisseur,
                'statut': statut,
                'responsable': responsable,
                'fournisseurs': [],
                'responsables': [],
                'total_livraisons': 0,
                'total_ttc': 0.0,
                'nb_recues': 0,
                'nb_retard': 0,
                'nb_en_cours': 0,
                'stats_statut': {'en_retard': 0, 'en_cours': 0, 'recue': 0},
            })

        domain = []
        if date_debut:
            domain.append(('date_order', '>=', date_debut))
        if date_fin:
            domain.append(('date_order', '<=', date_fin + ' 23:59:59'))

        records = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'purchase.order', 'search_read',
            [domain],
            {'fields': ['name', 'date_order', 'date_planned', 'partner_id', 'user_id', 'amount_total', 'state', 'currency_id'],
             'limit': 5000, 'order': 'date_order desc'},
        )

        today = date.today()
        for r in records:
            partner = r.get('partner_id')
            fourn_name = partner[1] if isinstance(partner, list) and len(partner) > 1 else '—'
            user = r.get('user_id')
            user_name = user[1] if isinstance(user, list) and len(user) > 1 else '—'
            cur = r.get('currency_id')
            devise = cur[1] if isinstance(cur, list) and len(cur) > 1 else 'MAD'
            state_raw = r.get('state') or 'draft'

            raw_order = (r.get('date_order') or '')[:10]
            raw_planned = (r.get('date_planned') or '')[:10]
            delay_days = None
            if raw_order and raw_planned and len(raw_order) == 10 and len(raw_planned) == 10:
                try:
                    d_order = _dt.strptime(raw_order, '%Y-%m-%d').date()
                    d_plan = _dt.strptime(raw_planned, '%Y-%m-%d').date()
                    delay_days = (d_plan - d_order).days
                except Exception:
                    delay_days = None

            is_received = state_raw in ('done',)
            is_cancel = state_raw in ('cancel',)
            is_late = False
            if raw_planned and len(raw_planned) == 10 and not is_received and not is_cancel:
                try:
                    d_plan = _dt.strptime(raw_planned, '%Y-%m-%d').date()
                    is_late = d_plan < today
                except Exception:
                    is_late = False

            if is_cancel:
                statut_livraison = 'Annulée'
            elif is_received:
                statut_livraison = 'Reçue'
            elif is_late:
                statut_livraison = 'En retard'
            else:
                statut_livraison = 'En cours'

            row = {
                'name': r.get('name') or '—',
                'date_order': _fmt_date(raw_order),
                'date_planned': _fmt_date(raw_planned),
                'fournisseur': fourn_name,
                'responsable': user_name,
                'montant_ttc': float(r.get('amount_total') or 0),
                'devise': devise,
                'state_raw': state_raw,
                'statut_livraison': statut_livraison,
                'delay_days': delay_days,
            }
            rows.append(row)

        fournisseurs = sorted({r['fournisseur'] for r in rows if r['fournisseur'] != '—'})
        responsables = sorted({r['responsable'] for r in rows if r['responsable'] != '—'})

        if fournisseur:
            rows = [r for r in rows if fournisseur.lower() in r['fournisseur'].lower()]
        if responsable:
            rows = [r for r in rows if responsable.lower() in r['responsable'].lower()]
        if statut:
            rows = [r for r in rows if r['statut_livraison'] == statut]

        total_livraisons = len(rows)
        total_ttc = round(sum(r['montant_ttc'] for r in rows), 2)
        nb_recues = sum(1 for r in rows if r['statut_livraison'] == 'Reçue')
        nb_retard = sum(1 for r in rows if r['statut_livraison'] == 'En retard')
        nb_en_cours = sum(1 for r in rows if r['statut_livraison'] == 'En cours')

        if export == 'csv':
            out = io.StringIO()
            w = csv.writer(out, delimiter=';')
            w.writerow(['N° Commande', 'Date commande', 'Date prévue', 'Fournisseur', 'Responsable',
                        'Montant TTC', 'Devise', 'Statut livraison', 'Délai (j)'])
            for r in rows:
                w.writerow([r['name'], r['date_order'], r['date_planned'], r['fournisseur'], r['responsable'],
                            r['montant_ttc'], r['devise'], r['statut_livraison'], r['delay_days'] if r['delay_days'] is not None else '—'])
            resp = HttpResponse(out.getvalue(), content_type='text/csv; charset=utf-8-sig')
            resp['Content-Disposition'] = 'attachment; filename=suivi_livraisons.csv'
            return resp

        if export == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Suivi livraisons'
            ws.append(['N° Commande', 'Date commande', 'Date prévue', 'Fournisseur', 'Responsable',
                       'Montant TTC', 'Devise', 'Statut livraison', 'Délai (j)'])
            hdr_fill = PatternFill('solid', fgColor='1A2C4E')
            hdr_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for r in rows:
                ws.append([r['name'], r['date_order'], r['date_planned'], r['fournisseur'], r['responsable'],
                          r['montant_ttc'], r['devise'], r['statut_livraison'], r['delay_days'] if r['delay_days'] is not None else ''])
            ws.append(['TOTAL', '', '', '', '', total_ttc, '', '', ''])
            last_row = ws.max_row
            for cell in ws[last_row]:
                cell.fill = PatternFill('solid', fgColor='1A2C4E')
                cell.font = Font(color='FFFFFF', bold=True)
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 28
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 18
            ws.column_dimensions['I'].width = 11
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(
                bio.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            resp['Content-Disposition'] = 'attachment; filename=suivi_livraisons.xlsx'
            return resp

        if export == 'pdf':
            bio_pdf = io.BytesIO()
            logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_somatrin.png'
            pw, ph = landscape(A4)

            class _PagedCanvasSL(rl_canvas.Canvas):
                def __init__(self, *args, **kw):
                    super().__init__(*args, **kw)
                    self._saved_states = []
                def showPage(self):
                    self._saved_states.append(dict(self.__dict__))
                    self._startPage()
                def save(self):
                    total = len(self._saved_states)
                    for state in self._saved_states:
                        self.__dict__.update(state)
                        self.setFont('Helvetica', 7.5)
                        self.setFillColor(colors.HexColor('#6B7280'))
                        self.drawRightString(pw - 15 * mm, 10 * mm,
                                             f'Page {self._pageNumber} / {total}')
                        super().showPage()
                    super().save()

            from datetime import datetime as _dtpdf
            gen_date = _dtpdf.now().strftime('%d/%m/%Y %H:%M')

            def _header_sl(c, d):
                c.saveState()
                if logo_path.is_file():
                    c.drawImage(str(logo_path), 15 * mm, ph - 22 * mm,
                                width=32 * mm, height=11 * mm,
                                preserveAspectRatio=True, mask='auto')
                c.setFont('Helvetica-Bold', 11)
                c.setFillColor(colors.HexColor('#1A2C4E'))
                c.drawString(52 * mm, ph - 13 * mm, 'SUIVI LIVRAISONS — SOMATRIN')
                c.setFont('Helvetica', 8)
                c.setFillColor(colors.HexColor('#6B7280'))
                c.drawString(52 * mm, ph - 20 * mm,
                             'Contrôle des délais et statut des réceptions fournisseurs')
                c.drawRightString(pw - 15 * mm, ph - 13 * mm, f'Généré le {gen_date}')
                c.drawRightString(pw - 15 * mm, ph - 20 * mm, 'Document Confidentiel')
                c.setStrokeColor(colors.HexColor('#E87722'))
                c.setLineWidth(1.2)
                c.line(15 * mm, ph - 26 * mm, pw - 15 * mm, ph - 26 * mm)
                c.restoreState()

            navy_sl = colors.HexColor('#1A2C4E')
            lgray_sl = colors.HexColor('#F8FAFC')

            pdf_data = [['N° Commande', 'Date cmd', 'Date prévue', 'Fournisseur',
                          'Responsable', 'Montant TTC', 'Devise', 'Statut', 'Délai (j)']]
            for r in rows:
                pdf_data.append([
                    r['name'], r['date_order'], r['date_planned'],
                    r['fournisseur'], r['responsable'],
                    f"{r['montant_ttc']:,.2f}", r['devise'],
                    r['statut_livraison'],
                    str(r['delay_days']) if r['delay_days'] is not None else '—',
                ])
            pdf_data.append(['TOTAL', '', '', '', '',
                              f"{total_ttc:,.2f}", '', '', ''])

            cw_sl = [30 * mm, 20 * mm, 20 * mm, 45 * mm, 35 * mm,
                     28 * mm, 14 * mm, 22 * mm, 18 * mm]
            tbl_sl = Table(pdf_data, colWidths=cw_sl, repeatRows=1)
            tbl_sl.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  navy_sl),
                ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, 0),  8),
                ('ALIGN',         (0, 0),  (-1, 0),  'CENTER'),
                ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
                ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
                ('FONTSIZE',      (0, 1),  (-1, -2), 7.5),
                ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, lgray_sl]),
                ('ALIGN',         (5, 1),  (5, -2),  'RIGHT'),
                ('ALIGN',         (8, 1),  (8, -2),  'CENTER'),
                ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor('#E5E7EB')),
                ('BACKGROUND',    (0, -1), (-1, -1), navy_sl),
                ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
                ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, -1), (-1, -1), 8),
                ('ALIGN',         (5, -1), (5, -1),  'RIGHT'),
            ]))

            doc_sl = SimpleDocTemplate(
                bio_pdf, pagesize=landscape(A4),
                leftMargin=15 * mm, rightMargin=15 * mm,
                topMargin=32 * mm, bottomMargin=22 * mm,
            )
            doc_sl.build([tbl_sl], onFirstPage=_header_sl, onLaterPages=_header_sl,
                         canvasmaker=_PagedCanvasSL)
            bio_pdf.seek(0)
            resp = HttpResponse(bio_pdf.getvalue(), content_type='application/pdf')
            resp['Content-Disposition'] = "attachment; filename=suivi_livraisons.pdf"
            return resp

    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    stats_statut = {
        'en_retard': nb_retard,
        'en_cours':  nb_en_cours,
        'recue':     nb_recues,
    }

    return render(request, 'achats/suivi_livraisons.html', {
        'rows': rows,
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'fournisseur': fournisseur,
        'statut': statut,
        'responsable': responsable,
        'fournisseurs': fournisseurs,
        'responsables': responsables,
        'total_livraisons': total_livraisons,
        'total_ttc': total_ttc,
        'nb_recues': nb_recues,
        'nb_retard': nb_retard,
        'nb_en_cours': nb_en_cours,
        'stats_statut': stats_statut,
    })


@login_required
def achats_suppliers(request):
    export = request.GET.get('export', '')
    nom_filter = request.GET.get('nom', '').strip()
    ville_filter = request.GET.get('ville', '').strip()
    pays_filter = request.GET.get('pays', '').strip()
    statut_filter = request.GET.get('statut', '').strip()

    _error_ctx = {
        'rows': [], 'total_fournisseurs': 0,
        'fournisseurs_actifs': 0, 'total_commandes': 0,
        'top_fournisseur': '—', 'nb_avec_email': 0, 'nb_avec_phone': 0,
        'stats_ville_json': '{}',
        'nom': nom_filter, 'ville': ville_filter, 'pays': pays_filter, 'statut': statut_filter,
        'villes': [], 'pays_list': [],
    }

    try:
        uid, models = get_odoo_connection()
    except Exception as e:
        _error_ctx['error'] = str(e)
        return render(request, 'achats/fournisseurs.html', _error_ctx)

    domain = [('is_company', '=', True), ('supplier_rank', '>', 0)]
    fields = ['id', 'name', 'phone', 'mobile', 'email', 'city', 'country_id',
              'supplier_rank', 'purchase_order_count', 'ref', 'vat', 'street', 'zip']

    try:
        records = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'res.partner', 'search_read',
            [domain], {'fields': fields, 'limit': 5000, 'order': 'name asc'}
        )
    except Exception as e:
        _error_ctx['error'] = str(e)
        return render(request, 'achats/fournisseurs.html', _error_ctx)

    # ── Dédoublonnage par (nom + ville) ──────────────────────────
    _seen = set()
    records_uniq = []
    for r in records:
        key = (
            (r.get('name') or '').strip().upper(),
            (r.get('city') or '').strip().upper(),
        )
        if key not in _seen:
            _seen.add(key)
            records_uniq.append(r)
    records = records_uniq

    # ── Traduction pays ──────────────────────────────────────────
    _PAYS_FR = {
        'Morocco': 'Maroc', 'France': 'France',
        'Algeria': 'Algérie', 'Tunisia': 'Tunisie',
        'Spain': 'Espagne', 'Belgium': 'Belgique',
        'Germany': 'Allemagne', 'Italy': 'Italie',
        'United Arab Emirates': 'Émirats Arabes Unis',
        'Saudi Arabia': 'Arabie Saoudite',
        'United States': 'États-Unis', 'China': 'Chine',
        'United Kingdom': 'Royaume-Uni', 'Netherlands': 'Pays-Bas',
        'Switzerland': 'Suisse', 'Portugal': 'Portugal',
        'Turkey': 'Turquie', 'Egypt': 'Égypte',
    }

    # ── Normalisation complète avant tout filtrage ───────────────
    def _clean(val):
        if not val or val is False:
            return ''
        return str(val).strip()

    all_rows = []
    for r in records:
        nom   = _clean(r.get('name')) or '—'
        ville = (_clean(r.get('city')) or '—').upper()
        pays_raw = r.get('country_id')
        if isinstance(pays_raw, (list, tuple)) and len(pays_raw) > 1:
            pays = _PAYS_FR.get(_clean(pays_raw[1]), _clean(pays_raw[1])) or '—'
        else:
            pays = '—'
        phone = _clean(r.get('phone')) or _clean(r.get('mobile')) or '—'
        email = _clean(r.get('email')) or '—'
        ref   = _clean(r.get('ref'))   or '—'
        vat   = _clean(r.get('vat'))   or '—'
        nb_cmd        = r.get('purchase_order_count') or 0
        supplier_rank = r.get('supplier_rank') or 0
        statut = 'Actif' if supplier_rank > 0 else 'Inactif'
        all_rows.append({
            'nom': nom, 'ville': ville, 'pays': pays,
            'phone': phone, 'email': email, 'ref': ref, 'vat': vat,
            'nb_commandes': nb_cmd, 'statut': statut,
        })

    # Totaux globaux et listes de choix calculés sur l'ensemble
    total_commandes = sum(r['nb_commandes'] for r in all_rows)
    villes_all = sorted({r['ville'].upper() for r in all_rows if r['ville'] != '—'})
    pays_all = sorted({r['pays'] for r in all_rows if r['pays'] != '—'})

    top_fournisseur = '—'
    if all_rows:
        top = max(all_rows, key=lambda r: r['nb_commandes'])
        top_fournisseur = top['nom']

    # ── Application des filtres ──────────────────────────────────
    rows = []
    for r in all_rows:
        if nom_filter and nom_filter.lower() not in r['nom'].lower():
            continue
        if ville_filter and ville_filter.upper() != r['ville'].upper():
            continue
        if pays_filter and pays_filter != r['pays']:
            continue
        if statut_filter and statut_filter != r['statut']:
            continue
        rows.append(r)

    total_fournisseurs = len(rows)
    fournisseurs_actifs = sum(1 for r in rows if r['statut'] == 'Actif')
    nb_avec_email = sum(1 for r in rows if r['email'] != '—')
    nb_avec_phone = sum(1 for r in rows if r['phone'] != '—')

    stats_ville: dict = defaultdict(int)
    for r in all_rows:
        if r['ville'] != '—':
            stats_ville[r['ville'].upper()] += 1
    top_villes = sorted(stats_ville.items(), key=lambda x: x[1], reverse=True)[:6]
    stats_ville_json = json.dumps({k: v for k, v in top_villes})

    # ── CSV ─────────────────────────────────────────────────────
    if export == 'csv':
        resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        resp['Content-Disposition'] = 'attachment; filename="fournisseurs.csv"'
        w = csv.writer(resp, delimiter=';')
        w.writerow(['Nom', 'Ville', 'Pays', 'Téléphone', 'Email', 'Réf.', 'TVA', 'Nb Commandes', 'Statut'])
        for r in rows:
            w.writerow([r['nom'], r['ville'], r['pays'], r['phone'], r['email'],
                        r['ref'], r['vat'], r['nb_commandes'], r['statut']])
        return resp

    # ── Excel ────────────────────────────────────────────────────
    if export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Fournisseurs'
        hdr_fill = PatternFill('solid', fgColor='1A2C4E')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        headers = ['Nom', 'Ville', 'Pays', 'Téléphone', 'Email', 'Réf.', 'TVA', 'Nb Commandes', 'Statut']
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
        for r in rows:
            ws.append([r['nom'], r['ville'], r['pays'], r['phone'], r['email'],
                       r['ref'], r['vat'], r['nb_commandes'], r['statut']])
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="fournisseurs.xlsx"'
        return resp

    # ── PDF ──────────────────────────────────────────────────────
    if export == 'pdf':
        bio_pdf = io.BytesIO()
        logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_somatrin.png'
        pw, ph = landscape(A4)

        class _PagedCanvasFN(rl_canvas.Canvas):
            def __init__(self, *args, **kw):
                super().__init__(*args, **kw)
                self._saved_states = []
            def showPage(self):
                self._saved_states.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved_states)
                for state in self._saved_states:
                    self.__dict__.update(state)
                    self.setFont('Helvetica', 7.5)
                    self.setFillColor(colors.HexColor('#6B7280'))
                    self.drawRightString(pw - 15 * mm, 10 * mm,
                                        f'Page {self._pageNumber} / {total}')
                    super().showPage()
                super().save()

        def _header_fn(c, d):
            c.saveState()
            if logo_path.is_file():
                c.drawImage(str(logo_path), 15 * mm, ph - 22 * mm,
                            width=32 * mm, height=11 * mm,
                            preserveAspectRatio=True, mask='auto')
            c.setFont('Helvetica-Bold', 11)
            c.setFillColor(colors.HexColor('#1A2C4E'))
            c.drawString(52 * mm, ph - 13 * mm, 'Référentiel Fournisseurs — SOMATRIN')
            c.setFont('Helvetica', 8)
            c.setFillColor(colors.HexColor('#6B7280'))
            c.drawRightString(pw - 15 * mm, ph - 13 * mm, 'Document Confidentiel')
            c.setStrokeColor(colors.HexColor('#E87722'))
            c.setLineWidth(1.2)
            c.line(15 * mm, ph - 26 * mm, pw - 15 * mm, ph - 26 * mm)
            c.restoreState()

        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle('cell', parent=styles['Normal'],
                                    fontSize=7.5, leading=10)

        col_widths = [60*mm, 35*mm, 35*mm, 38*mm, 60*mm, 25*mm, 25*mm]
        table_headers = ['Nom', 'Ville', 'Pays', 'Téléphone', 'Email', 'Nb Cmd', 'Statut']

        data = [table_headers]
        for r in rows:
            data.append([
                Paragraph(r['nom'], cell_style),
                r['ville'], r['pays'], r['phone'],
                Paragraph(r['email'], cell_style),
                str(r['nb_commandes']), r['statut'],
            ])

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A2C4E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))

        doc = SimpleDocTemplate(
            bio_pdf, pagesize=landscape(A4),
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=32*mm, bottomMargin=18*mm,
        )
        doc.build([tbl], onFirstPage=_header_fn, onLaterPages=_header_fn,
                  canvasmaker=_PagedCanvasFN)
        bio_pdf.seek(0)
        resp = HttpResponse(bio_pdf.read(), content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="fournisseurs.pdf"'
        return resp

    return render(request, 'achats/fournisseurs.html', {
        'rows': rows,
        'total_fournisseurs': total_fournisseurs,
        'fournisseurs_actifs': fournisseurs_actifs,
        'total_commandes': total_commandes,
        'top_fournisseur': top_fournisseur,
        'nb_avec_email': nb_avec_email,
        'nb_avec_phone': nb_avec_phone,
        'stats_ville_json': stats_ville_json,
        'nom': nom_filter,
        'ville': ville_filter,
        'pays': pays_filter,
        'statut': statut_filter,
        'villes': villes_all,
        'pays_list': pays_all,
    })


def _render_parc_module(request, current_key, page_title, page_subtitle, rows=None, extra_context=None):
    menu_items = [
        {'key': 'overview', 'label': "Vue d'ensemble", 'url': 'parc_overview', 'icon': 'bi-grid-1x2-fill'},
        {'key': 'equipements', 'label': "Équipements", 'url': 'parc_equipements', 'icon': 'bi-truck-front-fill'},
        {'key': 'disponibilite', 'label': 'Disponibilité', 'url': 'parc_disponibilite', 'icon': 'bi-speedometer2'},
        {'key': 'ordres_maintenance', 'label': 'Ordres maintenance', 'url': 'parc_ordres_maintenance', 'icon': 'bi-tools'},
        {'key': 'interventions', 'label': 'Interventions', 'url': 'parc_interventions', 'icon': 'bi-wrench-adjustable'},
        {'key': 'couts', 'label': 'Coûts maintenance', 'url': 'parc_couts', 'icon': 'bi-cash-coin'},
    ]
    ctx = {
        'page_title': page_title,
        'page_subtitle': page_subtitle,
        'module_key': current_key,
        'menu_items': menu_items,
        'rows': rows or [],
    }
    if extra_context:
        ctx.update(extra_context)
    return render(request, 'parc/module.html', ctx)


@login_required
def parc_overview(request):
    return _render_parc_module(
        request,
        current_key='overview',
        page_title="Vue d'ensemble",
        page_subtitle="Pilotage du parc matériel: disponibilité, maintenance et coûts.",
    )


@login_required
def parc_equipements(request):
    error = None
    rows = []
    total = actifs = 0
    nb_categories = nb_societes = 0
    try:
        uid, models = get_odoo_connection()
        fields_meta = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'maintenance.equipment', 'fields_get', [], {'attributes': ['type']},
        ) or {}
        available = set(fields_meta.keys())
        read_fields = ['name']
        for f in ('category_id', 'company_id', 'active', 'serial_no', 'technician_user_id'):
            if f in available:
                read_fields.append(f)
        records = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'maintenance.equipment', 'search_read',
            [[]],
            {'fields': sorted(set(read_fields)), 'limit': 2000, 'order': 'name asc'},
        )
        categories = set()
        societes = set()
        for r in records:
            cat = r.get('category_id')
            category_name = cat[1] if isinstance(cat, list) and len(cat) > 1 else '—'
            comp = r.get('company_id')
            company_name = comp[1] if isinstance(comp, list) and len(comp) > 1 else '—'
            tech = r.get('technician_user_id')
            tech_name = tech[1] if isinstance(tech, list) and len(tech) > 1 else '—'
            is_active = bool(r.get('active', True))
            rows.append({
                'values_display': [
                    r.get('name') or '—',
                    category_name,
                    company_name,
                    r.get('serial_no') or '—',
                    tech_name,
                    'Actif' if is_active else 'Inactif',
                ],
                'status_index': 5,
            })
            if category_name != '—':
                categories.add(category_name)
            if company_name != '—':
                societes.add(company_name)
            if is_active:
                actifs += 1
        total = len(rows)
        nb_categories = len(categories)
        nb_societes = len(societes)
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return _render_parc_module(
        request,
        current_key='equipements',
        page_title='Équipements',
        page_subtitle='Inventaire des équipements du parc et statut opérationnel.',
        rows=rows,
        extra_context={
            'error': error,
            'kpi_total': total,
            'kpi_actifs': actifs,
            'kpi_categories': nb_categories,
            'kpi_societes': nb_societes,
            'table_columns': ['Équipement', 'Catégorie', 'Société', 'Série', 'Technicien', 'Statut'],
        },
    )


@login_required
def parc_disponibilite(request):
    error = None
    rows = []
    kpi_disponibles = kpi_indisponibles = 0
    kpi_taux = 0.0
    try:
        uid, models = get_odoo_connection()
        eq_fields = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'maintenance.equipment', 'fields_get', [], {'attributes': ['type']},
        ) or {}
        eq_available = set(eq_fields.keys())
        eq_read = ['name']
        for f in ('active', 'category_id', 'company_id'):
            if f in eq_available:
                eq_read.append(f)
        equipments = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'maintenance.equipment', 'search_read',
            [[]],
            {'fields': sorted(set(eq_read)), 'limit': 2000, 'order': 'name asc'},
        )

        # Equipements avec ordre de maintenance ouvert => indisponibles.
        busy_ids = set()
        try:
            req_fields = models.execute_kw(
                settings.ODOO_DB, uid, settings.ODOO_PASS,
                'maintenance.request', 'fields_get', [], {'attributes': ['type']},
            ) or {}
            req_available = set(req_fields.keys())
            if 'equipment_id' in req_available and 'stage_id' in req_available:
                req_records = models.execute_kw(
                    settings.ODOO_DB, uid, settings.ODOO_PASS,
                    'maintenance.request', 'search_read',
                    [[('stage_id.done', '=', False)]],
                    {'fields': ['equipment_id'], 'limit': 5000},
                )
                for rr in req_records:
                    eq = rr.get('equipment_id')
                    if isinstance(eq, list) and eq:
                        busy_ids.add(eq[0])
        except Exception:
            busy_ids = set()

        for eq in equipments:
            eq_id = eq.get('id')
            is_active = bool(eq.get('active', True))
            is_busy = eq_id in busy_ids
            status = 'Indisponible' if (not is_active or is_busy) else 'Disponible'
            if status == 'Disponible':
                kpi_disponibles += 1
            else:
                kpi_indisponibles += 1
            cat = eq.get('category_id')
            comp = eq.get('company_id')
            rows.append({
                'values_display': [
                    eq.get('name') or '—',
                    cat[1] if isinstance(cat, list) and len(cat) > 1 else '—',
                    comp[1] if isinstance(comp, list) and len(comp) > 1 else '—',
                    status,
                ],
                'status_index': 3,
            })
        total = kpi_disponibles + kpi_indisponibles
        if total:
            kpi_taux = round((kpi_disponibles * 100.0) / total, 1)
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return _render_parc_module(
        request,
        current_key='disponibilite',
        page_title='Disponibilité',
        page_subtitle='Disponibilité opérationnelle du parc en temps réel.',
        rows=rows,
        extra_context={
            'error': error,
            'kpi_disponibles': kpi_disponibles,
            'kpi_indisponibles': kpi_indisponibles,
            'kpi_taux': kpi_taux,
            'table_columns': ['Équipement', 'Catégorie', 'Société', 'Disponibilité'],
        },
    )


@login_required
def parc_ordres_maintenance(request):
    error = None
    rows = []
    kpi_total = kpi_ouverts = kpi_clos = 0
    try:
        uid, models = get_odoo_connection()
        fields_meta = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'maintenance.request', 'fields_get', [], {'attributes': ['type']},
        ) or {}
        available = set(fields_meta.keys())
        date_field = 'request_date' if 'request_date' in available else ('create_date' if 'create_date' in available else None)
        fields = ['name']
        for f in ('equipment_id', 'maintenance_type', 'owner_user_id', 'stage_id', 'description'):
            if f in available:
                fields.append(f)
        if date_field:
            fields.append(date_field)
        records = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'maintenance.request', 'search_read',
            [[]],
            {'fields': sorted(set(fields)), 'limit': 2000, 'order': f'{date_field} desc' if date_field else 'id desc'},
        )
        for r in records:
            dt = (r.get(date_field) or '')[:10] if date_field else ''
            eq = r.get('equipment_id')
            own = r.get('owner_user_id')
            stage = r.get('stage_id')
            stage_name = stage[1] if isinstance(stage, list) and len(stage) > 1 else '—'
            is_closed = 'done' in stage_name.lower() or 'close' in stage_name.lower() or 'clôt' in stage_name.lower()
            if is_closed:
                kpi_clos += 1
            else:
                kpi_ouverts += 1
            rows.append({
                'values_display': [
                    r.get('name') or '—',
                    f"{dt[8:10]}/{dt[5:7]}/{dt[0:4]}" if len(dt) == 10 else '—',
                    eq[1] if isinstance(eq, list) and len(eq) > 1 else '—',
                    r.get('maintenance_type') or '—',
                    own[1] if isinstance(own, list) and len(own) > 1 else '—',
                    stage_name,
                    (r.get('description') or '—')[:100],
                ],
                'status_index': 5,
            })
        kpi_total = len(rows)
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    return _render_parc_module(
        request,
        current_key='ordres_maintenance',
        page_title='Ordres de maintenance',
        page_subtitle='Suivi des ordres de maintenance préventive et corrective.',
        rows=rows,
        extra_context={
            'error': error,
            'kpi_total': kpi_total,
            'kpi_ouverts': kpi_ouverts,
            'kpi_clos': kpi_clos,
            'table_columns': ['N° Ordre', 'Date', 'Équipement', 'Type', 'Responsable', 'Statut', 'Description'],
        },
    )


@login_required
def parc_interventions(request):
    return _render_parc_module(
        request,
        current_key='interventions',
        page_title='Interventions',
        page_subtitle='Planification des interventions techniques et suivi des priorités.',
    )


@login_required
def parc_couts(request):
    return _render_parc_module(
        request,
        current_key='couts',
        page_title='Coûts maintenance',
        page_subtitle='Pilotage des coûts d’entretien et d’immobilisation du parc.',
    )


@login_required
def transport_rentabilite(request):
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    revenus = []
    couts = []
    error = None
    try:
        uid, models = get_odoo_connection()

        invoice_domain = [('move_type', '=', 'out_invoice')]
        invoice_domain += _transport_domain_dates(date_debut, date_fin, 'invoice_date')
        invoices = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'account.move', 'search_read',
            [invoice_domain],
            {'fields': ['invoice_date', 'amount_untaxed'], 'limit': 5000}
        )
        revenus = [
            {'date': (inv.get('invoice_date') or '')[:10], 'montant': inv.get('amount_untaxed') or 0}
            for inv in invoices
            if inv.get('invoice_date')
        ]

        cost_domain = []
        cost_domain += _transport_domain_dates(date_debut, date_fin, 'date')
        cost_items = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'fleet.vehicle.cost', 'search_read',
            [cost_domain],
            {'fields': ['date', 'amount', 'cost_subtype_id'], 'limit': 5000}
        )
        couts = [
            {
                'date': (item.get('date') or '')[:10],
                'montant': item.get('amount') or 0,
                'nature': item['cost_subtype_id'][1] if item.get('cost_subtype_id') else '—',
            }
            for item in cost_items
            if item.get('date')
        ]
    except Exception as exc:
        error = f'Erreur de connexion Odoo : {exc}'

    total_revenus = round(sum(r['montant'] for r in revenus), 2)
    total_couts = round(sum(c['montant'] for c in couts), 2)
    resultat = round(total_revenus - total_couts, 2)
    marge_pct = round((resultat / total_revenus * 100), 2) if total_revenus else 0

    couts_par_nature = defaultdict(float)
    for cost in couts:
        couts_par_nature[c['nature']] += cost['montant']
    couts_agg = [
        {'nature': k, 'montant': round(v, 2)}
        for k, v in sorted(couts_par_nature.items(), key=lambda kv: -kv[1])
    ]

    return render(request, 'transport/rentabilite.html', {
        'error': error,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'total_revenus': total_revenus,
        'total_couts': total_couts,
        'resultat': resultat,
        'marge_pct': marge_pct,
        'couts_agg': couts_agg,
    })


# ─────────────────────────────────────────────
#  QHSE — ALERTES QUALITÉ
#  Modèle : quality.alert
# ─────────────────────────────────────────────

def _fetch_qhse_sites(uid, models):
    """Récupère la liste des sites depuis stock.location."""
    sites = []
    try:
        locs = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'stock.location', 'search_read',
            [[['usage', 'in', ['internal', 'customer', 'supplier']]]],
            {'fields': ['complete_name'], 'order': 'complete_name asc', 'limit': 300}
        )
        sites = [loc['complete_name'] for loc in locs if loc.get('complete_name')]
    except Exception:
        pass
    return sites


def _build_qhse_alert_domain(site=None):
    """Construit le domaine pour la recherche d'alertes QHSE."""
    domain = [('state', '!=', 'cancel')]
    if site:
        domain.append(('location_id.complete_name', 'ilike', site))
    return domain


def _fetch_qhse_alerts(uid, models, site=None, limit=200):
    """Récupère les alertes QHSE depuis Odoo."""
    alerts = []
    try:
        alerts = models.execute_kw(
            settings.ODOO_DB, uid, settings.ODOO_PASS,
            'quality.alert', 'search_read',
            [_build_qhse_alert_domain(site)],
            {
                'fields': ['name', 'date', 'state', 'company_id', 'location_id', 'team_id', 'category_id', 'user_id'],
                'order': 'date desc',
                'limit': limit
            }
        )
    except Exception:
        alerts = []
    return alerts


def _summarize_qhse_alerts(alerts):
    """Résume les alertes QHSE par état."""
    summary = {'total': 0, 'open': 0, 'in_progress': 0, 'done': 0}
    for alert in alerts:
        state = (alert.get('state') or '').lower()
        summary['total'] += 1
        if state in ('new', 'draft', 'proposed'):
            summary['open'] += 1
        elif state in ('in_progress', 'progress', 'doing'):
            summary['in_progress'] += 1
        elif state in ('done', 'closed', 'solved'):
            summary['done'] += 1
    return summary


def _load_qhse_context(request):
    """Charge le contexte pour les pages QHSE."""
    site = request.GET.get('site', '')
    error = None
    alerts = []
    sites = []
    summary = {'total': 0, 'open': 0, 'in_progress': 0, 'done': 0}
    try:
        uid, models = get_odoo_connection()
        sites = _fetch_qhse_sites(uid, models)
        alerts = _fetch_qhse_alerts(uid, models, site=site)
        summary = _summarize_qhse_alerts(alerts)
    except Exception as exc:
        error = str(exc)
    return {
        'error': error,
        'site': site,
        'sites': sites,
        'alerts': alerts,
        'summary': summary,
    }


@login_required
def qhse_bilan(request):
    """Bilan des alertes QHSE."""
    context = _load_qhse_context(request)
    context.update({
        'page_title': 'Bilan QHSE',
        'page_subtitle': 'Indicateurs QHSE récupérés depuis Odoo.',
    })
    return render(request, 'qhse/bilan.html', context)


@login_required
def qhse_entrees(request):
    """Entrées QHSE."""
    context = _load_qhse_context(request)
    context.update({
        'page_title': 'Entrées QHSE',
        'page_subtitle': 'Données de conformité et incidents entrés depuis Odoo.',
    })
    return render(request, 'qhse/entrees.html', context)


@login_required
def qhse_sorties(request):
    """Sorties QHSE."""
    context = _load_qhse_context(request)
    context.update({
        'page_title': 'Sorties QHSE',
        'page_subtitle': 'Rapports de sortie et actions QHSE provenant d\'Odoo.',
    })
    return render(request, 'qhse/sorties.html', context)


# ─────────────────────────────────────────────
#  GASOIL — EXPORTS (CSV/PDF alias for Excel)
# ─────────────────────────────────────────────

@login_required
def gasoil_sorties_csv(request):
    """Export CSV alias pour gasoil_sorties_export (Excel)."""
    return gasoil_sorties_export(request)


@login_required
def gasoil_rapport(request):
    """Page HTML du rapport sorties gasoil avec données d'exemple."""
    # Données d'exemple pour prévisualiser le rapport HTML.
    bons_data = [
        {
            'date': '2026-04-16',
            'name': 'LHMEK/MOI/09154',
            'societe': 'SOMATRIN',
            'site': 'LHMEK/Stock',
            'ouvrage': 'S00011-Manutention des MP vers concasseurs - Chargement transport Matières premières - LAFARGEHOLCIM MAROC',
            'engin': '59087-B-33/YV2XG30G3SB50467 6',
            'categorie': 'CAMION ENGIN',
            'chauffeur': 'MOHAMMED HADDAD',
            'cpt_initial': 1491,
            'cpt_actuel': 1498,
            'ecart': 7.0,
            'product_qty': 70.0,
            'consommation': 10.0,
            'anomalie': 'OK',
        },
        {
            'date': '2026-04-16',
            'name': 'LHMEK/MOI/09153',
            'societe': 'SOMATRIN',
            'site': 'LHMEK/Stock',
            'ouvrage': 'S00011-Manutention des MP vers concasseurs - Chargement transport Matières premières - LAFARGEHOLCIM MAROC',
            'engin': '59087-B-33/YV2XG30G3SB50468',
            'categorie': 'CAMION ENGIN',
            'chauffeur': 'ABDELOUAHAB B OUYGHRAQUINE',
            'cpt_initial': 1497,
            'cpt_actuel': 1504,
            'ecart': 7.0,
            'product_qty': 59.0,
            'consommation': 8.43,
            'anomalie': 'OK',
        },
        {
            'date': '2026-04-16',
            'name': 'LHMEK/MOI/09152',
            'societe': 'SOMATRIN',
            'site': 'LHMEK/Stock',
            'ouvrage': 'S00011-Manutention des MP vers concasseurs - Chargement transport Matières premières - LAFARGEHOLCIM MAROC',
            'engin': '59087-B-33/YV2XG30G3SB50469',
            'categorie': 'CAMION ENGIN',
            'chauffeur': 'MUSTAFA BOUAIOUN',
            'cpt_initial': 1538,
            'cpt_actuel': 1545,
            'ecart': 7.0,
            'product_qty': 70.0,
            'consommation': 10.0,
            'anomalie': 'OK',
        },
        {
            'date': '2026-04-16',
            'name': 'LHMEK/MOI/09151',
            'societe': 'SOMATRIN',
            'site': 'LHMEK/Stock',
            'ouvrage': 'S00011-Manutention des MP vers concasseurs - Chargement transport Matières premières - LAFARGEHOLCIM MAROC',
            'engin': '59087-B-33/YV2XG30G3SB50470',
            'categorie': 'CAMION ENGIN',
            'chauffeur': 'MOHAMMED EL MAKOUDI',
            'cpt_initial': 1500,
            'cpt_actuel': 1506,
            'ecart': 6.0,
            'product_qty': 60.0,
            'consommation': 10.0,
            'anomalie': 'OK',
        },
        {
            'date': '2026-04-16',
            'name': 'LHMEK/MOI/09150',
            'societe': 'SOMATRIN',
            'site': 'LHMEK/Stock',
            'ouvrage': 'S00011-Manutention des MP vers concasseurs - Chargement transport Matières premières - LAFARGEHOLCIM MAROC',
            'engin': '59087-B-33/YV2XG30G3SB50471',
            'categorie': 'CAMION ENGIN',
            'chauffeur': 'AZIZ EL FTATCHI',
            'cpt_initial': 1529,
            'cpt_actuel': 1535,
            'ecart': 6.0,
            'product_qty': 50.0,
            'consommation': 8.33,
            'anomalie': 'OK',
        },
        {
            'date': '2026-04-16',
            'name': 'LHMEK/MOI/09149',
            'societe': 'SOMATRIN',
            'site': 'LHMEK/Stock',
            'ouvrage': 'S00011-Manutention des MP vers concasseurs - Chargement transport Matières premières - LAFARGEHOLCIM MAROC',
            'engin': '59087-B-33/YV2XG30G3SB50472',
            'categorie': 'CAMION ENGIN',
            'chauffeur': 'MUSTAPHA MAHJOUB',
            'cpt_initial': 1561,
            'cpt_actuel': 1567,
            'ecart': 6.0,
            'product_qty': 52.0,
            'consommation': 8.67,
            'anomalie': 'OK',
        },
    ]

    total_qty = sum(b['product_qty'] for b in bons_data)

    return render(request, 'gasoil/rapport.html', {
        'bons': bons_data,
        'total_bons': len(bons_data),
        'total_qty': total_qty,
    })
    """Export PDF — redirige vers gasoil_sorties avec export=pdf."""
    # Passe par le GET en le modifiant localement
    request.GET = request.GET.copy()
    request.GET._mutable = True
    request.GET['export'] = 'pdf'
    request.GET._mutable = False
    return gasoil_sorties(request)
